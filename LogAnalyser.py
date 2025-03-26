'''
import os
import re
from datetime import datetime, timedelta
from collections import defaultdict
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QPushButton, QLabel,
    QLineEdit, QListWidget, QVBoxLayout, QHBoxLayout, QWidget, QProgressBar, QTabWidget,
    QRadioButton, QMessageBox, QSizePolicy, QDateEdit, QSpinBox, QGridLayout, QCheckBox,
    QComboBox, QDialog, QTreeView, QHeaderView, QScrollArea, QButtonGroup
)
from PyQt6.QtCore import Qt, QTimer, QSize, QPoint
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.gridspec as gridspec
from datetime import datetime
from PyQt6.QtGui import QPainter, QColor, QPixmap, QIcon, QFileSystemModel

class MultijetImgDialog(QDialog):
    def __init__(self, image_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("MultiJet Pipe")
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinMaxButtonsHint)  # 添加最小化和最大化按鈕

        # 載入圖片並獲取其大小
        pixmap = QPixmap(image_path)
        self.image_label = QLabel()
        self.image_label.setPixmap(pixmap)  # 設置圖片到 QLabel

        # 設置對話框大小為圖片大小
        self.resize(pixmap.size())  # 將對話框大小設置為圖片大小

        layout = QVBoxLayout()
        layout.addWidget(self.image_label)
        self.setLayout(layout)

class ValveImageLabel(QLabel):
    def __init__(self, pixmap, valve_positions, valve_close_img, parent=None):
        super().__init__(parent)
        self.pixmap = pixmap
        self.valve_positions = valve_positions  # 儲存所有閥門的位置
        self.valve_images = {valve_id: QPixmap(valve_close_img) for valve_id in valve_positions.keys()}  # 儲存閥門圖片

    def paintEvent(self, event):
        painter = QPainter(self)  # 在 QLabel 本身上繪製
        painter.drawPixmap(0, 0, self.pixmap)  # 繪製主圖片

        # 繪製所有閥門圖片
        for valve_id, position in self.valve_positions.items():
            # 確保 position 是 QPoint
            if isinstance(position, tuple):
                position = QPoint(*position)  # 使用 * 解包元組
            painter.drawPixmap(position, self.valve_images[valve_id])  # 繪製閥門圖片

    def sizeHint(self):
        return self.pixmap.size()  # 返回主圖片的大小

    def set_valve_image(self, valve_id, valve_image_path):
        self.valve_images[valve_id] = QPixmap(valve_image_path)
        self.update()  # 更新 QLabel 以顯示新的閥門圖片

class ValveStatusDialog(QDialog):
    def __init__(self, multijet_image_path, valve_open_image_path, valve_close_image_path, valve_positions, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Valve Status")
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinMaxButtonsHint)  # 添加最小化和最大化按鈕

        # 載入主圖片
        self.pixmap = QPixmap(multijet_image_path)
        self.valve_open_image_path = valve_open_image_path
        self.valve_close_image_path = valve_close_image_path
        self.valve_positions = valve_positions

        # 創建自定義 QLabel 來顯示圖片
        self.image_label = ValveImageLabel(self.pixmap, self.valve_positions, self.valve_close_image_path)

        # 創建滾動區域
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(False)  # 使滾動區域可調整大小
        self.scroll_area.setWidget(self.image_label)  # 將自定義 QLabel 設置為滾動區域的內容

        # 設置對話框大小為圖片大小
        self.resize(self.pixmap.size())

        # 創建顯示座標的 QLabel
        self.coordinate_label = QLabel("座標: (0, 0)")
        self.coordinate_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        layout = QVBoxLayout()
        layout.addWidget(self.scroll_area)
        layout.addWidget(self.coordinate_label)  # 添加座標顯示
        self.setLayout(layout)

        # 連接滑鼠移動事件
        self.image_label.setMouseTracking(True)  # 啟用滑鼠追蹤
        self.image_label.mouseMoveEvent = self.mouse_move_event  # 自定義滑鼠移動事件

        # 初始化閥門狀態
        self.valve_states = {valve_id: 0 for valve_id in self.valve_positions.keys()}  # 預設為關閉狀態

        # 繪製閥門
        self.update_valve_display()

    def closeEvent(self, event):
        # 在關閉對話框時釋放資源或進行清理
        print("Closing ValveStatusDialog")  # 可選，顯示關閉信息
        event.accept()  # 確保事件被接受

    def update_valve_display(self):
        for valve_id, position in self.valve_positions.items():
            status = self.valve_states[valve_id]
            valve_image = self.valve_close_image_path if status == 0 else self.valve_open_image_path
            self.image_label.set_valve_image(valve_id, valve_image)  # 更新閥門圖片

    def mouse_move_event(self, event):
        # 獲取滑鼠座標
        x = event.pos().x()
        y = event.pos().y()
        self.coordinate_label.setText(f"座標: ({x}, {y})")  # 更新座標顯示

class LogAnalyser(QMainWindow):
    def __init__(self):
        super().__init__()
        self.processed_files = 0  # 已解析的檔案數
        self.log_data = None
        self.min_time = None
        self.max_time = None
        self.valve_log_data = {}
        self.valve_status_dialog = None
        # 初始化計時器
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_processing_label)  # 連接計時器的超時信號到更新函數
        self.dot_count = 0
        self.mask_events = {}
        self.workflow_data = {}
        self.ctr_char_gen = False
        self.all_times = []
        self.protocol_data = {}
        #self.led = Led()  # 創建 LED 實例
        #self.led.set_on(False)  # 初始狀態為關閉

        self.initUI()

    def initUI(self):
        self.setWindowTitle("Log Analyser v1.0.0")
        self.setGeometry(100, 100, 1200, 800)

        # 主 Widget 和 Layout
        central_widget = QWidget()
        central_layout = QVBoxLayout(central_widget)
        self.setCentralWidget(central_widget)

        # 添加功能表
        self.create_menu_bar()

        # 顯示資料夾路徑
        self.folder_path_edit = QLineEdit()
        self.folder_path_edit.setFixedWidth(500)
        central_layout.addWidget(self.folder_path_edit)

        # 建立 Tab Widget
        self.tab_widget = QTabWidget()
        central_layout.addWidget(self.tab_widget)

        #### Tab 1 - 主要資料處理
        wf_widget = QWidget()
        wf_layout = QVBoxLayout(wf_widget)

        # Tab 1 完成，加入到 TabWidget
        self.tab_widget.addTab(wf_widget, "Work Flow")
        
        load_unload_label = QLabel("Load/Unload information")
        work_step_label = QLabel("Work Steps")

        self.lwt_mask_info = QListWidget()
        self.lwt_mask_info.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.lwt_mask_info.itemClicked.connect(self.display_workflow)

        self.lwt_wk_flow = QListWidget()
        self.lwt_wk_flow.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        left_v_layout = QVBoxLayout()
        left_v_layout.addWidget(load_unload_label)
        left_v_layout.addWidget(self.lwt_mask_info)

        right_v_layout = QVBoxLayout()
        right_v_layout.addWidget(work_step_label)
        right_v_layout.addWidget(self.lwt_wk_flow)

        list_h_layout = QHBoxLayout()
        list_h_layout.addLayout(left_v_layout)
        list_h_layout.addLayout(right_v_layout)
        list_h_layout.setStretchFactor(left_v_layout, 1)
        list_h_layout.setStretchFactor(right_v_layout, 1)

        # 創建 QFileSystemModel
        self.file_system_model = QFileSystemModel()

        # 獲取所有可用的磁碟驅動器
        drives = [f"{d}:\\" for d in range(65, 91) if os.path.exists(f"{chr(d)}:\\")]  # ASCII 65-90 對應 A-Z

        # 創建 QTreeView
        self.tree_view = QTreeView()
        self.tree_view.setFixedHeight(165)
        self.tree_view.setModel(self.file_system_model)

        # 添加「本機」作為根節點
        local_machine_index = self.file_system_model.setRootPath('')  # 設置根路徑為空以顯示所有磁碟驅動器
        self.tree_view.setRootIndex(local_machine_index)  # 設置樹的根節點

        # 設置樹狀結構的顯示屬性
        self.tree_view.setHeaderHidden(False)  # 顯示標題
        self.tree_view.setAlternatingRowColors(True)  # 交替行顏色

        # 設置欄位大小可調整
        header = self.tree_view.header()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # 設置欄位為可互動調整大小

        # 將樹狀結構添加到 wf_layout
        wf_layout.addWidget(self.tree_view)  # 將樹狀結構添加到工作流程佈局中

        # 將可用的磁碟驅動器添加到樹狀結構
        for drive in drives:
            drive_index = self.file_system_model.index(drive)  # 獲取驅動器的索引
            self.tree_view.setRootIndex(drive_index)  # 設置樹的根節點為該驅動器

        select_folder_button = QPushButton("Select Folder")
        select_folder_button.clicked.connect(self.select_folder_from_tree)

        wf_layout.addLayout(list_h_layout)
        # 將樹狀結構添加到 wf_layout
        wf_layout.addWidget(self.tree_view)  # 將樹狀結構添加到工作流程佈局中
        wf_layout.addWidget(select_folder_button, alignment=Qt.AlignmentFlag.AlignLeft)

        # 處理按鈕
        self.process_button = QPushButton("Process")
        self.process_button.setFixedSize(80, 40)
        self.process_button.setStyleSheet("""
            QPushButton {
                font-weight: bold;
                background-color: white;
                color: black; 
                border: 2px solid #000;
            }
            QPushButton:hover {
                font-weight: bold;
                background-color: lightblue;
                color: black;
            }
        """)
        self.process_button.clicked.connect(self.process_raw_log)
        central_layout.addWidget(self.process_button)

        # 新增進度條
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)  # 預設範圍為 0 到 100
        self.progress_bar.setValue(0)  # 初始值為 0
        self.progress_bar.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        self.progress_bar.setStyleSheet("QProgressBar::chunk { background-color: lightgreen; }")
        central_layout.addWidget(self.progress_bar)

        process_layout = QHBoxLayout()
        process_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        process_layout.addWidget(self.process_button, alignment=Qt.AlignmentFlag.AlignLeft)
        process_layout.addWidget(self.progress_bar, alignment=Qt.AlignmentFlag.AlignLeft)
        central_layout.addLayout(process_layout)

        # 完成訊息
        self.status_label = QLabel("")
        central_layout.addWidget(self.status_label, alignment=Qt.AlignmentFlag.AlignLeft)

        # Tab 2 - Mask ID Analysis
        mask_analysis_widget = QWidget()
        mask_analysis_v_layout = QVBoxLayout(mask_analysis_widget)
        self.tab_widget.addTab(mask_analysis_widget, "Mask ID Analysis")

        # Add buttons to generate charts
        load_unload_chart_button = QPushButton("Load/Unload Time Chart")
        load_unload_chart_button.clicked.connect(self.generate_load_unload_time_chart)

        duration_chart_button = QPushButton("Duration Time Chart")
        duration_chart_button.clicked.connect(self.generate_duration_time_chart)

        mask_analysis_btn_h_layout = QHBoxLayout()
        mask_analysis_btn_h_layout.addWidget(load_unload_chart_button)
        mask_analysis_btn_h_layout.addWidget(duration_chart_button)
        mask_analysis_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        mask_analysis_v_layout.addLayout(mask_analysis_btn_h_layout)

        # Canvas for displaying charts
        self.figure_mask_analysis = Figure(figsize=(10, 6))
        self.canvas_mask_analysis  = FigureCanvas(self.figure_mask_analysis)
        mask_analysis_v_layout.addWidget(self.canvas_mask_analysis)

        ### Tab 3 - Install Info
        install_widge = QWidget()
        install_layout = QVBoxLayout(install_widge)

        # Tab 3 完成，加入到 TabWidget
        self.tab_widget.addTab(install_widge, "Install Log")

        # Install Info 列表
        self.install_info_list = QListWidget()
        self.install_info_list.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        install_layout.addWidget(self.install_info_list)

        ### Tab 4 - MSC Info
        msc_widge = QWidget()
        msc_v_layout = QVBoxLayout(msc_widge)

        # Tab 4 完成，加入到 TabWidget
        self.tab_widget.addTab(msc_widge, "MSC Info")

        self.tab_widget.currentChanged.connect(self.on_tab_changed)

        # Radio Buttons for msc version 
        label_msc_ver = QLabel("MSC ver : ")
        self.msc_v2_radio = QRadioButton("MSC 2.x")
        self.msc_v2_radio.toggled.connect(self.on_format_selected)
        self.selected_format = "MSC 3.x"  # 預設選擇
        self.msc_v3_radio = QRadioButton("MSC 3.x")
        self.msc_v3_radio.setChecked(True)
        self.msc_v3_radio.toggled.connect(self.on_format_selected)

        msc_version_group = QButtonGroup(self)
        msc_version_group.addButton(self.msc_v2_radio)
        msc_version_group.addButton(self.msc_v3_radio)

        msc_ver_h_layout = QHBoxLayout()

        msc_ver_h_layout.addWidget(label_msc_ver)
        msc_ver_h_layout.addWidget(self.msc_v2_radio)
        msc_ver_h_layout.addWidget(self.msc_v3_radio)

        # Gauge checkbox
        label_gauge_tp = QLabel("   CTR : ")
        self.gauge_checkboxes = {}
        gauge_types = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2"]
        msc_ver_h_layout.addWidget(label_gauge_tp)
        for gauge in gauge_types:
            checkbox = QCheckBox(gauge)
            checkbox.setChecked(True)  # 預設為選中
            self.gauge_checkboxes[gauge] = checkbox  # 將複選框存儲在字典中
            msc_ver_h_layout.addWidget(checkbox)  # 將複選框添加到布局中

            # 連接複選框的狀態變化事件
            checkbox.stateChanged.connect(self.on_gauge_checkbox_changed)

        msc_ver_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(msc_ver_h_layout)

        # Radio Buttons for chart source 
        label_chart_source = QLabel("Chart source : ")
        self.multijet_chart_radio = QRadioButton("mjnxtdebug")
        self.multijet_chart_radio.setChecked(True)
        self.multijet_chart_radio.toggled.connect(self.on_chart_source_selected)
        self.selected_chart = "mjnxtdebug"  # 預設選擇
        self.protocol_chart_radio = QRadioButton("protocol")
        self.protocol_chart_radio.toggled.connect(self.on_chart_source_selected)

        chart_source_group = QButtonGroup(self)
        chart_source_group.addButton(self.multijet_chart_radio)
        chart_source_group.addButton(self.protocol_chart_radio)
        
        chart_source_h_layout = QHBoxLayout()
        chart_source_h_layout.addWidget(label_chart_source)
        chart_source_h_layout.addWidget(self.multijet_chart_radio)
        chart_source_h_layout.addWidget(self.protocol_chart_radio)

        chart_source_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(chart_source_h_layout)

        # 時間篩選器
        time_filter_layout = QGridLayout()
        #time_filter_layout.setHorizontalSpacing(1)  # 設置水平間距
        #time_filter_layout.setVerticalSpacing(1)     # 設置垂直間距
        #time_filter_layout.setContentsMargins(1, 1, 1, 1)  # 設置邊距為 0

        # 標題列
        header_labels = ["", "Date", "Hr", "Min", "Sec"]
        for col, text in enumerate(header_labels):
            label = QLabel(text)
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)  # 置中對齊
            time_filter_layout.addWidget(label, 0, col)

        # Start Time
        time_filter_layout.addWidget(QLabel("Start:"), 1, 0)

        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("yyyy/MM/dd")
        time_filter_layout.addWidget(self.start_date_edit, 1, 1)

        self.start_hour_spinbox = QSpinBox()
        self.start_hour_spinbox.setRange(0, 23)
        self.start_hour_spinbox.setFixedWidth(60)
        self.start_hour_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_hour_spinbox, 1, 2)

        self.start_minute_spinbox = QSpinBox()
        self.start_minute_spinbox.setRange(0, 59)
        self.start_minute_spinbox.setFixedWidth(60)
        self.start_minute_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_minute_spinbox, 1, 3)

        self.start_second_spinbox = QSpinBox()
        self.start_second_spinbox.setRange(0, 59)
        self.start_second_spinbox.setFixedWidth(60)
        self.start_second_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_second_spinbox, 1, 4)

        # End Time
        time_filter_layout.addWidget(QLabel("End:"), 2, 0)

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy/MM/dd")
        time_filter_layout.addWidget(self.end_date_edit, 2, 1)

        self.end_hour_spinbox = QSpinBox()
        self.end_hour_spinbox.setRange(0, 23)
        self.end_hour_spinbox.setFixedWidth(60)
        self.end_hour_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_hour_spinbox, 2, 2)

        self.end_minute_spinbox = QSpinBox()
        self.end_minute_spinbox.setRange(0, 59)
        self.end_minute_spinbox.setFixedWidth(60)
        self.end_minute_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_minute_spinbox, 2, 3)

        self.end_second_spinbox = QSpinBox()
        self.end_second_spinbox.setRange(0, 59)
        self.end_second_spinbox.setFixedWidth(60)
        self.end_second_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_second_spinbox, 2, 4)

        self.mask_project_combobox = QComboBox()
        self.mask_project_combobox.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents) # 自動調整大小
        self.mask_project_combobox.setMinimumWidth(280)
        self.protocol_combobox = QComboBox()
        self.protocol_combobox.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents) # 自動調整大小
        self.protocol_combobox.setMinimumWidth(220)
        self.mask_project_label = QLabel("Project:")
        self.protocol_label = QLabel("Protocol:")
        self.process_label = QLabel("Process:")
        self.u_label = QLabel("U: ")
        self.v_label = QLabel("V: ")

        protocol_h_layout = QHBoxLayout()
        protocol_h_layout.setContentsMargins(0, 0, 0, 0)
        protocol_h_layout.addWidget(self.mask_project_label)
        protocol_h_layout.addWidget(self.mask_project_combobox)
        protocol_h_layout.addWidget(self.protocol_label)
        protocol_h_layout.addWidget(self.protocol_combobox)
        protocol_h_layout.addWidget(self.process_label)
        protocol_h_layout.addWidget(self.u_label)
        protocol_h_layout.addWidget(self.v_label)
        protocol_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(protocol_h_layout)

        # Multijet Chart
        multijet_chart_h_layout = QHBoxLayout()
        multijet_pipe_button = QPushButton("MultiJet Pipe")
        multijet_pipe_button.clicked.connect(self.multijet_show)
        valve_status_button = QPushButton("Valve State")
        valve_status_button.clicked.connect(self.valve_state_show)

        multijet_chart_h_layout.addWidget(multijet_pipe_button)
        multijet_chart_h_layout.addWidget(valve_status_button)

        # Test for displaying time data
        vline_time_h_layout = QHBoxLayout()
        self.vline_time_label = QLabel("")
        vline_time_h_layout.addWidget(self.vline_time_label)

        # 讓時間篩選器靠左對齊
        time_filter_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(time_filter_layout)
        multijet_chart_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(multijet_chart_h_layout)
        vline_time_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(vline_time_h_layout)

        # 添加畫布以顯示圖表
        self.figure_ctr = Figure(figsize=(10, 6))
        self.canvas_ctr = FigureCanvas(self.figure_ctr)
        msc_v_layout.addWidget(self.canvas_ctr)

        # 添加按鈕以生成趨勢圖
        #gen_ctr_button = QPushButton("Chart")
        gen_ctr_button = QPushButton("")
        gen_ctr_button.setToolTip("Generate Chart")  # 設置工具提示
        line_chart_icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "line_chart_icon.png")  # 圖標路徑
        gen_ctr_button.setIcon(QIcon(line_chart_icon_path))  # 設置按鈕圖標
        gen_ctr_button.setIconSize(QSize(30, 30))  # 設置圖標大小
        gen_ctr_button.clicked.connect(lambda: self.create_chart(generate_chart=True))
        #expo_button = QPushButton("Export")
        expo_button = QPushButton("")
        expo_button.setToolTip("Export to Excel")  # 設置工具提示
        excel_icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_icon.png")  # 圖標路徑
        expo_button.setIcon(QIcon(excel_icon_path))  # 設置按鈕圖標
        expo_button.setIconSize(QSize(30, 30))  # 設置圖標大小
        expo_button.clicked.connect(lambda: self.create_chart(generate_chart=False))

        msc_btn_h_layout = QHBoxLayout()
        msc_btn_h_layout.addWidget(gen_ctr_button)
        msc_btn_h_layout.addWidget(expo_button)
        msc_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)

        # 添加 QComboBox 來選擇移動倍率
        self.shift_combo = QComboBox()
        self.shift_combo.addItems(["10%", "30%", "50%"])
        self.shift_combo.setCurrentIndex(0)  # 預設選擇 10%

        # 添加zoom按鈕
        zoom_in_button = QPushButton("-")
        zoom_in_button.clicked.connect(lambda: self.zoom_chart(zoom_in=False))
        zoom_out_button = QPushButton("+")
        zoom_out_button.clicked.connect(lambda: self.zoom_chart(zoom_in=True))

        # 添加平移按鈕
        left_shift_button = QPushButton("<--")
        left_shift_button.clicked.connect(lambda: self.shift_chart(left_shift=True))
        right_shift_button = QPushButton("-->")
        right_shift_button.clicked.connect(lambda: self.shift_chart(left_shift=False))

        # 將 zoom/shift 按鈕添加到佈局中
        msc_zoom_btn_h_layout = QHBoxLayout()
        msc_zoom_btn_h_layout.addWidget(zoom_in_button)
        msc_zoom_btn_h_layout.addWidget(zoom_out_button)
        msc_zoom_btn_h_layout.addWidget(left_shift_button)
        msc_zoom_btn_h_layout.addWidget(right_shift_button)
        msc_zoom_btn_h_layout.addWidget(self.shift_combo)
        msc_zoom_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignRight)

        msc_v_layout.addLayout(msc_zoom_btn_h_layout)
        msc_v_layout.addLayout(msc_btn_h_layout)        
        
    def create_menu_bar(self):
        menu_bar = self.menuBar()

        # File menu
        file_menu = menu_bar.addMenu("File")
        open_action = file_menu.addAction("Open Folder")
        open_action.triggered.connect(self.select_folder)

        # 新增 Log menu
        log_menu = menu_bar.addMenu("Log")
        self.log_actions = {
            "LogService": log_menu.addAction("LogService"),
            "Install": log_menu.addAction("Install"),
            "mjnxtdebug": log_menu.addAction("mjnxtdebug"),
            "Protocol": log_menu.addAction("Protocol")
        }

        for action in self.log_actions.values():
            action.setCheckable(True)  # 設置為可勾選
            action.setChecked(True)  # 預設為勾選

        # Info menu
        info_menu = menu_bar.addMenu("Info")
        about_action = info_menu.addAction("About")
        about_action.triggered.connect(self.show_about_dialog)

    def update_processing_label(self):
        self.dot_count = (self.dot_count + 1) % 4  # 使點數在 0 到 3 之間循環
        dots = '.' * self.dot_count  # 根據計數生成點數
        self.display_status(f"Processing {dots}", "ongoing")
    
    def open_file_dialog(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select a File", "", "All Files (*.*)")
        if file_path:
            QMessageBox.information(self, "Selected File", f"You selected:\n{file_path}")

    def show_about_dialog(self):
        QMessageBox.about(self, "Log Analyser", "Log Analyser v1.0.0\nAuthor : Davian Kuo\nE-mail : davian.kuo@zeiss.com")

    def display_status(self, msg, type):
        self.status_label.setText(msg)
        if type == "done":
            self.status_label.setStyleSheet("background-color: lightgreen;")
        elif type == "fail":
            self.status_label.setStyleSheet("background-color: lightpink;")
        elif type == "ongoing":
            self.status_label.setStyleSheet("background-color: lightyellow;")

    def select_folder_from_tree(self):
        # 獲取當前選中的索引
        selected_index = self.tree_view.currentIndex()
        if selected_index.isValid():  # 確保選中的索引有效
            # 獲取選中的資料夾路徑
            folder_path = self.file_system_model.filePath(selected_index)
            # 將路徑填入 folder_path_edit
            self.folder_path_edit.setText(folder_path)
            self.display_status("Folder selected !", "done")
        else:
            self.display_status("Please select folder !", "fail")

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.folder_path_edit.setText(folder)
            self.display_status("Folder selected !", "done")

    def on_chart_source_selected(self):
        if self.multijet_chart_radio.isChecked():
            self.selected_chart = "mjnxtdebug"
        elif self.protocol_chart_radio.isChecked():
            self.selected_chart = "protocol"
    
    def find_log_files(self):
        found_path = self.folder_path_edit.text()
        if not found_path:
            self.display_status("Please choose folder before parsing !", "fail")
            return None, None, None, None, None, None
        files = os.listdir(found_path)
        
        if self.log_actions["LogService"].isChecked():
            logsrvc_log = [
                os.path.join(found_path, file) for file in files
                if file == "LogService.txt" or re.match(r"LogService\d{4}-\d{2}-\d{2}_\d{6}\.txt", file)
            ]
            if not logsrvc_log:
                self.display_status("❌ Can't find valid LogService !", "fail")
                return None, None, None, None, None, None
        else:
            logsrvc_log = []
 
        if self.log_actions["Install"].isChecked():
            install_log = [
                os.path.join(found_path, file) for file in files
                if file == "Install.txt"
            ]
            if not install_log:
                self.display_status("❌ Can't find valid Install !", "fail")
                return None, None, None, None, None, None
        else:
            install_log = []
 
        if self.log_actions["mjnxtdebug"].isChecked():
            msc_log = [
                os.path.join(found_path, file) for file in files
                if re.match(r"^mjnxtdebug\d{8}\.log$", file)
            ]
            if not msc_log:
                self.display_status("❌ Can't find valid mjnxtdebug !", "fail")
                return None, None, None, None, None, None
        else:
            msc_log = []

        if self.log_actions["Protocol"].isChecked():
            protocol_log = [
                os.path.join(found_path, file) for file in files
                if re.match(r"Protocol_\d{4}-\d{2}-\d{2}_\d{4}_\d{4}\.txt", file)
            ]
            process_log = [
                os.path.join(found_path, file) for file in files
                if re.match(r"ProcessLog_\d{4}-\d{2}-\d{2}_\d{4}_\d{4}\.csv", file)
            ]

            if not protocol_log:
                self.display_status("❌ Can't find valid protocol file !", "fail")
                return None, None, None, None, None, None
            if not process_log:
                self.display_status("❌ Can't find valid ProcessLog file !", "fail")
                return None, None, None, None, None, None

            # 檢查 Protocol 和 ProcessLog 檔案是否成對存在
            protocol_ids = {re.search(r'_(\d{4}-\d{2}-\d{2}_\d{4}_\d{4})\.txt$', file).group(1): file for file in protocol_log}
            process_ids = {re.search(r'_(\d{4}-\d{2}-\d{2}_\d{4}_\d{4})\.csv$', file).group(1): file for file in process_log}

            missing_pairs = [pid for pid in protocol_ids if pid not in process_ids]
            if missing_pairs:
                self.display_status(f"❌ Missing ProcessLog files for: {', '.join(missing_pairs)}", "fail")
                return None, None, None, None, None, None
        else:
            protocol_log = []
            process_log = []

        # 進度條初始化
        total_files = len(logsrvc_log) + len(install_log) + len(msc_log) + len(protocol_log)
        self.progress_bar.setRange(0, total_files)
        self.progress_bar.setValue(0)

        return found_path, logsrvc_log, install_log, msc_log, protocol_log, process_log
    
    def process_raw_log(self):
        self.mask_events, self.workflow_data = {}, {}
        self.log_data, time_range, self.valve_log_data = {}, [], {}
        self.protocol_data = {}

        if not any(action.isChecked() for action in self.log_actions.values()):
            self.display_status("Please select at least one log to parse !", "fail")
            return

        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return

        self.process_button.setEnabled(False)
        self.process_button.setStyleSheet("background-color: lightgray; color: darkgray;")
        self.processed_files = 0
        self.display_status("Processing .", "ongoing")

        self.timer.start(100)

        # 1. LogService.txt 
        if logsrvc_file:
            parsed_result = self.parse_logsvr(folder_path, logsrvc_file)
            if parsed_result is None:
                self.timer.stop()
                self.display_status("LogService*.txt parse failed !", "fail")
                return
            self.mask_events, self.workflow_data = parsed_result

            # Show Mask ID Load/Unload Info
            self.display_load_unload_info()
        else:
            self.mask_events, self.workflow_data = {}, {}
            self.lwt_mask_info.clear()
            self.lwt_wk_flow.clear()

        # 2. Install.txt
        if install_log:
            self.parse_install(folder_path)
        else:
            self.install_info_list.clear()

        # 3. mjnxtdebug.log
        if msc_file:
            self.log_data, time_range, self.valve_log_data = self.parse_msc(folder_path, msc_file)
            if not self.log_data or not self.valve_log_data:
                self.timer.stop()
                self.display_status("mjnxtdebug*.log parse failed !", "fail")
                return

            # 獲取時間範圍
            self.min_time = min(time_range)
            self.max_time = max(time_range)
            self.start_date_edit.setMinimumDate(self.min_time.date())
            self.start_date_edit.setMaximumDate(self.max_time.date())
            self.end_date_edit.setMinimumDate(self.min_time.date())
            self.end_date_edit.setMaximumDate(self.max_time.date())
        else:
            self.log_data, time_range, self.valve_log_data = {}, [], {}
            self.figure_ctr.clear()

        # 4. Protocol.txt
        if protocol_file:
            self.parse_protocol(folder_path, protocol_file, process_log)
        else:
            self.protocol_data = {}
 
        self.timer.stop()
        self.display_status("Process done !", "done")
        QMessageBox.information(self, "Log Analyser", "Process done !", QMessageBox.StandardButton.Ok)
        self.process_button.setEnabled(True)
        self.process_button.setStyleSheet("""
            QPushButton {
                font-weight: bold;
                background-color: white;
                color: black; 
                border: 2px solid #000;
            }
            QPushButton:hover {
                font-weight: bold;
                background-color: lightblue;
                color: black;
            }
        """)

    def parse_logsvr(self, folder_path, logsrvc_file):
        logsrvc_file = sorted(
            logsrvc_file,
            key=lambda f: os.path.getmtime(os.path.join(folder_path, f)),
            reverse=True
        )

        # 儲存數據
        self.mask_events = defaultdict(list)
        self.workflow_data = defaultdict(dict)
        workflow_id = 1

        mask_event_pattern = r"Mask with Id: (\w+\.\d+) (loaded|unloaded)"
        workflow_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into '([^']+)'"
        end_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'Setup.ProtocolDefinition'"
        loaded_start_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'MaskLoading.LoadingMask'"
        unloaded_start_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'MaskUnloading.UnloadingMask'"

        loaded_start_time = None
        loaded_end_time = None
        unloaded_start_time = None
        unloaded_end_time = None

        current_workflow = []
        mask_id = None

        for i, file in enumerate(logsrvc_file):
            file_path = os.path.join(folder_path, file)
            with open(file_path, 'r', encoding='iso-8859-1') as f:
                lines = f.readlines()

            for line in reversed(lines):  # 倒序處理
                # 提取 Mask ID 事件
                mask_match = re.search(mask_event_pattern, line)
                if mask_match:
                    mask_id = mask_match.group(1)
                    event_type = mask_match.group(2)
                    timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    #self.mask_events[mask_id].append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, "N/A"))

                    if event_type == "loaded":
                        loaded_end_time = timestamp
                    elif event_type == "unloaded":
                        unloaded_end_time = timestamp

                # 提取 workflow_pattern
                workflow_match = re.search(workflow_pattern, line)
                if workflow_match and mask_id:
                    timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    current_workflow.append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], workflow_match.group(1)))

                    if re.search(loaded_start_pattern, line):
                        loaded_start_time = timestamp
                        if loaded_start_time and loaded_end_time and event_type == "loaded":
                            cost_time = str((loaded_end_time - loaded_start_time).total_seconds())
                        else:
                            cost_time = "N/A"
                        self.mask_events[mask_id].append((loaded_end_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, cost_time))
                    elif re.search(unloaded_start_pattern, line):
                        unloaded_start_time = timestamp
                        if unloaded_start_time and unloaded_end_time and event_type == "unloaded":
                            cost_time = str((unloaded_end_time - unloaded_start_time).total_seconds())
                        else:
                            cost_time = "N/A"
                        self.mask_events[mask_id].append((unloaded_end_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, cost_time))

                # 偵測到結束條件
                if re.search(end_pattern, line) and mask_id:
                    #timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    #timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    #current_workflow.append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], line.strip()))

                    # 確保 mask_id 的列表已初始化
                    if mask_id not in self.workflow_data[workflow_id]:
                        self.workflow_data[workflow_id][mask_id] = []
                    self.workflow_data[workflow_id][mask_id].extend(reversed(current_workflow))
                    
                    workflow_id += 1
                    current_workflow = []
                    mask_id = None
                    loaded_start_time = loaded_end_time = None
                    unloaded_start_time = unloaded_end_time = None

            # 更新進度條
            self.progress_bar.setValue(i + 1)
            self.processed_files += 1
            QApplication.processEvents()

        return self.mask_events, self.workflow_data
    
    def display_load_unload_info(self):
        # 清空 Mask ID 列表
        self.lwt_mask_info.clear()

        self.duration_dict = defaultdict(list)

        # 收集所有事件並排序
        all_events = [(mask_id, timestamp, event_type, cost_time)
                    for mask_id, events in self.mask_events.items()
                    for timestamp, event_type, cost_time in events]

        sorted_events = sorted(all_events, key=lambda x: x[1])  # 按時間排序

        loaded_timestamp = None  # 用於計算載入與卸載之間的時間

        for mask_id, timestamp, event_type, cost_time in sorted_events:
            self.lwt_mask_info.addItem(f"{mask_id} | {timestamp} | {event_type} | => cost {cost_time} secs")

            timestamp_dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S.%f")

            if event_type == "loaded":
                loaded_timestamp = timestamp_dt

            if event_type == "unloaded":
                if loaded_timestamp is None:
                    self.lwt_mask_info.addItem(f"duration : N/A")
                else:
                    duration_seconds = (timestamp_dt - loaded_timestamp).total_seconds()
                    hours, remainder = divmod(duration_seconds, 3600)  # 計算小時和剩餘秒數
                    minutes, seconds = divmod(remainder, 60)  # 計算分鐘和秒數
                    # hh:mm:ss 格式
                    duration_str = f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
                    self.lwt_mask_info.addItem(f"duration : {duration_str}")

                    # 將持續時間記錄到 duration_dict 中
                    self.duration_dict[mask_id].append(duration_seconds)

                    loaded_timestamp = None  # 重置載入時間
                self.lwt_mask_info.addItem(f"===========================================")

        if loaded_timestamp is not None:
            self.lwt_mask_info.addItem(f"duration : N/A")
        
    def display_workflow(self, item):
        selected_text = item.text()
        split_data = selected_text.split(" | ")
        if len(split_data) >= 4:
            mask_id, timestamp_str, event_type, cost_time = split_data
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S.%f")

            mask_event_loaded = r"MaskLoading.MaskLoaded"

            # 遍歷 workflow_data，找到與 mask_id 相符的記錄
            self.lwt_wk_flow.clear()
            wf_print_out = False
            for workflow_id, data in self.workflow_data.items():
                if mask_id in data and not wf_print_out:  # 確保 mask_id 存在於該 workflow_id 的數據中
                    for wf_timestamp_str, wf_line in data[mask_id]:  # 遍歷該 mask_id 的工作流
                        wf_timestamp = datetime.strptime(wf_timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
                        if wf_timestamp <= timestamp:  # 範圍條件
                            # 顯示第一組符合條件的資料後退出
                            #self.lwt_wk_flow.addItem(f"{workflow_id} | {wf_timestamp_str} | {wf_line}")
                            self.lwt_wk_flow.addItem(f"{wf_timestamp_str} | {wf_line}")
                            evnt_match = re.search(mask_event_loaded, wf_line)
                            if evnt_match:
                                self.lwt_wk_flow.addItem(f"-----------------------------------------------------------")
                            wf_print_out = True
            self.lwt_wk_flow.addItem(f"=================================")

    def generate_load_unload_time_chart(self):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        if not self.mask_events or not self.workflow_data:
            self.display_status("No parse for LogService.txt. !", "fail")
            return

        # 收集所有 loaded 事件
        all_loaded_events = [
            (mask_id, timestamp, event_type, cost_time)
            for mask_id, events in self.mask_events.items()
            for timestamp, event_type, cost_time in events if event_type == "loaded"
        ]
        # 收集所有 unloaded 事件
        all_unloaded_events = [
            (mask_id, timestamp, event_type, cost_time)
            for mask_id, events in self.mask_events.items()
            for timestamp, event_type, cost_time in events if event_type == "unloaded"
        ]

        # 按時間排序
        sorted_loaded_events = sorted(all_loaded_events, key=lambda x: datetime.strptime(x[1], "%Y-%m-%d %H:%M:%S.%f"))
        sorted_unloaded_events = sorted(all_unloaded_events, key=lambda x: datetime.strptime(x[1], "%Y-%m-%d %H:%M:%S.%f"))

        # 準備數據以繪製圖表
        loaded_times = {}
        for mask_id, timestamp, event_type, cost_time in sorted_loaded_events:
            if mask_id not in loaded_times:
                loaded_times[mask_id] = []
            loaded_times[mask_id].append(float(cost_time))

        unloaded_times = {}
        for mask_id, timestamp, event_type, cost_time in sorted_unloaded_events:
            if mask_id not in unloaded_times:
                unloaded_times[mask_id] = []
            unloaded_times[mask_id].append(float(cost_time)) 

        self.plot_load_unload_chart(loaded_times, unloaded_times)

    def plot_load_unload_chart(self, loaded_data, unloaded_data):
        self.figure_mask_analysis.clear()

        # 使用 gridspec 設置子圖的高度比例
        gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1])  # 上子圖佔 2， 下子圖佔 1

        ax1 = self.figure_mask_analysis.add_subplot(gs[0])  # 上半部子圖
        ax2 = self.figure_mask_analysis.add_subplot(gs[1])  # 下半部子圖

        # 繪製載入時間圖
        load_times = {mask_id: [cost_time for cost_time in cost_times if cost_time > 0] for mask_id, cost_times in loaded_data.items()}
        self.plot_chart(load_times, "Load Times", "Mask ID", "Cost Time (minutes)", event_type="loaded", ax=ax1)

        # 繪製卸載時間圖
        unload_times = {mask_id: [cost_time for cost_time in cost_times if cost_time > 0] for mask_id, cost_times in unloaded_data.items()}
        self.plot_chart(unload_times, "Unload Times", "Mask ID", "Cost Time (minutes)", event_type="unloaded", ax=ax2)

    def generate_duration_time_chart(self):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        if not self.mask_events or not self.workflow_data:
            self.display_status("No parse for LogService.txt. !", "fail")
            return
        
        self.plot_duration_chart(self.duration_dict)
    
    def plot_chart(self, data, title, xlabel, ylabel, event_type, ax):
        # 準備數據以繪製直條圖
        mask_ids = []
        duration_values = []
        mask_id_count = {}  # 用於計數每個 Mask ID 的出現次數

        for mask_id, cost_times in data.items():
            if cost_times:  # 確保有資料才繪製
                for cost_time in cost_times:
                    # 檢查 Mask ID 是否已經存在，並進行計數
                    if mask_id in mask_id_count:
                        mask_id_count[mask_id] += 1
                        new_mask_id = f"{mask_id}_{mask_id_count[mask_id]}"
                    else:
                        mask_id_count[mask_id] = 1
                        new_mask_id = mask_id

                    mask_ids.append(new_mask_id)
                    duration_values.append(cost_time / 60)  # 將秒轉換為分鐘

        if not mask_ids or not duration_values:
            print("No data to plot.")
            return  # 如果沒有數據，則不繪製圖表

        bar_width = 0.4

        # 繪製直條圖
        bars = ax.bar(mask_ids, duration_values, width=bar_width, color='skyblue')

        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)  # 單位為分鐘
        
        ax.grid(axis='y')  # 只顯示 y 軸的網格線

        # 在每根直條上顯示數值
        for bar in bars:
            yval = bar.get_height()  # 獲取直條的高度（即數值）
            ax.text(bar.get_x() + bar.get_width() / 2, yval, f'{yval:.2f}', ha='center', va='bottom')  # 在直條上方顯示數值

        # 在 y 軸為 10 的地方畫一條水平線
        #ax.axhline(y=10, color='red', linestyle='--', label='Threshold Line at 10') 

        # 調整 x 軸標籤的顯示
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 調整 y 軸範圍（如果需要）
        ax.set_ylim(bottom=0)  # 確保 y 軸從 0 開始

        # 設置 y 軸範圍
        ax.set_ylim(0, 20)

        # 調整邊距
        self.figure_mask_analysis.tight_layout()
        self.canvas_mask_analysis.draw()

    def plot_duration_chart(self, data):
        self.figure_mask_analysis.clear()
        ax = self.figure_mask_analysis.add_subplot(111)

        mask_ids = []
        duration_values = []
        mask_id_count = {}  # 用於計數每個 Mask ID 的出現次數

        for mask_id, durations in data.items():
            if durations:  # 確保有資料才繪製
                for duration in durations:
                    # 檢查 Mask ID 是否已經存在，並進行計數
                    if mask_id in mask_id_count:
                        mask_id_count[mask_id] += 1
                        new_mask_id = f"{mask_id}_{mask_id_count[mask_id]}"
                    else:
                        mask_id_count[mask_id] = 1
                        new_mask_id = mask_id

                    mask_ids.append(new_mask_id)
                    duration_values.append(duration)

        # 將持續時間從秒轉換為小時
        duration_values_in_hours = [duration / 3600 for duration in duration_values]
        
        # 繪製直條圖
        bars = ax.bar(mask_ids, duration_values_in_hours, color='skyblue')

        ax.set_title("Duration Time")
        ax.set_xlabel("Mask ID")
        ax.set_ylabel("Duration Time (hours)")
        
        ax.grid(axis='y')  # 只顯示 y 軸的網格線

        # 在每根直條上顯示數值
        for bar in bars:
            yval = bar.get_height()  # 獲取直條的高度（即數值）
            ax.text(bar.get_x() + bar.get_width() / 2, yval, f'{yval:.2f}', ha='center', va='bottom')  # 在直條上方顯示數值

        # 調整 x 軸標籤的顯示
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 調整 y 軸範圍（如果需要）
        ax.set_ylim(bottom=0)  # 確保 y 軸從 0 開始

        # 調整邊距
        self.figure_mask_analysis.tight_layout()

        self.canvas_mask_analysis.draw()

    def parse_install(self, folder_path):
        install_file = os.path.join(folder_path, "Install.txt")
        if os.path.exists(install_file):
            with open(install_file, 'r', encoding='iso-8859-1') as file:
                lines = file.readlines()
                for line in lines:
                    self.install_info_list.addItem(line.strip())
            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)

    def parse_msc(self, folder_path, msc_file):
        gauge_ids = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3"]
        valve_ids = ["V0-3", "V0-4", "V0-5", "V0-7", "V0-10", "V0-13", "V0-16", "V0-31", "V0-32", "V0-33", "V0-34", "V0-35",
                     "V1-3", "V1-4", "V1-9", "V1-10", "V1-11", "V1-13", "V1-21", "V1-22",
                     "V2-3", "V2-4", "V2-9", "V2-11", "V2-13", "V2-21",
                     "V3-3", "V3-4", "V3-9", "V3-11", "V3-13", "V3-21",
                     "V4-3", "V4-4", "V4-5", "V4-9", "V4-11", "V4-13", "V4-21",
                     "V9-3", "V9-4", "V9-5", "V9-7", "V9-8", "V9-11", "V9-12", "V9-13", "V9-14", "V9-15", "V9-16", "V9-17", "V9-21",
                     "V10-3", "V10-5", "V10-7", "V10-8", "V10-11", "V10-12", "V10-13", "V10-14", "V10-15", "V10-16", "V10-17", "V10-21"]

        selected_format = self.selected_format
        if selected_format == "MSC 2.x":
            ctr_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d+): \(\d+\) PressEvTh\(\): Sent MULTIJET_EVENT_CODE_CURRENT_PRESSURE_HAS_CHANGED\((\d+), (\d+), press=(-?\d+\.\d+), array=(-?\d+\.\d+)\)")
            valve_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d+): \((\d+)\) .*?: Calling SafeMediaCabinetSetValveDefaultState\((\d+),(\d+),(\d+)\)")
        elif selected_format == "MSC 3.x":
            ctr_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d+): \(.*?\) MultiJetImpl::MCPressCurrentValueChangedEvent\((\d+),(\d+)\), .*?pressure = (-?\d+\.\d+) mbar.*")
            valve_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d+): \(\d+\) .*?: Calling MediaCabinetSetValveDefaultState\((\d+),(\d+),(\d+)\)")
        else:
            self.display_status("Unsupported log format selected !", "fail")
            return None, None

        # 讀取 mjnxtdebugXXXXXXXX.log 文件以獲取初始時間
        initial_time = None

        for file_name in msc_file:
            file_path = os.path.join(folder_path, file_name)

            with open(file_path, 'r') as file:
                for line in file:
                    # 第一行包含時間信息
                    date_time_str = line.split(".")[0]
                    initial_time = datetime.strptime(date_time_str, "%Y/%m/%d, %H:%M:%S")
                    break  # 只需獲取第一行的時間
            break  # 只需處理第一個日誌文件

        valve_log_data = {valve_id: [(initial_time, 0)] for valve_id in valve_ids}  # 每個閥門的初始狀態為關閉
        
        time_stamps = []
        gauge_log_data = {gauge_id: [] for gauge_id in gauge_ids}

        for file_name in msc_file:
            file_path = os.path.join(folder_path, file_name)

            with open(file_path, 'r') as file:
                for line in file:
                    # 處理量測儀數據
                    gauge_match = ctr_pattern.search(line)
                    if gauge_match:
                        date_time_str, main_id, sub_id, press_value = gauge_match.groups()[:4]
                        parsed_time = datetime.strptime(date_time_str.split('.')[0], "%Y/%m/%d, %H:%M:%S")
                        time_stamps.append(parsed_time)
                        
                        gauge_id = f"P{main_id}-{sub_id}"
                        if gauge_id in gauge_log_data:
                            gauge_log_data[gauge_id].append((parsed_time, float(press_value)))

                    # 處理閥門狀態
                    valve_match = valve_pattern.search(line)
                    if valve_match:
                        date_time_str, main_id, sub_id, valve_status = valve_match.groups()[:4]
                        parsed_time = datetime.strptime(date_time_str.split('.')[0], "%Y/%m/%d, %H:%M:%S")

                        valve_id = f"V{main_id}-{sub_id}"
                        if valve_id in valve_log_data:
                            # 將新的狀態附加到閥門狀態列表中
                            valve_log_data[valve_id].append((parsed_time, int(valve_status)))

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
        
        # 檢查是否有任何閥門狀態或量測儀數據被記錄
        if not any(valve_log_data.values()) and not any(gauge_log_data.values()):
            return None, None, None
        
        return gauge_log_data, sorted(set(time_stamps)), valve_log_data
    
    def parse_protocol(self, folder_path, protocol_file, process_log):
        self.mask_project_combobox.clear()
        self.protocol_combobox.clear()
        
        gauge_ids = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3"]

        added_projects = set()  # 用於跟踪已添加的項目

        self.protocol_data = {}
        process_data = {gauge_id: [] for gauge_id in gauge_ids} 

        # Parse ProcessLog files
        for file in process_log:
            with open(file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                process_filename = os.path.splitext(os.path.basename(file))[0]  # 去除擴展名

                # 從檔名中提取基準時間
                base_time_str = process_filename.split("_")[1:3]  # 提取檔名中的日期和時間部分
                base_time_str = f"{base_time_str[0]} {base_time_str[1]}"  # 合併為 "YYYY-MM-DD HHMM"
                base_time = datetime.strptime(base_time_str, "%Y-%m-%d %H%M")  # 轉換為 datetime 對象

                gauge_values = {gauge_id: [] for gauge_id in gauge_ids}  # 用於存儲每個量測儀的值

                for line in lines:
                    line = line.strip()  # 去除行首和行尾的空白字符
                    if line.startswith("\"Elapsed Seconds\""):
                        continue  # 跳過標題行

                    parts = line.split(",")
                    if len(parts) < len(gauge_ids) + 2:  # 確保有足夠的欄位, 跳過不完整的行
                        continue

                    # 獲取 CSV 標題行中的量測儀欄位索引
                    header = [col.strip().strip('"') for col in lines[0].strip().split(",")]  # 去除引號
                    gauge_indices = {gauge_id: header.index(gauge_id) for gauge_id in gauge_ids if gauge_id in header}

                    # 依序提取量測儀的值
                    for gauge_id, index in gauge_indices.items():
                        value = float(parts[index].strip())  # 根據索引獲取對應的值
                        gauge_values[gauge_id].append(value)  # 將值存入對應的 gauge_id 列表

                # 計算每個量測儀的平均值並存入 process_data
                for gauge_id in gauge_ids:
                    if gauge_values[gauge_id]:  # 確保有值
                        average_value = sum(gauge_values[gauge_id]) / len(gauge_values[gauge_id])
                        average_value = round(average_value, 6)
                        process_data[gauge_id].append((base_time.strftime("%Y-%m-%d %H%M"), average_value))  # 存入時間戳和平均值

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
            
        # Parse Protocol.txt file
        for file in protocol_file:
            with open(file, 'r', encoding='utf-8') as f:
                lines = f.readlines()

                protocol_filename = os.path.splitext(os.path.basename(file))[0] # 去除擴展名
                self.protocol_combobox.addItem(protocol_filename)

                process_names = []
                u_value = None
                v_value = None
                current_recipe = None
                mask_project_name = None

                for line in lines:
                    line = line.strip()  # 去除行首和行尾的空白字符

                    if line.startswith("[Recipe_"):
                        current_recipe = line.split("[")[1].split("]")[0]  # 獲取 Recipe_#
                        continue

                    if current_recipe:
                        if "ApplicationModule=" in line:
                            am_path = line.split("ApplicationModule=")[1].strip()
                            process_name = os.path.basename(am_path)  # 獲取檔名，例如 Opaque.am
                            process_names.append(f"{current_recipe} => {process_name}")
                            current_recipe = None

                    if "U=" in line:  # 檢查行中是否包含 U=
                        u_match = re.search(r"U\s*=\s*([\d\.]+)", line)  # 匹配 U 的值
                        if u_match:
                            u_value = u_match.group(1)  # 更新 U 值

                    if "V=" in line:  # 檢查行中是否包含 V=
                        v_match = re.search(r"V\s*=\s*([\d\.]+)", line)  # 匹配 V 的值
                        if v_match:
                            v_value = v_match.group(1)  # 更新 V 值
                
                    # 提取 PreRepairImage 路徑中的項目
                    if "PreRepairImage=" in line:
                        image_path = line.split("PreRepairImage=")[1].strip()
                        project_match = re.search(r'\\([^\\]+)\\[^\\]+$', image_path)  # 匹配最後一個目錄名稱
                        if project_match:
                            mask_project_name = project_match.group(1)  # 提取目錄名稱
                            if mask_project_name not in added_projects:
                                self.mask_project_combobox.addItem(mask_project_name)  # 添加到 mask_project_combobox
                                added_projects.add(mask_project_name)  # 將項目添加到集合中
                
                # 從 protocol_filename 提取時間戳
                protocol_base_time_split = protocol_filename.split("_")[1:3]  # 提取檔名中的日期和時間部分
                protocol_base_time_str = f"{protocol_base_time_split[0]} {protocol_base_time_split[1]}"  # 合併為 "YYYY-MM-DD HHMM"

                # 將資料存入字典
                if mask_project_name:
                    if mask_project_name not in self.protocol_data:
                        self.protocol_data[mask_project_name] = {}
                    self.protocol_data[mask_project_name][protocol_filename] = {
                        "process": process_names,
                        "U": u_value,
                        "V": v_value,
                        "process_data": {}
                    }

                    # 只將與 protocol_filename 時間匹配的 gauge 平均值存入
                    for gauge_id in gauge_ids:
                        if gauge_id in process_data and process_data[gauge_id]:
                            for timestamp, value in process_data[gauge_id]:
                                if timestamp == protocol_base_time_str:
                                    self.protocol_data[mask_project_name][protocol_filename]["process_data"][gauge_id] = [(timestamp, value)]

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
            
        self.mask_project_combobox.currentIndexChanged.connect(self.update_protocol_combobox)
        self.protocol_combobox.currentIndexChanged.connect(self.update_protocol_info)

    def on_tab_changed(self, index):
        if index == 3:  # MSC Info 是第四個 Tab，索引為 3
            self.update_protocol_info()

    def update_protocol_combobox(self):
        selected_project = self.mask_project_combobox.currentText()  # 獲取選中的 mask_project_name
        self.protocol_combobox.clear()  # 清空 protocol_combobox

        if selected_project in self.protocol_data:
            for protocol_filename in self.protocol_data[selected_project]:
                self.protocol_combobox.addItem(protocol_filename)  # 添加相關的 protocol_filename

    def update_protocol_info(self):
        selected_project = self.mask_project_combobox.currentText()  # 獲取選中的 mask_project_name
        selected_protocol = self.protocol_combobox.currentText()  # 獲取選中的 Protocol 文件名

        if selected_project in self.protocol_data and selected_protocol in self.protocol_data[selected_project]:
            process_names = self.protocol_data[selected_project][selected_protocol]["process"]
            u_value = self.protocol_data[selected_project][selected_protocol]["U"]
            v_value = self.protocol_data[selected_project][selected_protocol]["V"]

            # 更新 Label
            process_display = "\n".join(process_names) if process_names else "N/A"
            self.process_label.setText(f"Processes:\n{process_display}")
            self.u_label.setText(f"\tU: {u_value if u_value is not None else 'N/A'}")
            self.v_label.setText(f"\tV: {v_value if v_value is not None else 'N/A'}")
        else:
            # 如果沒有找到對應的資料，顯示 N/A
            self.process_label.setText("\tProcess: N/A")
            self.u_label.setText("\tU: N/A")
            self.v_label.setText("\tV: N/A")

    def on_format_selected(self):
        if self.msc_v2_radio.isChecked():
            self.selected_format = "MSC 2.x"
        elif self.msc_v3_radio.isChecked():
            self.selected_format = "MSC 3.x"

    def on_gauge_checkbox_changed(self):
        self.create_chart(generate_chart=True)
    
    def create_chart(self, generate_chart=False):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        
        if self.selected_chart == "mjnxtdebug" and not self.log_data:
            self.display_status("mjnxtdebug*.log parse failed !", "fail")
            return
        elif self.selected_chart == "protocol" and not self.protocol_data:
            self.display_status("Protocol.txt or ProcessLog.csv parse failed !", "fail")
            return
        
        self.ctr_char_gen = False
    
        # 選擇開始和結束時間
        start_date = self.start_date_edit.text()
        start_time = f"{self.start_hour_spinbox.value():02}:{self.start_minute_spinbox.value():02}:{self.start_second_spinbox.value():02}"
        end_date = self.end_date_edit.text()
        end_time = f"{self.end_hour_spinbox.value():02}:{self.end_minute_spinbox.value():02}:{self.end_second_spinbox.value():02}"
        
        # 轉換為 Python datetime
        start_datetime = datetime.strptime(f"{start_date} {start_time}", "%Y/%m/%d %H:%M:%S")
        end_datetime = datetime.strptime(f"{end_date} {end_time}", "%Y/%m/%d %H:%M:%S")
        
        # 過濾數據
        filtered_data = self.filter_data(start_datetime, end_datetime)

        if self.selected_chart == "mjnxtdebug" and not filtered_data and not self.filtered_valve_data:
            self.display_status("No data in the selected time range !", "fail")
            return
        
        # 獲取選擇的量測儀類型
        selected_gauges = [gauge for gauge, checkbox in self.gauge_checkboxes.items() if checkbox.isChecked()]
        
        # 根據選擇的量測儀過濾數據
        if selected_gauges:
            filtered_data = {gauge: filtered_data.get(gauge, []) for gauge in selected_gauges}
        else:
            self.display_status("No gauge selected!", "fail")
            return
        
        if generate_chart:
            self.plot_ctr_chart(filtered_data)
            self.ctr_char_gen = True
            self.display_status("Trend chart is generated", "done")
        else:
            # 選擇儲存路徑
            save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel files (*.xlsx)")
            if not save_path:
                return

            self.save_to_excel(filtered_data, save_path)
            self.display_status("Data is exported to Excel.", "done")
        
    def filter_data(self, start_datetime, end_datetime):
        filtered_data = {}

        if self.selected_chart == "mjnxtdebug":
            for gauge_id, data in self.log_data.items():
                filtered_data[gauge_id] = [(dt, val) for dt, val in data if start_datetime <= dt <= end_datetime]
        elif self.selected_chart == "protocol":
            # 獲取當前選擇的 project file
            current_project = self.mask_project_combobox.currentText()

            # 整理所有 protocol file 的 gauge 平均值
            for _, protocol_info in self.protocol_data[current_project].items():
                process_data = protocol_info["process_data"]
                for gauge_id, gauge_values in process_data.items():
                    for timestamp, value in gauge_values:
                        # 將時間戳轉換為 datetime 對象
                        dt = datetime.strptime(timestamp, "%Y-%m-%d %H%M")
  
                        if gauge_id not in filtered_data:
                            filtered_data[gauge_id] = []
                        filtered_data[gauge_id].append((dt, value))

        return filtered_data
    
    def save_to_excel(self, filtered_data, save_path):
        # 建立新的 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "All Gauge Data"  # 設定工作表名稱

        # 設定表頭
        ws.append(["Date Time"] + [f"{gauge_id} Press Value" for gauge_id in filtered_data.keys()])  # 依序為每個量測儀添加壓力數據的標題

        # 收集所有時間點
        all_times = sorted(set(time for data in filtered_data.values() for time, _ in data))

        # 填充數據
        for row_idx, time in enumerate(all_times, start=2):
            ws.cell(row=row_idx, column=1, value=time)  # 寫入時間
            for col_idx, gauge_id in enumerate(filtered_data.keys(), start=2):
                # 嘗試找到該時間點對應的壓力值
                press_value = next((val for dt, val in filtered_data[gauge_id] if dt == time), None)
                if press_value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=press_value)
                else:
                    # 若無對應壓力值，計算前後壓力值的平均值
                    earlier_values = [val for dt, val in filtered_data[gauge_id] if dt < time]
                    later_values = [val for dt, val in filtered_data[gauge_id] if dt > time]

                    # 獲取最近的前後值
                    earlier_value = earlier_values[-1] if earlier_values else None
                    later_value = later_values[0] if later_values else None

                    # 計算平均值並填入
                    if earlier_value is not None and later_value is not None:
                        average_value = (earlier_value + later_value) / 2
                        ws.cell(row=row_idx, column=col_idx, value=average_value)
                    elif earlier_value is not None:  # 若只有前值
                        ws.cell(row=row_idx, column=col_idx, value=earlier_value)
                    elif later_value is not None:  # 若只有後值
                        ws.cell(row=row_idx, column=col_idx, value=later_value)

        # 建立趨勢圖
        chart = LineChart()
        chart.title = "CTR Pressure Trends"
        chart.x_axis.title = "Date Time"
        chart.y_axis.title = "Press Value (mbar)"

        # 設定數據範圍，從第二列開始，以包含所有量測儀的數據
        data_ref = Reference(ws, min_col=2, min_row=1, max_col=1 + len(filtered_data), max_row=len(all_times) + 1)
        chart.add_data(data_ref, titles_from_data=True)  # 包含標題

        # 設定 X 軸標籤（時間範圍）
        time_ref = Reference(ws, min_col=1, min_row=2, max_row=len(all_times) + 1)
        chart.set_categories(time_ref)

        # 設定 X 軸標籤格式
        chart.x_axis.number_format = "yyyy/mm/dd hh:mm:ss"
        chart.x_axis.majorTimeUnit = "days"
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickLblPos = "low"

        # 添加趨勢圖到工作表
        ws.add_chart(chart, "N2")  # 設定圖表顯示位置

        # 儲存 Excel 檔案
        wb.save(save_path)

    def plot_ctr_chart(self, filtered_data):
        # 清空畫布
        self.figure_ctr.clear()
        ax = self.figure_ctr.add_subplot(111)

        # 準備數據
        self.all_times = sorted(set(time.replace(tzinfo=None) for data in filtered_data.values() for time, _ in data))  # 將所有時間轉換為 offset-naive
        #print("All Times:", all_times)  # 調試輸出
        # 列印最小值和最大值
        #if self.all_times:  # 確保 all_times 不為空
        #    print("X 軸範圍: 最小值 =", min(self.all_times), ", 最大值 =", max(self.all_times))  # 列印 X 軸範圍
        #else:
        #    print("all_times 為空，無法計算最小值和最大值。")
        
        press_values = {gauge_id: [] for gauge_id in filtered_data.keys()}

        for gauge_id, data in filtered_data.items():
            for dt, val in data:
                press_values[gauge_id].append((dt.replace(tzinfo=None), val))  # 將 dt 轉換為 offset-naive

        # 繪製數據
        for gauge_id, values in press_values.items():
            if values:
                times, vals = zip(*values)
                if self.selected_chart == "protocol":
                    ax.plot(times, vals, label=gauge_id, marker='o')  # 使用 marker 繪製數據點
                    # 在每個數據點上顯示數值和時間
                    for time, val in zip(times, vals):
                        ax.text(time, val, f'{val:.6f}\n{time.strftime("%H:%M:%S")}', fontsize=8, ha='center', va='bottom', color='red')
                elif self.selected_chart == "mjnxtdebug":
                    ax.plot(times, vals, label=gauge_id)
    
        ax.set_title("CTR Pressure Trends")
        ax.set_xlabel("Date Time")
        ax.set_ylabel("Press Value (mbar)")
        
        # 調整圖例的位置
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1), framealpha=0.5)  # 將圖例放在右上角，並設置透明度
        ax.grid()

        # 設置 X 軸格式
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d %H:%M:%S"))  # 只顯示時間
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 添加可移動的縱線
        if self.all_times:
            self.vertical_line = ax.axvline(x=self.all_times[0], color='r', linestyle='--')  # 初始位置

            def on_mouse_move(event):
                try:
                    if event.inaxes == ax:  # 確保鼠標在正確的坐標軸上
                        if event.xdata is not None:  # 確保 xdata 是有效的
                            # 將 event.xdata 轉換為 datetime
                            date_time = mdates.num2date(event.xdata)

                            # 格式化 datetime 為字符串
                            formatted_time = date_time.strftime("%Y-%m-%d %H:%M:%S")

                            new_x_values = [event.xdata]  # 將 event.xdata 放入列表中
                            self.vertical_line.set_xdata(new_x_values)  # 更新縱線位置

                            # show the datatime for test
                            self.vline_time_label.setText(f'Time: {formatted_time}')

                            self.figure_ctr.canvas.draw()

                            # 更新閥門狀態顯示
                            if self.valve_status_dialog is not None:
                                current_valve_states = self.get_valve_states_at_time(date_time)
                                self.valve_status_dialog.valve_states = current_valve_states  # 更新 ValveStatusDialog 的閥門狀態
                                self.valve_status_dialog.update_valve_display()  # 調用 ValveStatusDialog 的更新方法

                except Exception as e:
                    print(f"An error occurred: {e}")  # 輸出錯誤信息

            self.figure_ctr.canvas.mpl_connect('motion_notify_event', on_mouse_move)

        # 自動調整 X 軸範圍
        if self.all_times:
            ax.set_xlim([min(self.all_times), max(self.all_times)])  # 設置 X 軸範圍

        # 調整邊距
        plt.subplots_adjust(bottom=0.2)  # 調整底部邊距

        # 顯示圖表
        self.figure_ctr.tight_layout()  # 自動調整子圖參數
        self.canvas_ctr.draw()

    def get_valve_states_at_time(self, date_time):
        # 確保 date_time 是 offset-naive
        if date_time.tzinfo is not None:
            date_time = date_time.replace(tzinfo=None)

        current_states = {}
        
        for valve_id, states in self.valve_log_data.items():
            # 遍歷狀態列表，找到最新的狀態
            for timestamp, state in reversed(states):
                # 確保 timestamp 是 offset-naive
                if timestamp.tzinfo is not None:
                    timestamp = timestamp.replace(tzinfo=None)
                    
                if timestamp <= date_time:
                    current_states[valve_id] = state
                    break  # 找到最新狀態後退出循環

        return current_states

    def zoom_chart(self, zoom_in=True):
        if self.ctr_char_gen == False:
            self.display_status("Trend Chart is unavailabe !", "fail")
            return
        
        ax = self.figure_ctr.get_axes()[0]  # 獲取第一個坐標軸
        xlim = ax.get_xlim()  # 獲取當前 x 軸範圍

        # 獲取選擇的倍率
        #selected_multiplier = self.zoom_combo.currentText().replace('%', '')
        #multiplier = int(selected_multiplier) / 100  # 轉換為小數
        multiplier = 0.1
        
        if zoom_in:
            ax.set_xlim([xlim[0] + (xlim[1] - xlim[0]) * multiplier, xlim[1] - (xlim[1] - xlim[0]) * multiplier])  # 放大 10%
        else:
            ax.set_xlim([xlim[0] - (xlim[1] - xlim[0]) * multiplier, xlim[1] + (xlim[1] - xlim[0]) * multiplier])  # 縮小 10%
        
        self.canvas_ctr.draw()  # 重新繪製圖形

    def shift_chart(self, left_shift=True):
        if self.ctr_char_gen == False:
            self.display_status("Trend Chart is unavailabe !", "fail")
            return
        
        ax = self.figure_ctr.get_axes()[0]  # 獲取第一個坐標軸
        xlim = ax.get_xlim()  # 獲取當前 x 軸範圍

        # 獲取選擇的倍率
        selected_multiplier = self.shift_combo.currentText().replace('%', '')
        multiplier = int(selected_multiplier) / 100  # 轉換為小數
        
        if left_shift:
            ax.set_xlim([xlim[0] - (xlim[1] - xlim[0]) * multiplier, xlim[1] - (xlim[1] - xlim[0]) * multiplier])  # 向左平移 10%
        else:
            ax.set_xlim([xlim[0] + (xlim[1] - xlim[0]) * multiplier, xlim[1] + (xlim[1] - xlim[0]) * multiplier])  # 向右平移 10%
        
        self.canvas_ctr.draw()  # 重新繪製圖形

    def multijet_show(self):
        # 獲取當前腳本的路徑
        current_directory = os.path.dirname(os.path.abspath(__file__))
        image_path = os.path.join(current_directory, "multijet_pipe_sample.png")  # 替換為您的圖片檔名

        dialog = MultijetImgDialog(image_path, self)
        dialog.show()

    def valve_state_show(self):
        # 獲取當前腳本的路徑
        current_directory = os.path.dirname(os.path.abspath(__file__))
        multijet_image_path = os.path.join(current_directory, "multijet_pipe.png")
        if not os.path.exists(multijet_image_path):
            self.display_status("multijet_pipe.png not found !", "fail")
            return

        valve_open_image_path = os.path.join(current_directory, "valve_open.png")
        if not os.path.exists(valve_open_image_path):
            self.display_status("valve_open.png not found !", "fail")
            return

        valve_close_image_path = os.path.join(current_directory, "valve_close.png")
        if not os.path.exists(valve_close_image_path):
            self.display_status("valve_close.png not found !", "fail")
            return

        valve_positions = {
            "V0-3": (174, 163),
            "V0-4": (174, 581),
            "V0-5": (118, 118),
            "V0-7": (118, 163),
            "V0-10": (174, 22),
            "V0-13": (118, 227),
            "V0-16": (174, 286),
            "V0-31": (80, 335),
            "V0-32": (80, 376),
            "V0-33": (80, 416),
            "V0-34": (80, 460),
            "V0-35": (80, 499),
            # CH1
            "V1-3": (896, 22),
            "V1-4": (520, 57),
            "V1-9": (588, 57),
            "V1-10": (804, 56),
            "V1-11": (247, 22),
            "V1-13": (304, 56),
            "V1-21": (956, 22),
            "V1-22": (956, 57),
            #CH2
            "V2-3": (896, 91),
            "V2-4": (520, 125),
            "V2-9": (588, 126),
            "V2-11": (247, 91),
            "V2-13": (304, 125),
            "V2-21": (956, 92),
            #CH3
            "V3-3": (896, 152),
            "V3-4": (520, 186),
            "V3-9": (588, 186),
            "V3-11": (247, 152),
            "V3-13": (304, 186),
            "V3-21": (956, 152),
            #CH4
            "V4-3": (896, 216),
            "V4-4": (478, 245),
            "V4-5": (386, 278),
            "V4-9": (588, 244),
            "V4-11": (247, 216),
            "V4-13": (304, 277),
            "V4-21": (956, 216),
            #CH9
            "V9-3": (896, 343),
            "V9-4": (804, 344),
            "V9-5": (459, 370),
            "V9-7": (626, 387),
            "V9-8": (705, 387),
            "V9-11": (247, 344),
            "V9-12": (784, 428),
            "V9-13": (304, 427),
            "V9-14": (386, 427),
            "V9-15": (386, 387),
            "V9-16": (386, 344),
            "V9-17": (495, 386),
            "V9-21": (956, 344),
            #CH10
            "V10-3": (896, 470),
            "V10-5": (458, 498),
            "V10-7": (626, 511),
            "V10-8": (705, 512),
            "V10-11": (247, 470),
            "V10-12": (788, 552),
            "V10-13": (304, 552),
            "V10-14": (386, 553),
            "V10-15": (386, 511),
            "V10-16": (386, 471),
            "V10-17": (495, 512),
            "V10-21": (956, 470),
        }

        # 創建 ValveStatusDialog 實例並儲存
        self.valve_status_dialog = ValveStatusDialog(multijet_image_path, valve_open_image_path, valve_close_image_path, valve_positions, self)
        self.valve_status_dialog.show()

if __name__ == "__main__":
    app = QApplication([])
    window = LogAnalyser()
    window.show()
    app.exec()
'''
'''
import os
import re
import win32com.client
from datetime import datetime
from collections import defaultdict
import subprocess
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QPushButton, QLabel,
    QLineEdit, QListWidget, QVBoxLayout, QHBoxLayout, QWidget, QProgressBar, QTabWidget,
    QRadioButton, QMessageBox, QSizePolicy, QDateEdit, QSpinBox, QGridLayout, QCheckBox,
    QComboBox, QDialog, QTreeView, QHeaderView, QScrollArea, QButtonGroup, QListWidgetItem
)
from PyQt6.QtCore import Qt, QTimer, QSize, QPoint
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.gridspec as gridspec
from datetime import datetime
from PyQt6.QtGui import QPainter, QPixmap, QIcon, QFileSystemModel, QColor

class MultijetImgDialog(QDialog):
    def __init__(self, image_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("MultiJet Pipe")
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinMaxButtonsHint)  # 添加最小化和最大化按鈕

        # 載入圖片並獲取其大小
        pixmap = QPixmap(image_path)
        self.image_label = QLabel()
        self.image_label.setPixmap(pixmap)  # 設置圖片到 QLabel

        # 設置對話框大小為圖片大小
        self.resize(pixmap.size())  # 將對話框大小設置為圖片大小

        layout = QVBoxLayout()
        layout.addWidget(self.image_label)
        self.setLayout(layout)

class ValveImageLabel(QLabel):
    def __init__(self, pixmap, valve_positions, valve_close_img, parent=None):
        super().__init__(parent)
        self.pixmap = pixmap
        self.valve_positions = valve_positions  # 儲存所有閥門的位置
        self.valve_images = {valve_id: QPixmap(valve_close_img) for valve_id in valve_positions.keys()}  # 儲存閥門圖片

    def paintEvent(self, event):
        painter = QPainter(self)  # 在 QLabel 本身上繪製
        painter.drawPixmap(0, 0, self.pixmap)  # 繪製主圖片

        # 繪製所有閥門圖片
        for valve_id, position in self.valve_positions.items():
            # 確保 position 是 QPoint
            if isinstance(position, tuple):
                position = QPoint(*position)  # 使用 * 解包元組
            painter.drawPixmap(position, self.valve_images[valve_id])  # 繪製閥門圖片

    def sizeHint(self):
        return self.pixmap.size()  # 返回主圖片的大小

    def set_valve_image(self, valve_id, valve_image_path):
        self.valve_images[valve_id] = QPixmap(valve_image_path)
        self.update()  # 更新 QLabel 以顯示新的閥門圖片

class ValveStatusDialog(QDialog):
    def __init__(self, multijet_image_path, valve_open_image_path, valve_close_image_path, valve_positions, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Valve Status")
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinMaxButtonsHint)  # 添加最小化和最大化按鈕

        # 載入主圖片
        self.pixmap = QPixmap(multijet_image_path)
        self.valve_open_image_path = valve_open_image_path
        self.valve_close_image_path = valve_close_image_path
        self.valve_positions = valve_positions

        # 創建自定義 QLabel 來顯示圖片
        self.image_label = ValveImageLabel(self.pixmap, self.valve_positions, self.valve_close_image_path)

        # 創建滾動區域
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(False)  # 使滾動區域可調整大小
        self.scroll_area.setWidget(self.image_label)  # 將自定義 QLabel 設置為滾動區域的內容

        # 設置對話框大小為圖片大小
        self.resize(self.pixmap.size())

        # 創建顯示座標的 QLabel
        self.coordinate_label = QLabel("座標: (0, 0)")
        self.coordinate_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        layout = QVBoxLayout()
        layout.addWidget(self.scroll_area)
        layout.addWidget(self.coordinate_label)  # 添加座標顯示
        self.setLayout(layout)

        # 連接滑鼠移動事件
        self.image_label.setMouseTracking(True)  # 啟用滑鼠追蹤
        self.image_label.mouseMoveEvent = self.mouse_move_event  # 自定義滑鼠移動事件

        # 初始化閥門狀態
        self.valve_states = {valve_id: 0 for valve_id in self.valve_positions.keys()}  # 預設為關閉狀態

        # 繪製閥門
        self.update_valve_display()

    def closeEvent(self, event):
        # 在關閉對話框時釋放資源或進行清理
        print("Closing ValveStatusDialog")  # 可選，顯示關閉信息
        event.accept()  # 確保事件被接受

    def update_valve_display(self):
        for valve_id, position in self.valve_positions.items():
            status = self.valve_states[valve_id]
            valve_image = self.valve_close_image_path if status == 0 else self.valve_open_image_path
            self.image_label.set_valve_image(valve_id, valve_image)  # 更新閥門圖片

    def mouse_move_event(self, event):
        # 獲取滑鼠座標
        x = event.pos().x()
        y = event.pos().y()
        self.coordinate_label.setText(f"座標: ({x}, {y})")  # 更新座標顯示

class LogAnalyser(QMainWindow):
    def __init__(self):
        super().__init__()
        # 設置窗口圖標
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "zeiss-logo.png")  # 替換為您的圖標文件名
        self.setWindowIcon(QIcon(icon_path))
        self.processed_files = 0  # 已解析的檔案數
        self.log_data = None
        self.min_time = None
        self.max_time = None
        self.valve_log_data = {}
        self.valve_status_dialog = None
        # 初始化計時器
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_processing_label)  # 連接計時器的超時信號到更新函數
        self.dot_count = 0
        self.mask_events = {}
        self.workflow_data = {}
        self.ctr_char_gen = False
        self.all_times = []
        self.protocol_data = {}
        self.prot_folder_path = None
        self.prot_start_date = None
        self.prot_end_date = None
        self.last_unchecked_gauge = None
        self.gauge_types = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3", "MFC9-5", "MFC10-5"]

        self.initUI()

    def initUI(self):
        self.setWindowTitle("Log Analyser v1.0.0")
        self.setGeometry(100, 100, 1200, 930)

        # 主 Widget 和 Layout
        central_widget = QWidget()
        central_layout = QVBoxLayout(central_widget)
        self.setCentralWidget(central_widget)
        
        # 創建 QLabel 用於顯示 logo
        logo_label = QLabel()
        logo_pixmap = QPixmap("zeiss-logo.png")
        logo_label.setPixmap(logo_pixmap)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignRight)

        # 將 logo 添加到佈局
        central_layout.addWidget(logo_label)

        # 功能表
        self.create_menu_bar()

        # 顯示資料夾路徑
        self.folder_path_edit = QLineEdit()
        self.folder_path_edit.setFixedHeight(30)
        self.folder_path_edit.setFixedWidth(550)
        central_layout.addWidget(self.folder_path_edit)

        # 建立 Tab
        self.tab_widget = QTabWidget()
        central_layout.addWidget(self.tab_widget)

        #### Tab 1 - 主要資料處理
        wf_widget = QWidget()
        wf_layout = QVBoxLayout(wf_widget)

        # Tab 1 完成，加入到 TabWidget
        self.tab_widget.addTab(wf_widget, "Work Flow")
        
        load_unload_label = QLabel("Load/Unload information")
        work_step_label = QLabel("Work Steps")

        self.lwt_mask_info = QListWidget()
        self.lwt_mask_info.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.lwt_mask_info.itemClicked.connect(self.display_workflow)

        self.lwt_wk_flow = QListWidget()
        self.lwt_wk_flow.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        left_v_layout = QVBoxLayout()
        left_v_layout.addWidget(load_unload_label)
        left_v_layout.addWidget(self.lwt_mask_info)

        right_v_layout = QVBoxLayout()
        right_v_layout.addWidget(work_step_label)
        right_v_layout.addWidget(self.lwt_wk_flow)

        list_h_layout = QHBoxLayout()
        list_h_layout.addLayout(left_v_layout)
        list_h_layout.addLayout(right_v_layout)
        list_h_layout.setStretchFactor(left_v_layout, 1)
        list_h_layout.setStretchFactor(right_v_layout, 1)

        # 創建 QFileSystemModel
        self.file_system_model = QFileSystemModel()

        # 獲取所有可用的磁碟驅動器
        drives = [f"{d}:\\" for d in range(65, 91) if os.path.exists(f"{chr(d)}:\\")]  # ASCII 65-90 對應 A-Z

        # 創建 QTreeView
        self.tree_view = QTreeView()
        self.tree_view.setFixedHeight(170)
        self.tree_view.setModel(self.file_system_model)

        # 添加「本機」作為根節點
        local_machine_index = self.file_system_model.setRootPath('')  # 設置根路徑為空以顯示所有磁碟驅動器
        self.tree_view.setRootIndex(local_machine_index)  # 設置樹的根節點

        # 設置樹狀結構的顯示屬性
        self.tree_view.setHeaderHidden(False)  # 顯示標題
        self.tree_view.setAlternatingRowColors(True)  # 交替行顏色

        # 設置欄位大小可調整
        header = self.tree_view.header()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # 設置欄位為可互動調整大小

        # 將可用的磁碟驅動器添加到樹狀結構
        for drive in drives:
            drive_index = self.file_system_model.index(drive)  # 獲取驅動器的索引
            self.tree_view.setRootIndex(drive_index)  # 設置樹的根節點為該驅動器

        select_folder_button = QPushButton("Select Folder")
        select_folder_button.clicked.connect(self.select_folder_from_tree)

        wf_layout.addLayout(list_h_layout)
        wf_layout.addWidget(self.tree_view)  # 將樹狀結構添加到工作流程佈局中
        wf_layout.addWidget(select_folder_button, alignment=Qt.AlignmentFlag.AlignLeft)

        # 處理按鈕
        self.process_button = QPushButton("Process")
        self.process_button.setFixedSize(80, 40)
        self.process_button.setStyleSheet("""
            QPushButton {
                font-weight: bold;
                background-color: white;
                color: black; 
                border: 2px solid #000;
            }
            QPushButton:hover {
                font-weight: bold;
                background-color: lightblue;
                color: black;
            }
        """)
        self.process_button.clicked.connect(self.process_raw_log)
        central_layout.addWidget(self.process_button)

        # 新增進度條
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)  # 預設範圍為 0 到 100
        self.progress_bar.setValue(0)  # 初始值為 0
        self.progress_bar.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        self.progress_bar.setStyleSheet("QProgressBar::chunk { background-color: lightgreen; }")
        central_layout.addWidget(self.progress_bar)

        process_layout = QHBoxLayout()
        process_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        process_layout.addWidget(self.process_button, alignment=Qt.AlignmentFlag.AlignLeft)
        process_layout.addWidget(self.progress_bar, alignment=Qt.AlignmentFlag.AlignLeft)
        central_layout.addLayout(process_layout)

        # 完成訊息
        self.status_label = QLabel("")
        central_layout.addWidget(self.status_label, alignment=Qt.AlignmentFlag.AlignLeft)

        # Tab 2 - Mask ID Analysis
        mask_analysis_widget = QWidget()
        mask_analysis_v_layout = QVBoxLayout(mask_analysis_widget)
        self.tab_widget.addTab(mask_analysis_widget, "Mask ID Analysis")

        # Add buttons to generate charts
        load_unload_chart_button = QPushButton("Load/Unload Time Chart")
        load_unload_chart_button.clicked.connect(self.generate_load_unload_time_chart)

        duration_chart_button = QPushButton("Duration Time Chart")
        duration_chart_button.clicked.connect(self.generate_duration_time_chart)

        mask_analysis_btn_h_layout = QHBoxLayout()
        mask_analysis_btn_h_layout.addWidget(load_unload_chart_button)
        mask_analysis_btn_h_layout.addWidget(duration_chart_button)
        mask_analysis_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        mask_analysis_v_layout.addLayout(mask_analysis_btn_h_layout)

        # Canvas for displaying charts
        self.figure_mask_analysis = Figure(figsize=(10, 6))
        self.canvas_mask_analysis  = FigureCanvas(self.figure_mask_analysis)
        mask_analysis_v_layout.addWidget(self.canvas_mask_analysis)

        ### Tab 3 - Install Info
        install_widge = QWidget()
        install_layout = QVBoxLayout(install_widge)

        # Tab 3 完成，加入到 TabWidget
        self.tab_widget.addTab(install_widge, "Install Log")

        # Install Info 列表
        self.install_info_list = QListWidget()
        self.install_info_list.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        install_layout.addWidget(self.install_info_list)

        ### Tab 4 - MSC Info
        msc_widge = QWidget()
        msc_v_layout = QVBoxLayout(msc_widge)

        # Tab 4 完成，加入到 TabWidget
        self.tab_widget.addTab(msc_widge, "MSC Info")

        self.tab_widget.currentChanged.connect(self.on_tab_changed)

        # Radio Buttons for msc version 
        label_msc_ver = QLabel("MSC ver : ")
        self.msc_v2_radio = QRadioButton("MSC 2.x")
        self.msc_v2_radio.toggled.connect(self.on_format_selected)
        self.selected_format = "MSC 3.x"  # 預設選擇
        self.msc_v3_radio = QRadioButton("MSC 3.x")
        self.msc_v3_radio.setChecked(True)
        self.msc_v3_radio.toggled.connect(self.on_format_selected)

        msc_version_group = QButtonGroup(self)
        msc_version_group.addButton(self.msc_v2_radio)
        msc_version_group.addButton(self.msc_v3_radio)

        msc_ver_h_layout = QHBoxLayout()

        msc_ver_h_layout.addWidget(label_msc_ver)
        msc_ver_h_layout.addWidget(self.msc_v2_radio)
        msc_ver_h_layout.addWidget(self.msc_v3_radio)

        # Gauge checkbox
        label_gauge_tp = QLabel("   CTR : ")
        self.gauge_checkboxes = {}
        gauge_types = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3", "MFC9-5", "MFC10-5"]
        msc_ver_h_layout.addWidget(label_gauge_tp)
        for gauge in gauge_types:
            checkbox = QCheckBox(gauge)
            checkbox.setChecked(True)  # 預設為選中
            self.gauge_checkboxes[gauge] = checkbox  # 將複選框存儲在字典中
            msc_ver_h_layout.addWidget(checkbox)  # 將複選框添加到布局中

            # 連接複選框的狀態變化事件
            checkbox.stateChanged.connect(self.on_gauge_checkbox_changed)

        msc_ver_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(msc_ver_h_layout)

        # Radio Buttons for chart source 
        label_chart_source = QLabel("Chart source : ")
        self.multijet_chart_radio = QRadioButton("mjnxtdebug")
        self.multijet_chart_radio.setChecked(True)
        self.multijet_chart_radio.toggled.connect(self.on_chart_source_selected)
        self.selected_chart = "mjnxtdebug"  # 預設選擇
        self.protocol_chart_radio = QRadioButton("protocol")
        self.protocol_chart_radio.toggled.connect(self.on_chart_source_selected)

        chart_source_group = QButtonGroup(self)
        chart_source_group.addButton(self.multijet_chart_radio)
        chart_source_group.addButton(self.protocol_chart_radio)
        
        chart_source_h_layout = QHBoxLayout()
        chart_source_h_layout.addWidget(label_chart_source)
        chart_source_h_layout.addWidget(self.multijet_chart_radio)
        chart_source_h_layout.addWidget(self.protocol_chart_radio)

        chart_source_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(chart_source_h_layout)

        # 時間篩選器
        time_filter_layout = QGridLayout()
        #time_filter_layout.setHorizontalSpacing(1)  # 設置水平間距
        #time_filter_layout.setVerticalSpacing(1)     # 設置垂直間距
        #time_filter_layout.setContentsMargins(1, 1, 1, 1)  # 設置邊距為 0

        # 標題列
        header_labels = ["", "Date", "Hr", "Min", "Sec"]
        for col, text in enumerate(header_labels):
            label = QLabel(text)
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)  # 置中對齊
            time_filter_layout.addWidget(label, 0, col)

        # Start Time
        time_filter_layout.addWidget(QLabel("Start:"), 1, 0)

        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("yyyy/MM/dd")
        time_filter_layout.addWidget(self.start_date_edit, 1, 1)

        self.start_hour_spinbox = QSpinBox()
        self.start_hour_spinbox.setRange(0, 23)
        self.start_hour_spinbox.setFixedWidth(60)
        self.start_hour_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_hour_spinbox, 1, 2)

        self.start_minute_spinbox = QSpinBox()
        self.start_minute_spinbox.setRange(0, 59)
        self.start_minute_spinbox.setFixedWidth(60)
        self.start_minute_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_minute_spinbox, 1, 3)

        self.start_second_spinbox = QSpinBox()
        self.start_second_spinbox.setRange(0, 59)
        self.start_second_spinbox.setFixedWidth(60)
        self.start_second_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_second_spinbox, 1, 4)

        # End Time
        time_filter_layout.addWidget(QLabel("End:"), 2, 0)

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy/MM/dd")
        time_filter_layout.addWidget(self.end_date_edit, 2, 1)

        self.end_hour_spinbox = QSpinBox()
        self.end_hour_spinbox.setRange(0, 23)
        self.end_hour_spinbox.setFixedWidth(60)
        self.end_hour_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_hour_spinbox, 2, 2)

        self.end_minute_spinbox = QSpinBox()
        self.end_minute_spinbox.setRange(0, 59)
        self.end_minute_spinbox.setFixedWidth(60)
        self.end_minute_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_minute_spinbox, 2, 3)

        self.end_second_spinbox = QSpinBox()
        self.end_second_spinbox.setRange(0, 59)
        self.end_second_spinbox.setFixedWidth(60)
        self.end_second_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_second_spinbox, 2, 4)

        # Protocol 
        self.mask_project_combobox = QComboBox()
        self.mask_project_combobox.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents) # 自動調整大小
        self.mask_project_combobox.setMinimumWidth(280)
        self.protocol_combobox = QComboBox()
        self.protocol_combobox.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents) # 自動調整大小
        self.protocol_combobox.setMinimumWidth(220)
        self.mask_project_label = QLabel("Project:")
        self.protocol_label = QLabel("Protocol:")
        self.process_label = QLabel("Process:")
        self.u_label = QLabel("U: ")
        self.v_label = QLabel("V: ")

        protocol_h_layout = QHBoxLayout()
        protocol_h_layout.setContentsMargins(0, 0, 0, 0)
        protocol_h_layout.addWidget(self.mask_project_label)
        protocol_h_layout.addWidget(self.mask_project_combobox)
        protocol_h_layout.addWidget(self.protocol_label)
        protocol_h_layout.addWidget(self.protocol_combobox)
        protocol_h_layout.addWidget(self.process_label)
        protocol_h_layout.addWidget(self.u_label)
        protocol_h_layout.addWidget(self.v_label)
        protocol_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(protocol_h_layout)

        # Protocol reader
        protocol_reader_h_layout = QHBoxLayout()
        protocol_reader_button = QPushButton("Protocol Reader V5")
        protocol_reader_button.clicked.connect(lambda: self.open_protocol_reader(self.prot_start_date, self.prot_end_date, self.prot_folder_path))

        protocol_reader_h_layout.addWidget(protocol_reader_button)
        protocol_reader_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(protocol_reader_h_layout)
     
        # Multijet Chart
        multijet_chart_h_layout = QHBoxLayout()
        multijet_pipe_button = QPushButton("MultiJet Pipe")
        multijet_pipe_button.clicked.connect(self.multijet_show)
        valve_status_button = QPushButton("Valve State")
        valve_status_button.clicked.connect(self.valve_state_show)

        multijet_chart_h_layout.addWidget(multijet_pipe_button)
        multijet_chart_h_layout.addWidget(valve_status_button)

        vline_h_layout = QHBoxLayout()
        # Display time data
        self.vline_time_label = QLabel("")
        # Display CTR data
        self.vline_P1_1_label = QLabel("")
        self.vline_P2_1_label = QLabel("")
        self.vline_P3_1_label = QLabel("")
        self.vline_P4_1_label = QLabel("")
        self.vline_P9_1_label = QLabel("")
        self.vline_P9_2_label = QLabel("")
        self.vline_P9_3_label = QLabel("")
        self.vline_P10_1_label = QLabel("")
        self.vline_P10_2_label = QLabel("")
        self.vline_P10_3_label = QLabel("")
        self.vline_MFC9_5_label = QLabel("")
        self.vline_MFC10_5_label = QLabel("")
        vline_h_layout.addWidget(self.vline_time_label)
        vline_h_layout.addWidget(self.vline_P1_1_label)
        vline_h_layout.addWidget(self.vline_P2_1_label)
        vline_h_layout.addWidget(self.vline_P3_1_label)
        vline_h_layout.addWidget(self.vline_P4_1_label)
        vline_h_layout.addWidget(self.vline_P9_1_label)
        vline_h_layout.addWidget(self.vline_P9_2_label)
        vline_h_layout.addWidget(self.vline_P9_3_label)
        vline_h_layout.addWidget(self.vline_P10_1_label)
        vline_h_layout.addWidget(self.vline_P10_2_label)
        vline_h_layout.addWidget(self.vline_P10_3_label)
        vline_h_layout.addWidget(self.vline_MFC9_5_label)
        vline_h_layout.addWidget(self.vline_MFC10_5_label)

        # 讓時間篩選器靠左對齊
        time_filter_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(time_filter_layout)
        multijet_chart_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(multijet_chart_h_layout)
        vline_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(vline_h_layout)

        # 添加畫布以顯示圖表
        self.figure_ctr = Figure(figsize=(10, 6))
        self.canvas_ctr = FigureCanvas(self.figure_ctr)
        msc_v_layout.addWidget(self.canvas_ctr)

        # 添加按鈕以生成趨勢圖
        gen_ctr_button = QPushButton("")
        gen_ctr_button.setToolTip("Generate Chart")  # 設置工具提示
        line_chart_icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "line_chart_icon.png")
        gen_ctr_button.setIcon(QIcon(line_chart_icon_path))  # 設置按鈕圖標
        gen_ctr_button.setIconSize(QSize(30, 30))  # 設置圖標大小
        gen_ctr_button.clicked.connect(lambda: self.create_chart(generate_chart=True))
        expo_button = QPushButton("")
        expo_button.setToolTip("Export to Excel")  # 設置工具提示
        excel_icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_icon.png")
        expo_button.setIcon(QIcon(excel_icon_path))  # 設置按鈕圖標
        expo_button.setIconSize(QSize(30, 30))  # 設置圖標大小
        expo_button.clicked.connect(lambda: self.create_chart(generate_chart=False))

        msc_btn_h_layout = QHBoxLayout()
        msc_btn_h_layout.addWidget(gen_ctr_button)
        msc_btn_h_layout.addWidget(expo_button)
        msc_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)

        # 添加 QComboBox 來選擇移動倍率
        self.shift_combo = QComboBox()
        self.shift_combo.addItems(["10%", "30%", "50%"])
        self.shift_combo.setCurrentIndex(0)  # 預設選擇 10%

        # 添加zoom按鈕
        zoom_in_button = QPushButton("-")
        zoom_in_button.clicked.connect(lambda: self.zoom_chart(zoom_in=False))
        zoom_out_button = QPushButton("+")
        zoom_out_button.clicked.connect(lambda: self.zoom_chart(zoom_in=True))

        # 添加平移按鈕
        left_shift_button = QPushButton("<--")
        left_shift_button.clicked.connect(lambda: self.shift_chart(left_shift=True))
        right_shift_button = QPushButton("-->")
        right_shift_button.clicked.connect(lambda: self.shift_chart(left_shift=False))

        # 將 zoom/shift 按鈕添加到佈局中
        msc_zoom_btn_h_layout = QHBoxLayout()
        msc_zoom_btn_h_layout.addWidget(zoom_in_button)
        msc_zoom_btn_h_layout.addWidget(zoom_out_button)
        msc_zoom_btn_h_layout.addWidget(left_shift_button)
        msc_zoom_btn_h_layout.addWidget(right_shift_button)
        msc_zoom_btn_h_layout.addWidget(self.shift_combo)
        msc_zoom_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignRight)

        msc_v_layout.addLayout(msc_zoom_btn_h_layout)
        msc_v_layout.addLayout(msc_btn_h_layout)        
        
    def create_menu_bar(self):
        menu_bar = self.menuBar()

        # File menu
        file_menu = menu_bar.addMenu("File")
        open_action = file_menu.addAction("Open Folder")
        open_action.triggered.connect(self.select_folder)

        # 新增 Log menu
        log_menu = menu_bar.addMenu("Log")
        self.log_actions = {
            "LogService": log_menu.addAction("LogService"),
            "Install": log_menu.addAction("Install"),
            "mjnxtdebug": log_menu.addAction("mjnxtdebug"),
            "Protocol": log_menu.addAction("Protocol")
        }

        for action in self.log_actions.values():
            action.setCheckable(True)  # 設置為可勾選
            action.setChecked(True)  # 預設為勾選

        # Info menu
        info_menu = menu_bar.addMenu("Info")
        about_action = info_menu.addAction("About")
        about_action.triggered.connect(self.show_about_dialog)

    def update_processing_label(self):
        self.dot_count = (self.dot_count + 1) % 4  # 使點數在 0 到 3 之間循環
        dots = '.' * self.dot_count  # 根據計數生成點數
        self.display_status(f"Processing {dots}", "ongoing")
    
    def show_about_dialog(self):
        QMessageBox.about(self, "About", "Log Analyser v1.0.0\nAuthor : Davian Kuo\nE-mail : davian.kuo@zeiss.com")

    def display_status(self, msg, type):
        self.status_label.setText(msg)
        if type == "done":
            self.status_label.setStyleSheet("background-color: lightgreen;")
        elif type == "fail":
            self.status_label.setStyleSheet("background-color: lightpink;")
        elif type == "ongoing":
            self.status_label.setStyleSheet("background-color: lightyellow;")

    def select_folder_from_tree(self):
        # 獲取當前選中的索引
        selected_index = self.tree_view.currentIndex()
        if selected_index.isValid():  # 確保選中的索引有效
            # 獲取選中的資料夾路徑
            folder_path = self.file_system_model.filePath(selected_index)
            # 將路徑填入 folder_path_edit
            self.folder_path_edit.setText(folder_path)
            self.display_status("Folder selected !", "done")
        else:
            self.display_status("Please select folder !", "fail")
            QMessageBox.information(self, "Log Analyser", "Please select folder !", QMessageBox.StandardButton.Ok)

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.folder_path_edit.setText(folder)
            self.display_status("Folder selected !", "done")

    def set_time_filter_ctrl(self, en):
        self.start_date_edit.setEnabled(en)
        self.start_hour_spinbox.setEnabled(en)
        self.start_minute_spinbox.setEnabled(en)
        self.start_second_spinbox.setEnabled(en)

        self.end_date_edit.setEnabled(en)
        self.end_hour_spinbox.setEnabled(en)
        self.end_minute_spinbox.setEnabled(en)
        self.end_second_spinbox.setEnabled(en)

    def on_chart_source_selected(self):
        if self.multijet_chart_radio.isChecked():
            self.selected_chart = "mjnxtdebug"
            self.set_time_filter_ctrl(True)
        elif self.protocol_chart_radio.isChecked():
            self.selected_chart = "protocol"
            self.set_time_filter_ctrl(False)

        for gauge_id in self.gauge_types:
            self.gauge_checkboxes[gauge_id].setEnabled(True)
            self.gauge_checkboxes[gauge_id].setChecked(True)
    
    def open_protocol_reader(self, start_date, end_date, folder_path):
        if start_date == None or end_date == None or folder_path == None:
            self.display_status("Protocol date is None !", "fail")
            QMessageBox.information(self, "Log Analyser", "Protocol date is None !", QMessageBox.StandardButton.Ok)
            return

        # 獲取當前腳本的路徑
        excel_file_path = os.path.join(folder_path, "Protocol reader_V5.xlsm")  # 替換為你的檔案名

        # 檢查檔案是否存在
        if os.path.exists(excel_file_path):
            try:
                # 啟動 Excel 應用程式
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True

                workbook = excel.Workbooks.Open(excel_file_path)

                excel.Application.Run("SetInformationInputValues", start_date, end_date, folder_path)
                excel.Application.Run("Log_analysis")

                # 保存並關閉工作簿
                #workbook.Save()
                #workbook.Close()
                #excel.Quit()
            except Exception as e:
                print(f"Failed to fill Excel file: {str(e)}")
        else:
            print("The specified Excel file does not exist !")
            self.display_status("The specified Excel file does not exist !", "fail")
    
    def find_log_files(self):
        found_path = self.folder_path_edit.text()
        if not found_path:
            self.display_status("Please choose folder before parsing !", "fail")
            QMessageBox.information(self, "Log Analyser", "Please choose folder before parsing !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None

        logsrvc_log = []
        install_log = []
        msc_log = []
        protocol_log = []
        process_log = []

        # 遞迴搜尋所有子資料夾
        for root, dirs, files in os.walk(found_path):
            for file in files:
                if self.log_actions["LogService"].isChecked():
                    if file == "LogService.txt" or re.match(r"LogService\d{4}-\d{2}-\d{2}_\d{6}\.txt", file):
                        logsrvc_log.append(os.path.join(root, file))

                if self.log_actions["Install"].isChecked():
                    if file == "Install.txt":
                        install_log.append(os.path.join(root, file))

                if self.log_actions["mjnxtdebug"].isChecked():
                    if re.match(r"^mjnxtdebug\d{8}\.log$", file):
                        msc_log.append(os.path.join(root, file))

                if self.log_actions["Protocol"].isChecked():
                    if re.match(r"Protocol_\d{4}-\d{2}-\d{2}_\d{4}_\d{4}\.txt", file):
                        protocol_log.append(os.path.join(root, file))
                    if re.match(r"ProcessLog_\d{4}-\d{2}-\d{2}_\d{4}_\d{4}\.csv", file):
                        process_log.append(os.path.join(root, file))

        # 檢查是否找到有效的日誌文件
        if self.log_actions["LogService"].isChecked() and not logsrvc_log:
            self.display_status("❌ Can't find valid LogService !", "fail")
            QMessageBox.information(self, "Log Analyser", "❌ Can't find valid LogService !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None
        if self.log_actions["Install"].isChecked() and not install_log:
            self.display_status("❌ Can't find valid Install !", "fail")
            QMessageBox.information(self, "Log Analyser", "❌ Can't find valid Install !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None
        if self.log_actions["mjnxtdebug"].isChecked() and not msc_log:
            self.display_status("❌ Can't find valid mjnxtdebug !", "fail")
            QMessageBox.information(self, "Log Analyser", "❌ Can't find valid mjnxtdebug !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None
        if self.log_actions["Protocol"].isChecked():
            if not protocol_log:
                self.display_status("❌ Can't find valid protocol file !", "fail")
                QMessageBox.information(self, "Log Analyser", "❌ Can't find valid protocol file !", QMessageBox.StandardButton.Ok)
                return None, None, None, None, None, None
            if not process_log:
                self.display_status("❌ Can't find valid ProcessLog file !", "fail")
                QMessageBox.information(self, "Log Analyser", "❌ Can't find valid ProcessLog file !", QMessageBox.StandardButton.Ok)
                return None, None, None, None, None, None

        # 進度條初始化
        total_files = len(logsrvc_log) + len(install_log) + len(msc_log) + len(protocol_log) + len(process_log)
        self.progress_bar.setRange(0, total_files)
        self.progress_bar.setValue(0)

        return found_path, logsrvc_log, install_log, msc_log, protocol_log, process_log

    def process_raw_log(self):
        self.mask_events, self.workflow_data = {}, {}
        self.log_data, time_range, self.valve_log_data = {}, [], {}
        self.protocol_data = {}

        if not any(action.isChecked() for action in self.log_actions.values()):
            self.display_status("Please select at least one log to parse !", "fail")
            QMessageBox.information(self, "Log Analyser", "Please select at least one log to parse !", QMessageBox.StandardButton.Ok)
            return

        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        
        self.prot_folder_path = os.path.normpath(folder_path)
        self.process_button.setEnabled(False)
        self.process_button.setStyleSheet("background-color: lightgray; color: darkgray;")
        self.processed_files = 0
        self.display_status("Processing .", "ongoing")

        self.timer.start(100)

        # 1. LogService.txt 
        if logsrvc_file:
            parsed_result = self.parse_logsvr(folder_path, logsrvc_file)
            if parsed_result is None:
                self.timer.stop()
                self.display_status("LogService*.txt parse failed !", "fail")
                QMessageBox.information(self, "Log Analyser", "LogService*.txt parse failed !", QMessageBox.StandardButton.Ok)
                return
            self.mask_events, self.workflow_data = parsed_result

            # Show Mask ID Load/Unload Info
            self.display_load_unload_info()
        else:
            self.mask_events, self.workflow_data = {}, {}
            self.lwt_mask_info.clear()
            self.lwt_wk_flow.clear()

        # 2. Install.txt
        if install_log:
            self.parse_install(folder_path)
        else:
            self.install_info_list.clear()

        # 3. mjnxtdebug.log
        if msc_file:
            self.log_data, time_range, self.valve_log_data = self.parse_msc(folder_path, msc_file)
            if not self.log_data or not self.valve_log_data:
                self.timer.stop()
                self.display_status("mjnxtdebug*.log parse failed !", "fail")
                QMessageBox.information(self, "Log Analyser", "mjnxtdebug*.log parse failed !", QMessageBox.StandardButton.Ok)
                return

            # 獲取時間範圍
            self.min_time = min(time_range)
            self.max_time = max(time_range)
            self.start_date_edit.setMinimumDate(self.min_time.date())
            self.start_date_edit.setMaximumDate(self.max_time.date())
            self.end_date_edit.setMinimumDate(self.min_time.date())
            self.end_date_edit.setMaximumDate(self.max_time.date())

        else:
            self.log_data, time_range, self.valve_log_data = {}, [], {}
            self.figure_ctr.clear()

        # 4. Protocol.txt
        if protocol_file and process_log:
            self.prot_start_date, self.prot_end_date = self.parse_protocol(folder_path, protocol_file, process_log)
        else:
            self.protocol_data = {}
 
        self.timer.stop()
        self.display_status("Process done !", "done")
        QMessageBox.information(self, "Log Analyser", "Process done !", QMessageBox.StandardButton.Ok)
        self.process_button.setEnabled(True)
        self.process_button.setStyleSheet("""
            QPushButton {
                font-weight: bold;
                background-color: white;
                color: black; 
                border: 2px solid #000;
            }
            QPushButton:hover {
                font-weight: bold;
                background-color: lightblue;
                color: black;
            }
        """)

    def parse_logsvr(self, folder_path, logsrvc_file):
        logsrvc_file = sorted(
            logsrvc_file,
            key=lambda f: os.path.getmtime(os.path.join(folder_path, f)),
            reverse=True
        )

        # 儲存數據
        self.mask_events = defaultdict(list)
        self.workflow_data = defaultdict(dict)
        workflow_id = 1

        mask_event_pattern = r"Mask with Id: (\w+\.\d+) (loaded|unloaded)"
        workflow_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into '([^']+)'"
        end_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'Setup.ProtocolDefinition'"
        loaded_start_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'MaskLoading.LoadingMask'"
        unloaded_start_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'MaskUnloading.UnloadingMask'"

        loaded_start_time = None
        loaded_end_time = None
        unloaded_start_time = None
        unloaded_end_time = None

        current_workflow = []
        mask_id = None

        for i, file in enumerate(logsrvc_file):
            file_path = os.path.join(folder_path, file)
            with open(file_path, 'r', encoding='iso-8859-1') as f:
                lines = f.readlines()

            for line in reversed(lines):  # 倒序處理
                # 提取 Mask ID 事件
                mask_match = re.search(mask_event_pattern, line)
                if mask_match:
                    mask_id = mask_match.group(1)
                    event_type = mask_match.group(2)
                    timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    #self.mask_events[mask_id].append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, "N/A"))

                    if event_type == "loaded":
                        loaded_end_time = timestamp
                    elif event_type == "unloaded":
                        unloaded_end_time = timestamp

                # 提取 workflow_pattern
                workflow_match = re.search(workflow_pattern, line)
                if workflow_match and mask_id:
                    timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    current_workflow.append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], workflow_match.group(1)))

                    if re.search(loaded_start_pattern, line):
                        loaded_start_time = timestamp
                        if loaded_start_time and loaded_end_time and event_type == "loaded":
                            cost_time = str((loaded_end_time - loaded_start_time).total_seconds())
                        else:
                            cost_time = "N/A"
                        self.mask_events[mask_id].append((loaded_end_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, cost_time))
                    elif re.search(unloaded_start_pattern, line):
                        unloaded_start_time = timestamp
                        if unloaded_start_time and unloaded_end_time and event_type == "unloaded":
                            cost_time = str((unloaded_end_time - unloaded_start_time).total_seconds())
                        else:
                            cost_time = "N/A"
                        self.mask_events[mask_id].append((unloaded_end_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, cost_time))

                # 偵測到結束條件
                if re.search(end_pattern, line) and mask_id:
                    #timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    #timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    #current_workflow.append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], line.strip()))

                    # 確保 mask_id 的列表已初始化
                    if mask_id not in self.workflow_data[workflow_id]:
                        self.workflow_data[workflow_id][mask_id] = []
                    self.workflow_data[workflow_id][mask_id].extend(reversed(current_workflow))
                    
                    workflow_id += 1
                    current_workflow = []
                    mask_id = None
                    loaded_start_time = loaded_end_time = None
                    unloaded_start_time = unloaded_end_time = None

            # 更新進度條
            self.progress_bar.setValue(i + 1)
            self.processed_files += 1
            QApplication.processEvents()

        return self.mask_events, self.workflow_data
    
    def display_load_unload_info(self):
        # 清空 Mask ID 列表
        self.lwt_mask_info.clear()

        self.duration_dict = defaultdict(list)

        # 收集所有事件並排序
        all_events = [(mask_id, timestamp, event_type, cost_time)
                    for mask_id, events in self.mask_events.items()
                    for timestamp, event_type, cost_time in events]

        sorted_events = sorted(all_events, key=lambda x: x[1])  # 按時間排序

        loaded_timestamp = None  # 用於計算載入與卸載之間的時間

        for mask_id, timestamp, event_type, cost_time in sorted_events:
            self.lwt_mask_info.addItem(f"{mask_id} | {timestamp} | {event_type} | => cost {cost_time} secs")

            timestamp_dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S.%f")

            if event_type == "loaded":
                loaded_timestamp = timestamp_dt

            if event_type == "unloaded":
                if loaded_timestamp is None:
                    self.lwt_mask_info.addItem(f"duration : N/A")
                else:
                    duration_seconds = (timestamp_dt - loaded_timestamp).total_seconds()
                    hours, remainder = divmod(duration_seconds, 3600)  # 計算小時和剩餘秒數
                    minutes, seconds = divmod(remainder, 60)  # 計算分鐘和秒數
                    # hh:mm:ss 格式
                    duration_str = f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
                    self.lwt_mask_info.addItem(f"duration : {duration_str}")

                    # 將持續時間記錄到 duration_dict 中
                    self.duration_dict[mask_id].append(duration_seconds)

                    loaded_timestamp = None  # 重置載入時間
                self.lwt_mask_info.addItem(f"===========================================")

        if loaded_timestamp is not None:
            self.lwt_mask_info.addItem(f"duration : N/A")
        
    def display_workflow(self, item):
        selected_text = item.text()
        split_data = selected_text.split(" | ")
        if len(split_data) >= 4:
            mask_id, timestamp_str, event_type, cost_time = split_data
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S.%f")

            mask_event_loaded = r"MaskLoading.MaskLoaded"

            # 遍歷 workflow_data，找到與 mask_id 相符的記錄
            self.lwt_wk_flow.clear()
            wf_print_out = False
            for workflow_id, data in self.workflow_data.items():
                if mask_id in data and not wf_print_out:  # 確保 mask_id 存在於該 workflow_id 的數據中
                    for wf_timestamp_str, wf_line in data[mask_id]:  # 遍歷該 mask_id 的工作流
                        wf_timestamp = datetime.strptime(wf_timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
                        if wf_timestamp <= timestamp:  # 範圍條件
                            # 顯示第一組符合條件的資料後退出
                            #self.lwt_wk_flow.addItem(f"{workflow_id} | {wf_timestamp_str} | {wf_line}")
                            self.lwt_wk_flow.addItem(f"{wf_timestamp_str} | {wf_line}")
                            evnt_match = re.search(mask_event_loaded, wf_line)
                            if evnt_match:
                                self.lwt_wk_flow.addItem(f"-----------------------------------------------------------")
                            wf_print_out = True
            self.lwt_wk_flow.addItem(f"=================================")

    def generate_load_unload_time_chart(self):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        if not self.mask_events or not self.workflow_data:
            self.display_status("No parse for LogService.txt !", "fail")
            QMessageBox.information(self, "Log Analyser", "No parse for LogService.txt !", QMessageBox.StandardButton.Ok)
            return

        # 收集所有 loaded 事件
        all_loaded_events = [
            (mask_id, timestamp, event_type, cost_time)
            for mask_id, events in self.mask_events.items()
            for timestamp, event_type, cost_time in events if event_type == "loaded"
        ]
        # 收集所有 unloaded 事件
        all_unloaded_events = [
            (mask_id, timestamp, event_type, cost_time)
            for mask_id, events in self.mask_events.items()
            for timestamp, event_type, cost_time in events if event_type == "unloaded"
        ]

        # 按時間排序
        sorted_loaded_events = sorted(all_loaded_events, key=lambda x: datetime.strptime(x[1], "%Y-%m-%d %H:%M:%S.%f"))
        sorted_unloaded_events = sorted(all_unloaded_events, key=lambda x: datetime.strptime(x[1], "%Y-%m-%d %H:%M:%S.%f"))

        # 準備數據以繪製圖表
        loaded_times = {}
        for mask_id, timestamp, event_type, cost_time in sorted_loaded_events:
            if mask_id not in loaded_times:
                loaded_times[mask_id] = []
            loaded_times[mask_id].append(float(cost_time))

        unloaded_times = {}
        for mask_id, timestamp, event_type, cost_time in sorted_unloaded_events:
            if mask_id not in unloaded_times:
                unloaded_times[mask_id] = []
            unloaded_times[mask_id].append(float(cost_time)) 

        self.plot_load_unload_chart(loaded_times, unloaded_times)

    def plot_load_unload_chart(self, loaded_data, unloaded_data):
        self.figure_mask_analysis.clear()

        # 使用 gridspec 設置子圖的高度比例
        gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1])  # 上子圖佔 2， 下子圖佔 1

        ax1 = self.figure_mask_analysis.add_subplot(gs[0])  # 上半部子圖
        ax2 = self.figure_mask_analysis.add_subplot(gs[1])  # 下半部子圖

        # 繪製載入時間圖
        load_times = {mask_id: [cost_time for cost_time in cost_times if cost_time > 0] for mask_id, cost_times in loaded_data.items()}
        self.plot_chart(load_times, "Load Times", "Mask ID", "Cost Time (minutes)", event_type="loaded", ax=ax1)

        # 繪製卸載時間圖
        unload_times = {mask_id: [cost_time for cost_time in cost_times if cost_time > 0] for mask_id, cost_times in unloaded_data.items()}
        self.plot_chart(unload_times, "Unload Times", "Mask ID", "Cost Time (minutes)", event_type="unloaded", ax=ax2)

    def generate_duration_time_chart(self):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        if not self.mask_events or not self.workflow_data:
            self.display_status("No parse for LogService.txt !", "fail")
            QMessageBox.information(self, "Log Analyser", "No parse for LogService.txt !", QMessageBox.StandardButton.Ok)
            return
        
        self.plot_duration_chart(self.duration_dict)
    
    def plot_chart(self, data, title, xlabel, ylabel, event_type, ax):
        # 準備數據以繪製直條圖
        mask_ids = []
        duration_values = []
        mask_id_count = {}  # 用於計數每個 Mask ID 的出現次數

        for mask_id, cost_times in data.items():
            if cost_times:  # 確保有資料才繪製
                for cost_time in cost_times:
                    # 檢查 Mask ID 是否已經存在，並進行計數
                    if mask_id in mask_id_count:
                        mask_id_count[mask_id] += 1
                        new_mask_id = f"{mask_id}_{mask_id_count[mask_id]}"
                    else:
                        mask_id_count[mask_id] = 1
                        new_mask_id = mask_id

                    mask_ids.append(new_mask_id)
                    duration_values.append(cost_time / 60)  # 將秒轉換為分鐘

        if not mask_ids or not duration_values:
            print("No data to plot.")
            return  # 如果沒有數據，則不繪製圖表

        bar_width = 0.4

        # 繪製直條圖
        bars = ax.bar(mask_ids, duration_values, width=bar_width, color='skyblue')

        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)  # 單位為分鐘
        
        ax.grid(axis='y')  # 只顯示 y 軸的網格線

        # 在每根直條上顯示數值
        for bar in bars:
            yval = bar.get_height()  # 獲取直條的高度（即數值）
            ax.text(bar.get_x() + bar.get_width() / 2, yval, f'{yval:.2f}', ha='center', va='bottom')  # 在直條上方顯示數值

        # 在 y 軸為 10 的地方畫一條水平線
        #ax.axhline(y=10, color='red', linestyle='--', label='Threshold Line at 10') 

        # 調整 x 軸標籤的顯示
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 調整 y 軸範圍（如果需要）
        ax.set_ylim(bottom=0)  # 確保 y 軸從 0 開始

        # 設置 y 軸範圍
        ax.set_ylim(0, 20)

        # 調整邊距
        self.figure_mask_analysis.tight_layout()
        self.canvas_mask_analysis.draw()

    def plot_duration_chart(self, data):
        self.figure_mask_analysis.clear()
        ax = self.figure_mask_analysis.add_subplot(111)

        mask_ids = []
        duration_values = []
        mask_id_count = {}  # 用於計數每個 Mask ID 的出現次數

        for mask_id, durations in data.items():
            if durations:  # 確保有資料才繪製
                for duration in durations:
                    # 檢查 Mask ID 是否已經存在，並進行計數
                    if mask_id in mask_id_count:
                        mask_id_count[mask_id] += 1
                        new_mask_id = f"{mask_id}_{mask_id_count[mask_id]}"
                    else:
                        mask_id_count[mask_id] = 1
                        new_mask_id = mask_id

                    mask_ids.append(new_mask_id)
                    duration_values.append(duration)

        # 將持續時間從秒轉換為小時
        duration_values_in_hours = [duration / 3600 for duration in duration_values]
        
        # 繪製直條圖
        bars = ax.bar(mask_ids, duration_values_in_hours, color='skyblue')

        ax.set_title("Duration Time")
        ax.set_xlabel("Mask ID")
        ax.set_ylabel("Duration Time (hours)")
        
        ax.grid(axis='y')  # 只顯示 y 軸的網格線

        # 在每根直條上顯示數值
        for bar in bars:
            yval = bar.get_height()  # 獲取直條的高度（即數值）
            ax.text(bar.get_x() + bar.get_width() / 2, yval, f'{yval:.2f}', ha='center', va='bottom')  # 在直條上方顯示數值

        # 調整 x 軸標籤的顯示
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 調整 y 軸範圍（如果需要）
        ax.set_ylim(bottom=0)  # 確保 y 軸從 0 開始

        # 調整邊距
        self.figure_mask_analysis.tight_layout()

        self.canvas_mask_analysis.draw()

    def parse_install(self, folder_path):
        install_file = os.path.join(folder_path, "Install.txt")
        if os.path.exists(install_file):
            with open(install_file, 'r', encoding='iso-8859-1') as file:
                lines = file.readlines()
                for line in lines:
                    item = QListWidgetItem(line.strip())
                    self.install_info_list.addItem(item)

                # 將最後一行設置為黃色
                if self.install_info_list.count() > 0:  # 確保有項目存在
                    last_item = self.install_info_list.item(self.install_info_list.count() - 1)
                    last_item.setBackground(QColor("yellow"))  # 設置背景顏色為黃色

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)

    def parse_msc(self, folder_path, msc_file):
        gauge_ids = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3", "MFC9-5", "MFC10-5"]
        valve_ids = ["V0-3", "V0-4", "V0-5", "V0-7", "V0-10", "V0-13", "V0-16", "V0-31", "V0-32", "V0-33", "V0-34", "V0-35",
                     "V1-3", "V1-4", "V1-9", "V1-10", "V1-11", "V1-13", "V1-21", "V1-22",
                     "V2-3", "V2-4", "V2-9", "V2-11", "V2-13", "V2-21",
                     "V3-3", "V3-4", "V3-9", "V3-11", "V3-13", "V3-21",
                     "V4-3", "V4-4", "V4-5", "V4-9", "V4-11", "V4-13", "V4-21",
                     "V9-3", "V9-4", "V9-5", "V9-7", "V9-8", "V9-11", "V9-12", "V9-13", "V9-14", "V9-15", "V9-16", "V9-17", "V9-21",
                     "V10-3", "V10-5", "V10-7", "V10-8", "V10-11", "V10-12", "V10-13", "V10-14", "V10-15", "V10-16", "V10-17", "V10-21"]

        selected_format = self.selected_format
        if selected_format == "MSC 2.x":
            ctr_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(\d+\) PressEvTh\(\): Sent MULTIJET_EVENT_CODE_CURRENT_PRESSURE_HAS_CHANGED\((\d+), (\d+), press=(-?\d+\.\d+), array=(-?\d+\.\d+)\)")
            valve_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \((\d+)\) .*?: Calling SafeMediaCabinetSetValveDefaultState\((\d+),(\d+),(\d+)\)")
        elif selected_format == "MSC 3.x":
            ctr_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(.*?\) MultiJetImpl::MCPressCurrentValueChangedEvent\((\d+),(\d+)\), .*?pressure = (-?\d+\.\d+) mbar.*")
            mfc_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(\d+\) MCMFCCurrValueChangedEvent\(ch=(\d+), row=(\d+), value=([\d\.-]+), ValueSent=[\d\.-]+\)")
            valve_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(\d+\) .*?: Calling MediaCabinetSetValveDefaultState\((\d+),(\d+),(\d+)\)")
        else:
            self.display_status("Unsupported log format selected !", "fail")
            QMessageBox.information(self, "Log Analyser", "Unsupported log format selected !", QMessageBox.StandardButton.Ok)
            return None, None

        # 讀取 mjnxtdebugXXXXXXXX.log 文件以獲取初始時間
        initial_time = None

        for file_name in msc_file:
            file_path = os.path.join(folder_path, file_name)
            with open(file_path, 'r') as file:
                for line in file:
                    # 使用正則表達式提取時間字符串
                    time_pattern = re.search(r'(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3})', line)
                    if time_pattern:
                        date_time_str = time_pattern.group(1)  # 獲取匹配的時間字符串
                        initial_time = datetime.strptime(date_time_str, "%Y/%m/%d, %H:%M:%S.%f")
                        break  # 只需獲取第一行的時間
            break  # 只需處理第一個日誌文件

        valve_log_data = {valve_id: [(initial_time, 0)] for valve_id in valve_ids}  # 每個閥門的初始狀態為關閉
        
        time_stamps = []
        gauge_log_data = {gauge_id: [] for gauge_id in gauge_ids}

        for file_name in msc_file:
            file_path = os.path.join(folder_path, file_name)

            with open(file_path, 'r') as file:
                for line in file:
                    # CTR/MFC value
                    gauge_match = ctr_pattern.search(line) or mfc_pattern.search(line)
                    if gauge_match:
                        date_time_str, main_id, sub_id, press_value = gauge_match.groups()[:4]
                        parsed_time = datetime.strptime(date_time_str, "%Y/%m/%d, %H:%M:%S.%f")
                        time_stamps.append(parsed_time)
                        
                        if (main_id == '9' and sub_id == '5') or (main_id == '10' and sub_id == '5'):
                            gauge_id = f"MFC{main_id}-{sub_id}"
                        else:
                            gauge_id = f"P{main_id}-{sub_id}"
                        if gauge_id in gauge_log_data:
                            gauge_log_data[gauge_id].append((parsed_time, float(press_value)))

                    # valve status
                    valve_match = valve_pattern.search(line)
                    if valve_match:
                        date_time_str, main_id, sub_id, valve_status = valve_match.groups()[:4]
                        parsed_time = datetime.strptime(date_time_str, "%Y/%m/%d, %H:%M:%S.%f")

                        valve_id = f"V{main_id}-{sub_id}"
                        if valve_id in valve_log_data:
                            # 將新的狀態附加到閥門狀態列表中
                            valve_log_data[valve_id].append((parsed_time, int(valve_status)))

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
        
        # 檢查是否有任何閥門狀態或量測儀數據被記錄
        if not any(valve_log_data.values()) and not any(gauge_log_data.values()):
            return None, None, None
        
        return gauge_log_data, sorted(set(time_stamps)), valve_log_data
    
    def parse_protocol(self, folder_path, protocol_file, process_log):
        self.mask_project_combobox.clear()
        self.protocol_combobox.clear()
        
        start_date = None
        end_date = None
        gauge_ids = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3"]
        

        added_projects = set()  # 用於跟踪已添加的項目

        self.protocol_data = {}
        process_data = {gauge_id: [] for gauge_id in gauge_ids}

        # 設定閥門的起始值
        valve_ids = {
            "P1-1": "V1-21|22",
            "P2-1": "V2-21",
            "P3-1": "V3-21",
            "P4-1": "V4-21",
            "P9-1": "V9-21",
            "P9-2": "V9-21",
            "P9-3": "V9-21",
            "P10-1": "V10-21",
            "P10-2": "V10-21",
            "P10-3": "V10-21",
        }

        # Parse ProcessLog files
        for file in process_log:
            date_match = re.search(r'ProcessLog_(\d{4})-(\d{2})-(\d{2})', file)
            if date_match:
                date_str = f"{date_match.group(1)}/{date_match.group(2)}/{date_match.group(3)}"  # yyyy/mm/dd
                if start_date is None:
                    start_date = date_str
                end_date = date_str
            with open(file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                process_filename = os.path.splitext(os.path.basename(file))[0]  # 去除擴展名

                # 從檔名中提取基準時間
                base_time_str = process_filename.split("_")[1:3]  # 提取檔名中的日期和時間部分
                base_time_str = f"{base_time_str[0]} {base_time_str[1]}"  # 合併為 "YYYY-MM-DD HHMM"
                base_time = datetime.strptime(base_time_str, "%Y-%m-%d %H%M")  # 轉換為 datetime 對象

                gauge_values = {gauge_id: [] for gauge_id in gauge_ids}  # 用於存儲每個量測儀的值
                valve_values = {valve_id: [] for valve_id in valve_ids.values()}  # 用於存儲閥門的值

                for line in lines:
                    line = line.strip()  # 去除行首和行尾的空白字符
                    if line.startswith("\"Elapsed Seconds\""):
                        continue  # 跳過標題行

                    parts = line.split(",")
                    if len(parts) < len(gauge_ids) + 2:  # 確保有足夠的欄位, 跳過不完整的行
                        continue

                    # 獲取 CSV 標題行中的量測儀欄位索引
                    header = [col.strip().strip('"') for col in lines[0].strip().split(",")]  # 去除引號
                    gauge_indices = {gauge_id: header.index(gauge_id) for gauge_id in gauge_ids if gauge_id in header}
                    valve_indices = {valve_id: header.index(valve_id) for valve_id in valve_ids.values() if valve_id in header}

                    # 依序提取量測儀的值
                    for gauge_id, index in gauge_indices.items():
                        value = float(parts[index].strip())  # 根據索引獲取對應的值
                        gauge_values[gauge_id].append(value)  # 將值存入對應的 gauge_id 列表

                    # 依序提取閥門的值
                    for valve_id, index in valve_indices.items():
                        valve_value = int(parts[index].strip())  # 根據索引獲取對應的閥門值
                        valve_values[valve_id].append(valve_value)  # 將值存入對應的閥門列表

                # 計算每個量測儀的平均值並存入 process_data
                for gauge_id in gauge_ids:
                    if gauge_values[gauge_id]:  # 確保有值
                        # 獲取對應閥門的值
                        valve_id = valve_ids[gauge_id]
                        start_calculating = False
                        total_value = 0
                        count = 0

                        for i in range(len(valve_values[valve_id])):
                            if valve_values[valve_id][i] == 1:
                                start_calculating = True  # 開始計算
                            if start_calculating:
                                total_value += gauge_values[gauge_id][i]
                                count += 1

                        if count > 0:
                            average_value = total_value / count  # 計算平均值
                            average_value = round(average_value, 6)
                            process_data[gauge_id].append((base_time.strftime("%Y-%m-%d %H%M"), average_value))  # 存入時間戳和平均值

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
            
        # Parse Protocol.txt file
        for file in protocol_file:
            with open(file, 'r', encoding='utf-8') as f:
                lines = f.readlines()

                protocol_filename = os.path.splitext(os.path.basename(file))[0] # 去除擴展名
                self.protocol_combobox.addItem(protocol_filename)

                process_names = []
                u_value = None
                v_value = None
                current_recipe = None
                mask_project_name = None

                for line in lines:
                    line = line.strip()  # 去除行首和行尾的空白字符

                    if line.startswith("[Recipe_"):
                        current_recipe = line.split("[")[1].split("]")[0]  # 獲取 Recipe_#
                        continue

                    if current_recipe:
                        if "ApplicationModule=" in line:
                            am_path = line.split("ApplicationModule=")[1].strip()
                            process_name = os.path.basename(am_path)  # 獲取檔名，例如 Opaque.am
                            process_names.append(f"{current_recipe} => {process_name}")
                            current_recipe = None

                    if "U=" in line:  # 檢查行中是否包含 U=
                        u_match = re.search(r"U\s*=\s*([\d\.]+)", line)  # 匹配 U 的值
                        if u_match:
                            u_value = u_match.group(1)  # 更新 U 值

                    if "V=" in line:  # 檢查行中是否包含 V=
                        v_match = re.search(r"V\s*=\s*([\d\.]+)", line)  # 匹配 V 的值
                        if v_match:
                            v_value = v_match.group(1)  # 更新 V 值
                
                    # 提取 PreRepairImage 路徑中的項目
                    if "PreRepairImage=" in line:
                        image_path = line.split("PreRepairImage=")[1].strip()
                        project_match = re.search(r'\\([^\\]+)\\[^\\]+$', image_path)  # 匹配最後一個目錄名稱
                        if project_match:
                            mask_project_name = project_match.group(1)  # 提取目錄名稱
                            if mask_project_name not in added_projects:
                                self.mask_project_combobox.addItem(mask_project_name)  # 添加到 mask_project_combobox
                                added_projects.add(mask_project_name)  # 將項目添加到集合中
                
                # 從 protocol_filename 提取時間戳
                protocol_base_time_split = protocol_filename.split("_")[1:3]  # 提取檔名中的日期和時間部分
                protocol_base_time_str = f"{protocol_base_time_split[0]} {protocol_base_time_split[1]}"  # 合併為 "YYYY-MM-DD HHMM"

                # 將資料存入字典
                if mask_project_name:
                    if mask_project_name not in self.protocol_data:
                        self.protocol_data[mask_project_name] = {}
                    self.protocol_data[mask_project_name][protocol_filename] = {
                        "process": process_names,
                        "U": u_value,
                        "V": v_value,
                        "process_data": {}
                    }

                    # 只將與 protocol_filename 時間匹配的 gauge 平均值存入
                    for gauge_id in gauge_ids:
                        if gauge_id in process_data and process_data[gauge_id]:
                            for timestamp, value in process_data[gauge_id]:
                                if timestamp == protocol_base_time_str:
                                    self.protocol_data[mask_project_name][protocol_filename]["process_data"][gauge_id] = [(timestamp, value)]

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
            
        self.mask_project_combobox.currentIndexChanged.connect(self.update_protocol_combobox)
        self.protocol_combobox.currentIndexChanged.connect(self.update_protocol_info)

        return start_date, end_date

    def on_tab_changed(self, index):
        if index == 3:  # MSC Info 是第四個 Tab，索引為 3
            self.update_protocol_info()

    def update_protocol_combobox(self):
        selected_project = self.mask_project_combobox.currentText()  # 獲取選中的 mask_project_name
        self.protocol_combobox.clear()  # 清空 protocol_combobox

        if selected_project in self.protocol_data:
            for protocol_filename in self.protocol_data[selected_project]:
                self.protocol_combobox.addItem(protocol_filename)  # 添加相關的 protocol_filename

    def update_protocol_info(self):
        selected_project = self.mask_project_combobox.currentText()  # 獲取選中的 mask_project_name
        selected_protocol = self.protocol_combobox.currentText()  # 獲取選中的 Protocol 文件名

        if selected_project in self.protocol_data and selected_protocol in self.protocol_data[selected_project]:
            process_names = self.protocol_data[selected_project][selected_protocol]["process"]
            u_value = self.protocol_data[selected_project][selected_protocol]["U"]
            v_value = self.protocol_data[selected_project][selected_protocol]["V"]

            # 更新 Label
            process_display = "\n".join(process_names) if process_names else "N/A"
            self.process_label.setText(f"Processes:\n{process_display}")
            self.u_label.setText(f"\tU: {u_value if u_value is not None else 'N/A'}")
            self.v_label.setText(f"\tV: {v_value if v_value is not None else 'N/A'}")
        else:
            # 如果沒有找到對應的資料，顯示 N/A
            self.process_label.setText("\tProcess: N/A")
            self.u_label.setText("\tU: N/A")
            self.v_label.setText("\tV: N/A")

    def on_format_selected(self):
        if self.msc_v2_radio.isChecked():
            self.selected_format = "MSC 2.x"
        elif self.msc_v3_radio.isChecked():
            self.selected_format = "MSC 3.x"

    def on_gauge_checkbox_changed(self):
        # 獲取當前觸發事件的複選框
        #checkbox = self.sender()
        
        # 獲取複選框的名稱或 ID
        #gauge_id = None
        #for gauge, cb in self.gauge_checkboxes.items():
        #    if cb == checkbox:
        #        gauge_id = gauge
        #        break

        # 檢查複選框的狀態
        #if checkbox.isChecked():
            #print(f"{gauge_id} 被選中")  # 當前複選框被選中
        #    self.last_unchecked_gauge = None  # 如果選中，重置最後一個未選中的 gauge
        #else:
            #print(f"{gauge_id} 被取消選擇")  # 當前複選框被取消選擇
        #    self.last_unchecked_gauge = gauge_id  # 更新最後一個被取消選擇的 gauge

        #self.create_chart(generate_chart=True)

        pass
    
    def create_chart(self, generate_chart=False):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        
        if self.selected_chart == "mjnxtdebug" and not self.log_data:
            self.display_status("mjnxtdebug*.log parse failed !", "fail")
            QMessageBox.information(self, "Log Analyser", "mjnxtdebug*.log parse failed !", QMessageBox.StandardButton.Ok)
            return
        elif self.selected_chart == "protocol" and not self.protocol_data:
            self.display_status("Protocol.txt or ProcessLog.csv parse failed !", "fail")
            QMessageBox.information(self, "Log Analyser", "Protocol.txt or ProcessLog.csv parse failed !", QMessageBox.StandardButton.Ok)
            return
            
        # 選擇開始和結束時間
        start_date = self.start_date_edit.text()
        start_time = f"{self.start_hour_spinbox.value():02}:{self.start_minute_spinbox.value():02}:{self.start_second_spinbox.value():02}"
        end_date = self.end_date_edit.text()
        end_time = f"{self.end_hour_spinbox.value():02}:{self.end_minute_spinbox.value():02}:{self.end_second_spinbox.value():02}"
        
        # 轉換為 Python datetime
        start_datetime = datetime.strptime(f"{start_date} {start_time}", "%Y/%m/%d %H:%M:%S")
        end_datetime = datetime.strptime(f"{end_date} {end_time}", "%Y/%m/%d %H:%M:%S")
        
        # 過濾數據
        filtered_data = self.filter_data(start_datetime, end_datetime)

        if self.selected_chart == "mjnxtdebug" and all(not values for values in filtered_data.values()):
            self.display_status("No data in selected range !", "fail")
            QMessageBox.information(self, "Log Analyser", "No CTR or Valve data in selected range !\n Please set the time filter.", QMessageBox.StandardButton.Ok)
            
            for gauge_id in self.gauge_types:
                self.gauge_checkboxes[gauge_id].setEnabled(True)
                self.gauge_checkboxes[gauge_id].setChecked(True)
            
            return
        
        # 獲取選擇的量測儀類型
        selected_gauges = [gauge for gauge, checkbox in self.gauge_checkboxes.items() if checkbox.isChecked()]
        
        # 根據選擇的量測儀過濾數據
        if selected_gauges:
            filtered_data = {gauge: filtered_data.get(gauge, []) for gauge in selected_gauges}
        else:
            self.display_status("No gauge selected !", "fail")
            QMessageBox.information(self, "Log Analyser", "No gauge selected !", QMessageBox.StandardButton.Ok)

            # 恢復最後一個被取消選擇的 gauge
            #if self.last_unchecked_gauge:
            #    self.gauge_checkboxes[self.last_unchecked_gauge].setChecked(True)  # 恢復該複選框為選中狀態
                #print(f"恢復選中: {self.last_unchecked_gauge}")  # 顯示恢復的 gauge

            return
        
        if generate_chart:
            self.plot_ctr_chart(filtered_data)
        else:
            self.save_to_excel(filtered_data)
        
    def filter_data(self, start_datetime, end_datetime):
        filtered_data = {}
        gauges_to_disable = []

        if self.selected_chart == "mjnxtdebug":
            #gauges_to_disable = []
            for gauge_id, data in self.log_data.items():
                #filtered_data[gauge_id] = [(dt, val) for dt, val in data if start_datetime <= dt <= end_datetime]
                filtered_values = [(dt, val) for dt, val in data if start_datetime <= dt <= end_datetime]
                filtered_data[gauge_id] = filtered_values

                if not filtered_values:
                    gauges_to_disable.append(gauge_id)
            
        elif self.selected_chart == "protocol":
            # 獲取當前選擇的 project file
            current_project = self.mask_project_combobox.currentText()

            # 整理所有 protocol file 的 gauge 平均值
            for _, protocol_info in self.protocol_data[current_project].items():
                process_data = protocol_info["process_data"]
                for gauge_id, gauge_values in process_data.items():
                    for timestamp, value in gauge_values:
                        # 將時間戳轉換為 datetime 對象
                        dt = datetime.strptime(timestamp, "%Y-%m-%d %H%M")
  
                        if gauge_id not in filtered_data:
                            filtered_data[gauge_id] = []
                        filtered_data[gauge_id].append((dt, value))

            # 檢查 filtered_data 中的 gauge_id 是否存在，若不存在則添加到 gauges_to_disable
            for dis_gauge in self.gauge_checkboxes.keys():
                if dis_gauge not in filtered_data or not filtered_data[gauge_id]:
                    gauges_to_disable.append(dis_gauge)

        # 禁用不符合條件的複選框並設置為未選中
        for gauge_id in self.gauge_checkboxes.keys():
            if gauge_id in gauges_to_disable:
                self.gauge_checkboxes[gauge_id].setEnabled(False)
                self.gauge_checkboxes[gauge_id].setChecked(False)
            else:
                self.gauge_checkboxes[gauge_id].setEnabled(True)
                #self.gauge_checkboxes[gauge_id].setChecked(True)

        return filtered_data
    
    def save_to_excel(self, filtered_data):
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel files (*.xlsx)")
        # 檢查文件是否存在且是否可寫入
        if os.path.exists(save_path):
            try:
                # 嘗試以寫入模式打開文件
                with open(save_path, 'a'):
                    pass  # 如果成功，則文件未被佔用
            except PermissionError:
                # 如果出現 PermissionError，則顯示警告
                QMessageBox.warning(self, "Warning", f"The file '{save_path}' is currently open or cannot be accessed. Please close it and try again.", QMessageBox.StandardButton.Ok)
                self.display_status("Excel save error !", "fail")
                return

        # 建立新的 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "All Gauge Data"  # 設定工作表名稱

        # 設定表頭
        ws.append(["Date Time"] + [f"{gauge_id} Press Value" for gauge_id in filtered_data.keys()])  # 依序為每個量測儀添加壓力數據的標題

        # 收集所有時間點
        all_times = sorted(set(time for data in filtered_data.values() for time, _ in data))

        # 填充數據
        for row_idx, time in enumerate(all_times, start=2):
            # 將時間格式化為字符串，包含毫秒
            formatted_time = time.strftime("%Y/%m/%d, %H:%M:%S.%f")[:-3]  # 只保留到毫秒
            ws.cell(row=row_idx, column=1, value=formatted_time)  # 寫入時間
            for col_idx, gauge_id in enumerate(filtered_data.keys(), start=2):
                # 嘗試找到該時間點對應的壓力值
                press_value = next((val for dt, val in filtered_data[gauge_id] if dt == time), None)
                if press_value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=press_value)
                else:
                    # 若無對應壓力值，計算前後壓力值的平均值
                    earlier_values = [val for dt, val in filtered_data[gauge_id] if dt < time]
                    later_values = [val for dt, val in filtered_data[gauge_id] if dt > time]

                    # 獲取最近的前後值
                    earlier_value = earlier_values[-1] if earlier_values else None
                    later_value = later_values[0] if later_values else None

                    # 計算平均值並填入
                    if earlier_value is not None and later_value is not None:
                        average_value = (earlier_value + later_value) / 2
                        ws.cell(row=row_idx, column=col_idx, value=average_value)
                    elif earlier_value is not None:  # 若只有前值
                        ws.cell(row=row_idx, column=col_idx, value=earlier_value)
                    elif later_value is not None:  # 若只有後值
                        ws.cell(row=row_idx, column=col_idx, value=later_value)

        # 建立趨勢圖
        chart = LineChart()
        chart.title = "CTR Pressure Trends"
        chart.x_axis.title = "Date Time"
        chart.y_axis.title = "Press Value (mbar)"

        # 設定數據範圍，從第二列開始，以包含所有量測儀的數據
        data_ref = Reference(ws, min_col=2, min_row=1, max_col=1 + len(filtered_data), max_row=len(all_times) + 1)
        chart.add_data(data_ref, titles_from_data=True)  # 包含標題

        # 設定 X 軸標籤（時間範圍）
        time_ref = Reference(ws, min_col=1, min_row=2, max_row=len(all_times) + 1)
        chart.set_categories(time_ref)

        # 設定 X 軸標籤格式
        chart.x_axis.number_format = "yyyy/mm/dd hh:mm:ss.000"
        chart.x_axis.majorTimeUnit = "days"
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickLblPos = "low"

        # 添加趨勢圖到工作表
        ws.add_chart(chart, "N2")  # 設定圖表顯示位置

        # 儲存 Excel 檔案
        wb.save(save_path)

        QMessageBox.information(self, "Log Analyser", "Data is exported to Excel.", QMessageBox.StandardButton.Ok)
        self.display_status("Data is exported to Excel.", "done")

    def plot_ctr_chart(self, filtered_data):
        self.ctr_char_gen = False
        # 清空畫布
        self.figure_ctr.clear()
        ax = self.figure_ctr.add_subplot(111)
        #ax2 = ax.twinx()

        # 準備數據
        self.all_times = sorted(set(time.replace(tzinfo=None) for data in filtered_data.values() for time, _ in data))  # 將所有時間轉換為 offset-naive
        #print("All Times:", all_times)  # 調試輸出
        # 列印最小值和最大值
        #if self.all_times:  # 確保 all_times 不為空
        #    print("X 軸範圍: 最小值 =", min(self.all_times), ", 最大值 =", max(self.all_times))  # 列印 X 軸範圍
        #else:
        #    print("all_times 為空，無法計算最小值和最大值。")
        
        press_values = {gauge_id: [] for gauge_id in filtered_data.keys()}

        for gauge_id, data in filtered_data.items():
            for dt, val in data:
                press_values[gauge_id].append((dt.replace(tzinfo=None), val))  # 將 dt 轉換為 offset-naive

        # 繪製數據
        for gauge_id, values in press_values.items():
            if values:
                times, vals = zip(*values)
                if self.selected_chart == "protocol":
                    ax.plot(times, vals, label=gauge_id, marker='o')  # 使用 marker 繪製數據點
                    # 在每個數據點上顯示數值和時間
                    for time, val in zip(times, vals):
                        ax.text(time, val, f'{val:.6f}\n{time.strftime("%H:%M:%S.%f")}', fontsize=8, ha='center', va='bottom', color='red')
                elif self.selected_chart == "mjnxtdebug":
                    ax.plot(times, vals, label=gauge_id)
                    #if gauge_id in ["MFC9-5", "MFC10-5"]:
                    #    ax2.plot(times, vals, label=gauge_id, linestyle='--')  # 使用虛線繪製 SCCM 數據
    
        ax.set_title("MultiJet Trend Chart")
        ax.set_xlabel("Date Time")
        if (self.gauge_checkboxes["MFC9-5"].isChecked() and not self.gauge_checkboxes["MFC10-5"].isChecked() and
            all(not self.gauge_checkboxes[gauge].isChecked() for gauge in self.gauge_checkboxes if gauge not in ["MFC9-5", "MFC10-5"])):
            ax.set_ylabel("Flow Rate (SCCM)")  # 只有當 MFC9-5 被選中且其他都未選中時設置為 SCCM
        elif (self.gauge_checkboxes["MFC10-5"].isChecked() and not self.gauge_checkboxes["MFC9-5"].isChecked() and
            all(not self.gauge_checkboxes[gauge].isChecked() for gauge in self.gauge_checkboxes if gauge not in ["MFC9-5", "MFC10-5"])):
            ax.set_ylabel("Flow Rate (SCCM)")  # 只有當 MFC10-5 被選中且其他都未選中時設置為 SCCM
        elif (self.gauge_checkboxes["MFC10-5"].isChecked() and self.gauge_checkboxes["MFC9-5"].isChecked() and
            all(not self.gauge_checkboxes[gauge].isChecked() for gauge in self.gauge_checkboxes if gauge not in ["MFC9-5", "MFC10-5"])):
            ax.set_ylabel("Flow Rate (SCCM)")
        else:
            ax.set_ylabel("Press Value (mbar)")  # 否則設置為 mbar
        #ax2.set_ylabel("Flow Rate (SCCM)")
        
        # 調整圖例的位置
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1), framealpha=0.5)
        #ax2.legend(loc='lower right', bbox_to_anchor=(1, 1), framealpha=0.5)
        ax.grid()

        # 設置 X 軸格式
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d %H:%M:%S"))  # 只顯示時間
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 添加可移動的縱線
        if self.all_times:
            self.vertical_line = ax.axvline(x=self.all_times[0], color='r', linestyle='--')  # 初始位置

            def on_mouse_move(event):
                try:
                    if event.inaxes == ax:  # 確保鼠標在正確的坐標軸上
                        if event.xdata is not None:  # 確保 xdata 是有效的
                            # 將 event.xdata 轉換為 datetime
                            date_time = mdates.num2date(event.xdata).replace(tzinfo=None)

                            # 格式化 datetime 為字符串
                            formatted_time = date_time.strftime("%Y-%m-%d %H:%M:%S.%f")

                            new_x_values = [event.xdata]  # 將 event.xdata 放入列表中
                            self.vertical_line.set_xdata(new_x_values)  # 更新縱線位置

                            # show the datatime
                            self.vline_time_label.setText(f'<b>Time:</b> {formatted_time}')
                            self.figure_ctr.canvas.draw()

		                    # 獲取當前時間點的 gauge 值
                            for gauge_id in self.gauge_checkboxes.keys():
                                if gauge_id in filtered_data:
                                    # 獲取該 gauge 的數據
                                    gauge_data = filtered_data[gauge_id]
                                    # 找到最近的兩個點
                                    lower_point = None
                                    upper_point = None

                                    for dt, val in gauge_data:
                                        dt = dt.replace(tzinfo=None)
                                        if dt < date_time:
                                            lower_point = (dt, val)
                                        elif dt > date_time and upper_point is None:
                                            upper_point = (dt, val)
                                            break

                                    # 如果找到了兩個點，進行插值
                                    if lower_point and upper_point:
                                        dt1, val1 = lower_point
                                        dt2, val2 = upper_point

                                        # 線性插值計算
                                        slope = (val2 - val1) / (dt2 - dt1).total_seconds()  # 計算斜率
                                        interpolated_value = val1 + slope * (date_time - dt1).total_seconds()  # 計算插值

                                        # 更新相應的 QLabel
                                        getattr(self, f'vline_{gauge_id.replace("-", "_")}_label').setText(f'<b>{gauge_id}:</b> {interpolated_value:.2f} ')
                                    else:
                                        # 若沒有找到足夠的點，則顯示 N/A
                                        getattr(self, f'vline_{gauge_id.replace("-", "_")}_label').setText(f'<b>{gauge_id}:</b> N/A ')

                            # 更新閥門狀態顯示
                            if self.valve_status_dialog is not None:
                                current_valve_states = self.get_valve_states_at_time(date_time)
                                self.valve_status_dialog.valve_states = current_valve_states  # 更新 ValveStatusDialog 的閥門狀態
                                self.valve_status_dialog.update_valve_display()  # 調用 ValveStatusDialog 的更新方法

                except Exception as e:
                    print(f"An error occurred: {e}")  # 輸出錯誤信息

            self.figure_ctr.canvas.mpl_connect('motion_notify_event', on_mouse_move)

        # 自動調整 X 軸範圍
        if self.all_times:
            ax.set_xlim([min(self.all_times), max(self.all_times)])  # 設置 X 軸範圍

        # 調整邊距
        plt.subplots_adjust(bottom=0.2)  # 調整底部邊距

        # 顯示圖表
        self.figure_ctr.tight_layout()  # 自動調整子圖參數
        self.canvas_ctr.draw()

        self.ctr_char_gen = True
        self.display_status("Trend chart is generated", "done")

    def get_valve_states_at_time(self, date_time):
        # 確保 date_time 是 offset-naive
        if date_time.tzinfo is not None:
            date_time = date_time.replace(tzinfo=None)

        current_states = {}
        
        for valve_id, states in self.valve_log_data.items():
            # 遍歷狀態列表，找到最新的狀態
            for timestamp, state in reversed(states):
                # 確保 timestamp 是 offset-naive
                if timestamp.tzinfo is not None:
                    timestamp = timestamp.replace(tzinfo=None)
                    
                if timestamp <= date_time:
                    current_states[valve_id] = state
                    break  # 找到最新狀態後退出循環

        return current_states

    def zoom_chart(self, zoom_in=True):
        if self.ctr_char_gen == False:
            self.display_status("Trend Chart is unavailabe !", "fail")
            QMessageBox.information(self, "Log Analyser", "Trend Chart is unavailabe !", QMessageBox.StandardButton.Ok)
            return
        
        ax = self.figure_ctr.get_axes()[0]  # 獲取第一個坐標軸
        xlim = ax.get_xlim()  # 獲取當前 x 軸範圍

        # 獲取選擇的倍率
        #selected_multiplier = self.zoom_combo.currentText().replace('%', '')
        #multiplier = int(selected_multiplier) / 100  # 轉換為小數
        multiplier = 0.1
        
        if zoom_in:
            ax.set_xlim([xlim[0] + (xlim[1] - xlim[0]) * multiplier, xlim[1] - (xlim[1] - xlim[0]) * multiplier])  # 放大 10%
        else:
            ax.set_xlim([xlim[0] - (xlim[1] - xlim[0]) * multiplier, xlim[1] + (xlim[1] - xlim[0]) * multiplier])  # 縮小 10%
        
        self.canvas_ctr.draw()  # 重新繪製圖形

    def shift_chart(self, left_shift=True):
        if self.ctr_char_gen == False:
            self.display_status("Trend Chart is unavailabe !", "fail")
            QMessageBox.information(self, "Log Analyser", "Trend Chart is unavailabe !", QMessageBox.StandardButton.Ok)
            return
        
        ax = self.figure_ctr.get_axes()[0]  # 獲取第一個坐標軸
        xlim = ax.get_xlim()  # 獲取當前 x 軸範圍

        # 獲取選擇的倍率
        selected_multiplier = self.shift_combo.currentText().replace('%', '')
        multiplier = int(selected_multiplier) / 100  # 轉換為小數
        
        if left_shift:
            ax.set_xlim([xlim[0] - (xlim[1] - xlim[0]) * multiplier, xlim[1] - (xlim[1] - xlim[0]) * multiplier])  # 向左平移 10%
        else:
            ax.set_xlim([xlim[0] + (xlim[1] - xlim[0]) * multiplier, xlim[1] + (xlim[1] - xlim[0]) * multiplier])  # 向右平移 10%
        
        self.canvas_ctr.draw()  # 重新繪製圖形

    def multijet_show(self):
        # 獲取當前腳本的路徑
        current_directory = os.path.dirname(os.path.abspath(__file__))
        image_path = os.path.join(current_directory, "multijet_pipe_sample.png")  # 替換為您的圖片檔名

        dialog = MultijetImgDialog(image_path, self)
        dialog.show()

    def valve_state_show(self):
        # 獲取當前腳本的路徑
        current_directory = os.path.dirname(os.path.abspath(__file__))
        multijet_image_path = os.path.join(current_directory, "multijet_pipe.png")
        if not os.path.exists(multijet_image_path):
            #self.display_status("multijet_pipe.png not found !", "fail")
            print("multijet_pipe.png not found")
            return

        valve_open_image_path = os.path.join(current_directory, "valve_open.png")
        if not os.path.exists(valve_open_image_path):
            #self.display_status("valve_open.png not found !", "fail")
            print("valve_open.png not found")
            return

        valve_close_image_path = os.path.join(current_directory, "valve_close.png")
        if not os.path.exists(valve_close_image_path):
            #self.display_status("valve_close.png not found !", "fail")
            print("valve_close.png not found")
            return

        valve_positions = {
            "V0-3": (174, 163),
            "V0-4": (174, 581),
            "V0-5": (118, 118),
            "V0-7": (118, 163),
            "V0-10": (174, 22),
            "V0-13": (118, 227),
            "V0-16": (174, 286),
            "V0-31": (80, 335),
            "V0-32": (80, 376),
            "V0-33": (80, 416),
            "V0-34": (80, 460),
            "V0-35": (80, 499),
            # CH1
            "V1-3": (896, 22),
            "V1-4": (520, 57),
            "V1-9": (588, 57),
            "V1-10": (804, 56),
            "V1-11": (247, 22),
            "V1-13": (304, 56),
            "V1-21": (956, 22),
            "V1-22": (956, 57),
            #CH2
            "V2-3": (896, 91),
            "V2-4": (520, 125),
            "V2-9": (588, 126),
            "V2-11": (247, 91),
            "V2-13": (304, 125),
            "V2-21": (956, 92),
            #CH3
            "V3-3": (896, 152),
            "V3-4": (520, 186),
            "V3-9": (588, 186),
            "V3-11": (247, 152),
            "V3-13": (304, 186),
            "V3-21": (956, 152),
            #CH4
            "V4-3": (896, 216),
            "V4-4": (478, 245),
            "V4-5": (386, 278),
            "V4-9": (588, 244),
            "V4-11": (247, 216),
            "V4-13": (304, 277),
            "V4-21": (956, 216),
            #CH9
            "V9-3": (896, 343),
            "V9-4": (804, 344),
            "V9-5": (459, 370),
            "V9-7": (626, 387),
            "V9-8": (705, 387),
            "V9-11": (247, 344),
            "V9-12": (784, 428),
            "V9-13": (304, 427),
            "V9-14": (386, 427),
            "V9-15": (386, 387),
            "V9-16": (386, 344),
            "V9-17": (495, 386),
            "V9-21": (956, 344),
            #CH10
            "V10-3": (896, 470),
            "V10-5": (458, 498),
            "V10-7": (626, 511),
            "V10-8": (705, 512),
            "V10-11": (247, 470),
            "V10-12": (788, 552),
            "V10-13": (304, 552),
            "V10-14": (386, 553),
            "V10-15": (386, 511),
            "V10-16": (386, 471),
            "V10-17": (495, 512),
            "V10-21": (956, 470),
        }

        # 創建 ValveStatusDialog 實例並儲存
        self.valve_status_dialog = ValveStatusDialog(multijet_image_path, valve_open_image_path, valve_close_image_path, valve_positions, self)
        self.valve_status_dialog.show()

if __name__ == "__main__":
    app = QApplication([])
    window = LogAnalyser()
    window.show()
    app.exec()'
'''
'''
import sys
import os
import re
import win32com.client
from datetime import datetime
from collections import defaultdict
import subprocess
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QPushButton, QLabel,
    QLineEdit, QListWidget, QVBoxLayout, QHBoxLayout, QWidget, QProgressBar, QTabWidget,
    QRadioButton, QMessageBox, QSizePolicy, QDateEdit, QSpinBox, QGridLayout, QCheckBox,
    QComboBox, QDialog, QTreeView, QHeaderView, QScrollArea, QButtonGroup, QListWidgetItem
)
from PyQt6.QtCore import Qt, QTimer, QSize, QPoint
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.gridspec as gridspec
from datetime import datetime
from PyQt6.QtGui import QPainter, QPixmap, QIcon, QFileSystemModel, QColor

class MultijetImgDialog(QDialog):
    def __init__(self, image_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("MultiJet Pipe")
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinMaxButtonsHint)  # 添加最小化和最大化按鈕

        # 載入圖片並獲取其大小
        pixmap = QPixmap(image_path)
        self.image_label = QLabel()
        self.image_label.setPixmap(pixmap)  # 設置圖片到 QLabel

        # 設置對話框大小為圖片大小
        self.resize(pixmap.size())  # 將對話框大小設置為圖片大小

        layout = QVBoxLayout()
        layout.addWidget(self.image_label)
        self.setLayout(layout)

class ValveImageLabel(QLabel):
    def __init__(self, pixmap, valve_positions, valve_close_img, parent=None):
        super().__init__(parent)
        self.pixmap = pixmap
        self.valve_positions = valve_positions  # 儲存所有閥門的位置
        self.valve_images = {valve_id: QPixmap(valve_close_img) for valve_id in valve_positions.keys()}  # 儲存閥門圖片

    def paintEvent(self, event):
        painter = QPainter(self)  # 在 QLabel 本身上繪製
        painter.drawPixmap(0, 0, self.pixmap)  # 繪製主圖片

        # 繪製所有閥門圖片
        for valve_id, position in self.valve_positions.items():
            # 確保 position 是 QPoint
            if isinstance(position, tuple):
                position = QPoint(*position)  # 使用 * 解包元組
            painter.drawPixmap(position, self.valve_images[valve_id])  # 繪製閥門圖片

    def sizeHint(self):
        return self.pixmap.size()  # 返回主圖片的大小

    def set_valve_image(self, valve_id, valve_image_path):
        self.valve_images[valve_id] = QPixmap(valve_image_path)
        self.update()  # 更新 QLabel 以顯示新的閥門圖片

class ValveStatusDialog(QDialog):
    def __init__(self, multijet_image_path, valve_open_image_path, valve_close_image_path, valve_positions, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Valve Status")
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinMaxButtonsHint)  # 添加最小化和最大化按鈕

        # 載入主圖片
        self.pixmap = QPixmap(multijet_image_path)
        self.valve_open_image_path = valve_open_image_path
        self.valve_close_image_path = valve_close_image_path
        self.valve_positions = valve_positions

        # 創建自定義 QLabel 來顯示圖片
        self.image_label = ValveImageLabel(self.pixmap, self.valve_positions, self.valve_close_image_path)

        # 創建滾動區域
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(False)  # 使滾動區域可調整大小
        self.scroll_area.setWidget(self.image_label)  # 將自定義 QLabel 設置為滾動區域的內容

        # 設置對話框大小為圖片大小
        self.resize(self.pixmap.size())

        # 創建顯示座標的 QLabel
        self.coordinate_label = QLabel("座標: (0, 0)")
        self.coordinate_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        layout = QVBoxLayout()
        layout.addWidget(self.scroll_area)
        layout.addWidget(self.coordinate_label)  # 添加座標顯示
        self.setLayout(layout)

        # 連接滑鼠移動事件
        self.image_label.setMouseTracking(True)  # 啟用滑鼠追蹤
        self.image_label.mouseMoveEvent = self.mouse_move_event  # 自定義滑鼠移動事件

        # 初始化閥門狀態
        self.valve_states = {valve_id: 0 for valve_id in self.valve_positions.keys()}  # 預設為關閉狀態

        # 繪製閥門
        self.update_valve_display()

    def closeEvent(self, event):
        # 在關閉對話框時釋放資源或進行清理
        print("Closing ValveStatusDialog")  # 可選，顯示關閉信息
        event.accept()  # 確保事件被接受

    def update_valve_display(self):
        for valve_id, position in self.valve_positions.items():
            status = self.valve_states[valve_id]
            valve_image = self.valve_close_image_path if status == 0 else self.valve_open_image_path
            self.image_label.set_valve_image(valve_id, valve_image)  # 更新閥門圖片

    def mouse_move_event(self, event):
        # 獲取滑鼠座標
        x = event.pos().x()
        y = event.pos().y()
        self.coordinate_label.setText(f"座標: ({x}, {y})")  # 更新座標顯示

class LogAnalyser(QMainWindow):
    def __init__(self):
        super().__init__()
        # 設置窗口圖標
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "zeiss-logo.png")  # 替換為您的圖標文件名
        self.setWindowIcon(QIcon(icon_path))
        self.processed_files = 0  # 已解析的檔案數
        self.log_data = None
        self.min_time = None
        self.max_time = None
        self.valve_log_data = {}
        self.valve_status_dialog = None
        # 初始化計時器
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_processing_label)  # 連接計時器的超時信號到更新函數
        self.dot_count = 0
        self.mask_events = {}
        self.workflow_data = {}
        self.ctr_char_gen = False
        self.all_times = []
        self.protocol_data = {}
        self.prot_folder_path = None
        self.prot_start_date = None
        self.prot_end_date = None
        self.last_unchecked_gauge = None
        self.gauge_types = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3", "MFC9-5", "MFC10-5"]

        self.initUI()

    def initUI(self):
        self.setWindowTitle("Log Analyser v1.0.0")
        self.setGeometry(100, 100, 1200, 930)

        # 主 Widget 和 Layout
        central_widget = QWidget()
        central_layout = QVBoxLayout(central_widget)
        self.setCentralWidget(central_widget)
        
        # 創建 QLabel 用於顯示 logo
        logo_label = QLabel()
        if getattr(sys, 'frozen', False):
            # 如果應用程序是從可執行檔運行
            logo_pixmap = QPixmap(os.path.join(sys._MEIPASS, "zeiss-logo.png"))
        else:
            # 如果應用程序是從源代碼運行
            logo_pixmap = QPixmap("zeiss-logo.png")
        logo_label.setPixmap(logo_pixmap)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignRight)

        # 將 logo 添加到佈局
        central_layout.addWidget(logo_label)

        # 功能表
        self.create_menu_bar()

        # 顯示資料夾路徑
        self.folder_path_edit = QLineEdit()
        self.folder_path_edit.setFixedHeight(30)
        self.folder_path_edit.setFixedWidth(550)
        central_layout.addWidget(self.folder_path_edit)

        # 建立 Tab
        self.tab_widget = QTabWidget()
        central_layout.addWidget(self.tab_widget)

        #### Tab 1 - 主要資料處理
        wf_widget = QWidget()
        wf_layout = QVBoxLayout(wf_widget)

        # Tab 1 完成，加入到 TabWidget
        self.tab_widget.addTab(wf_widget, "Work Flow")
        
        load_unload_label = QLabel("Load/Unload information")
        work_step_label = QLabel("Work Steps")

        self.lwt_mask_info = QListWidget()
        self.lwt_mask_info.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.lwt_mask_info.itemClicked.connect(self.display_workflow)

        self.lwt_wk_flow = QListWidget()
        self.lwt_wk_flow.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        left_v_layout = QVBoxLayout()
        left_v_layout.addWidget(load_unload_label)
        left_v_layout.addWidget(self.lwt_mask_info)

        right_v_layout = QVBoxLayout()
        right_v_layout.addWidget(work_step_label)
        right_v_layout.addWidget(self.lwt_wk_flow)

        list_h_layout = QHBoxLayout()
        list_h_layout.addLayout(left_v_layout)
        list_h_layout.addLayout(right_v_layout)
        list_h_layout.setStretchFactor(left_v_layout, 1)
        list_h_layout.setStretchFactor(right_v_layout, 1)

        # 創建 QFileSystemModel
        self.file_system_model = QFileSystemModel()

        # 獲取所有可用的磁碟驅動器
        drives = [f"{d}:\\" for d in range(65, 91) if os.path.exists(f"{chr(d)}:\\")]  # ASCII 65-90 對應 A-Z

        # 創建 QTreeView
        self.tree_view = QTreeView()
        self.tree_view.setFixedHeight(170)
        self.tree_view.setModel(self.file_system_model)

        # 添加「本機」作為根節點
        local_machine_index = self.file_system_model.setRootPath('')  # 設置根路徑為空以顯示所有磁碟驅動器
        self.tree_view.setRootIndex(local_machine_index)  # 設置樹的根節點

        # 設置樹狀結構的顯示屬性
        self.tree_view.setHeaderHidden(False)  # 顯示標題
        self.tree_view.setAlternatingRowColors(True)  # 交替行顏色

        # 設置欄位大小可調整
        header = self.tree_view.header()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # 設置欄位為可互動調整大小

        # 將可用的磁碟驅動器添加到樹狀結構
        for drive in drives:
            drive_index = self.file_system_model.index(drive)  # 獲取驅動器的索引
            self.tree_view.setRootIndex(drive_index)  # 設置樹的根節點為該驅動器

        select_folder_button = QPushButton("Select Folder")
        select_folder_button.clicked.connect(self.select_folder_from_tree)

        wf_layout.addLayout(list_h_layout)
        wf_layout.addWidget(self.tree_view)  # 將樹狀結構添加到工作流程佈局中
        wf_layout.addWidget(select_folder_button, alignment=Qt.AlignmentFlag.AlignLeft)

        # 處理按鈕
        self.process_button = QPushButton("Process")
        self.process_button.setFixedSize(80, 40)
        self.process_button.setStyleSheet("""
            QPushButton {
                font-weight: bold;
                background-color: white;
                color: black; 
                border: 2px solid #000;
            }
            QPushButton:hover {
                font-weight: bold;
                background-color: lightblue;
                color: black;
            }
        """)
        self.process_button.clicked.connect(self.process_raw_log)
        central_layout.addWidget(self.process_button)

        # 新增進度條
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)  # 預設範圍為 0 到 100
        self.progress_bar.setValue(0)  # 初始值為 0
        self.progress_bar.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        self.progress_bar.setStyleSheet("QProgressBar::chunk { background-color: lightgreen; }")
        central_layout.addWidget(self.progress_bar)

        process_layout = QHBoxLayout()
        process_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        process_layout.addWidget(self.process_button, alignment=Qt.AlignmentFlag.AlignLeft)
        process_layout.addWidget(self.progress_bar, alignment=Qt.AlignmentFlag.AlignLeft)
        central_layout.addLayout(process_layout)

        # 完成訊息
        self.status_label = QLabel("")
        central_layout.addWidget(self.status_label, alignment=Qt.AlignmentFlag.AlignLeft)

        # Tab 2 - Mask ID Analysis
        mask_analysis_widget = QWidget()
        mask_analysis_v_layout = QVBoxLayout(mask_analysis_widget)
        self.tab_widget.addTab(mask_analysis_widget, "Mask ID Analysis")

        # Add buttons to generate charts
        load_unload_chart_button = QPushButton("Load/Unload Time Chart")
        load_unload_chart_button.clicked.connect(self.generate_load_unload_time_chart)

        duration_chart_button = QPushButton("Duration Time Chart")
        duration_chart_button.clicked.connect(self.generate_duration_time_chart)

        mask_analysis_btn_h_layout = QHBoxLayout()
        mask_analysis_btn_h_layout.addWidget(load_unload_chart_button)
        mask_analysis_btn_h_layout.addWidget(duration_chart_button)
        mask_analysis_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        mask_analysis_v_layout.addLayout(mask_analysis_btn_h_layout)

        # Canvas for displaying charts
        self.figure_mask_analysis = Figure(figsize=(10, 6))
        self.canvas_mask_analysis  = FigureCanvas(self.figure_mask_analysis)
        mask_analysis_v_layout.addWidget(self.canvas_mask_analysis)

        ### Tab 3 - Install Info
        install_widge = QWidget()
        install_layout = QVBoxLayout(install_widge)

        # Tab 3 完成，加入到 TabWidget
        self.tab_widget.addTab(install_widge, "Install Log")

        # Install Info 列表
        self.install_info_list = QListWidget()
        self.install_info_list.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        install_layout.addWidget(self.install_info_list)

        ### Tab 4 - MSC Info
        msc_widge = QWidget()
        msc_v_layout = QVBoxLayout(msc_widge)

        # Tab 4 完成，加入到 TabWidget
        self.tab_widget.addTab(msc_widge, "MSC Info")

        self.tab_widget.currentChanged.connect(self.on_tab_changed)

        # Radio Buttons for msc version 
        label_msc_ver = QLabel("MSC ver : ")
        self.msc_v2_radio = QRadioButton("MSC 2.x")
        self.msc_v2_radio.toggled.connect(self.on_format_selected)
        self.selected_format = "MSC 3.x"  # 預設選擇
        self.msc_v3_radio = QRadioButton("MSC 3.x")
        self.msc_v3_radio.setChecked(True)
        self.msc_v3_radio.toggled.connect(self.on_format_selected)

        msc_version_group = QButtonGroup(self)
        msc_version_group.addButton(self.msc_v2_radio)
        msc_version_group.addButton(self.msc_v3_radio)

        msc_ver_h_layout = QHBoxLayout()

        msc_ver_h_layout.addWidget(label_msc_ver)
        msc_ver_h_layout.addWidget(self.msc_v2_radio)
        msc_ver_h_layout.addWidget(self.msc_v3_radio)

        # Gauge checkbox
        label_gauge_tp = QLabel("   CTR : ")
        self.gauge_checkboxes = {}
        gauge_types = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3", "MFC9-5", "MFC10-5"]
        msc_ver_h_layout.addWidget(label_gauge_tp)
        for gauge in gauge_types:
            checkbox = QCheckBox(gauge)
            checkbox.setChecked(True)  # 預設為選中
            self.gauge_checkboxes[gauge] = checkbox  # 將複選框存儲在字典中
            msc_ver_h_layout.addWidget(checkbox)  # 將複選框添加到布局中

            # 連接複選框的狀態變化事件
            checkbox.stateChanged.connect(self.on_gauge_checkbox_changed)

        msc_ver_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(msc_ver_h_layout)

        # Radio Buttons for chart source 
        label_chart_source = QLabel("Chart source : ")
        self.multijet_chart_radio = QRadioButton("mjnxtdebug")
        self.multijet_chart_radio.setChecked(True)
        self.multijet_chart_radio.toggled.connect(self.on_chart_source_selected)
        self.selected_chart = "mjnxtdebug"  # 預設選擇
        self.protocol_chart_radio = QRadioButton("protocol")
        self.protocol_chart_radio.toggled.connect(self.on_chart_source_selected)

        chart_source_group = QButtonGroup(self)
        chart_source_group.addButton(self.multijet_chart_radio)
        chart_source_group.addButton(self.protocol_chart_radio)
        
        chart_source_h_layout = QHBoxLayout()
        chart_source_h_layout.addWidget(label_chart_source)
        chart_source_h_layout.addWidget(self.multijet_chart_radio)
        chart_source_h_layout.addWidget(self.protocol_chart_radio)

        chart_source_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(chart_source_h_layout)

        # 時間篩選器
        time_filter_layout = QGridLayout()
        #time_filter_layout.setHorizontalSpacing(1)  # 設置水平間距
        #time_filter_layout.setVerticalSpacing(1)     # 設置垂直間距
        #time_filter_layout.setContentsMargins(1, 1, 1, 1)  # 設置邊距為 0

        # 標題列
        header_labels = ["", "Date", "Hr", "Min", "Sec"]
        for col, text in enumerate(header_labels):
            label = QLabel(text)
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)  # 置中對齊
            time_filter_layout.addWidget(label, 0, col)

        # Start Time
        time_filter_layout.addWidget(QLabel("Start:"), 1, 0)

        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("yyyy/MM/dd")
        time_filter_layout.addWidget(self.start_date_edit, 1, 1)

        self.start_hour_spinbox = QSpinBox()
        self.start_hour_spinbox.setRange(0, 23)
        self.start_hour_spinbox.setFixedWidth(60)
        self.start_hour_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_hour_spinbox, 1, 2)

        self.start_minute_spinbox = QSpinBox()
        self.start_minute_spinbox.setRange(0, 59)
        self.start_minute_spinbox.setFixedWidth(60)
        self.start_minute_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_minute_spinbox, 1, 3)

        self.start_second_spinbox = QSpinBox()
        self.start_second_spinbox.setRange(0, 59)
        self.start_second_spinbox.setFixedWidth(60)
        self.start_second_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_second_spinbox, 1, 4)

        # End Time
        time_filter_layout.addWidget(QLabel("End:"), 2, 0)

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy/MM/dd")
        time_filter_layout.addWidget(self.end_date_edit, 2, 1)

        self.end_hour_spinbox = QSpinBox()
        self.end_hour_spinbox.setRange(0, 23)
        self.end_hour_spinbox.setFixedWidth(60)
        self.end_hour_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_hour_spinbox, 2, 2)

        self.end_minute_spinbox = QSpinBox()
        self.end_minute_spinbox.setRange(0, 59)
        self.end_minute_spinbox.setFixedWidth(60)
        self.end_minute_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_minute_spinbox, 2, 3)

        self.end_second_spinbox = QSpinBox()
        self.end_second_spinbox.setRange(0, 59)
        self.end_second_spinbox.setFixedWidth(60)
        self.end_second_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_second_spinbox, 2, 4)

        # Protocol 
        self.mask_project_combobox = QComboBox()
        self.mask_project_combobox.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents) # 自動調整大小
        self.mask_project_combobox.setMinimumWidth(280)
        self.protocol_combobox = QComboBox()
        self.protocol_combobox.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents) # 自動調整大小
        self.protocol_combobox.setMinimumWidth(220)
        self.mask_project_label = QLabel("Project:")
        self.protocol_label = QLabel("Protocol:")
        self.process_label = QLabel("Process:")
        self.u_label = QLabel("U: ")
        self.v_label = QLabel("V: ")

        protocol_h_layout = QHBoxLayout()
        protocol_h_layout.setContentsMargins(0, 0, 0, 0)
        protocol_h_layout.addWidget(self.mask_project_label)
        protocol_h_layout.addWidget(self.mask_project_combobox)
        protocol_h_layout.addWidget(self.protocol_label)
        protocol_h_layout.addWidget(self.protocol_combobox)
        protocol_h_layout.addWidget(self.process_label)
        protocol_h_layout.addWidget(self.u_label)
        protocol_h_layout.addWidget(self.v_label)
        protocol_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(protocol_h_layout)

        # Protocol reader
        protocol_reader_h_layout = QHBoxLayout()
        protocol_reader_button = QPushButton("Protocol Reader V5")
        protocol_reader_button.clicked.connect(lambda: self.open_protocol_reader(self.prot_start_date, self.prot_end_date, self.prot_folder_path))

        protocol_reader_h_layout.addWidget(protocol_reader_button)
        protocol_reader_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(protocol_reader_h_layout)
     
        # Multijet Chart
        multijet_chart_h_layout = QHBoxLayout()
        multijet_pipe_button = QPushButton("MultiJet Pipe")
        multijet_pipe_button.clicked.connect(self.multijet_show)
        valve_status_button = QPushButton("Valve State")
        valve_status_button.clicked.connect(self.valve_state_show)

        multijet_chart_h_layout.addWidget(multijet_pipe_button)
        multijet_chart_h_layout.addWidget(valve_status_button)

        vline_h_layout = QHBoxLayout()
        # Display time data
        self.vline_time_label = QLabel("")
        # Display CTR data
        self.vline_P1_1_label = QLabel("")
        self.vline_P2_1_label = QLabel("")
        self.vline_P3_1_label = QLabel("")
        self.vline_P4_1_label = QLabel("")
        self.vline_P9_1_label = QLabel("")
        self.vline_P9_2_label = QLabel("")
        self.vline_P9_3_label = QLabel("")
        self.vline_P10_1_label = QLabel("")
        self.vline_P10_2_label = QLabel("")
        self.vline_P10_3_label = QLabel("")
        self.vline_MFC9_5_label = QLabel("")
        self.vline_MFC10_5_label = QLabel("")
        vline_h_layout.addWidget(self.vline_time_label)
        vline_h_layout.addWidget(self.vline_P1_1_label)
        vline_h_layout.addWidget(self.vline_P2_1_label)
        vline_h_layout.addWidget(self.vline_P3_1_label)
        vline_h_layout.addWidget(self.vline_P4_1_label)
        vline_h_layout.addWidget(self.vline_P9_1_label)
        vline_h_layout.addWidget(self.vline_P9_2_label)
        vline_h_layout.addWidget(self.vline_P9_3_label)
        vline_h_layout.addWidget(self.vline_P10_1_label)
        vline_h_layout.addWidget(self.vline_P10_2_label)
        vline_h_layout.addWidget(self.vline_P10_3_label)
        vline_h_layout.addWidget(self.vline_MFC9_5_label)
        vline_h_layout.addWidget(self.vline_MFC10_5_label)

        # 讓時間篩選器靠左對齊
        time_filter_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(time_filter_layout)
        multijet_chart_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(multijet_chart_h_layout)
        vline_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(vline_h_layout)

        # 添加畫布以顯示圖表
        self.figure_ctr = Figure(figsize=(10, 6))
        self.canvas_ctr = FigureCanvas(self.figure_ctr)
        msc_v_layout.addWidget(self.canvas_ctr)

        # 添加按鈕以生成趨勢圖
        gen_ctr_button = QPushButton("")
        gen_ctr_button.setToolTip("Generate Chart")  # 設置工具提示
        line_chart_icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "line_chart_icon.png")
        gen_ctr_button.setIcon(QIcon(line_chart_icon_path))  # 設置按鈕圖標
        gen_ctr_button.setIconSize(QSize(30, 30))  # 設置圖標大小
        gen_ctr_button.clicked.connect(lambda: self.create_chart(generate_chart=True))
        expo_button = QPushButton("")
        expo_button.setToolTip("Export to Excel")  # 設置工具提示
        excel_icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_icon.png")
        expo_button.setIcon(QIcon(excel_icon_path))  # 設置按鈕圖標
        expo_button.setIconSize(QSize(30, 30))  # 設置圖標大小
        expo_button.clicked.connect(lambda: self.create_chart(generate_chart=False))

        msc_btn_h_layout = QHBoxLayout()
        msc_btn_h_layout.addWidget(gen_ctr_button)
        msc_btn_h_layout.addWidget(expo_button)
        msc_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)

        # 添加 QComboBox 來選擇移動倍率
        self.shift_combo = QComboBox()
        self.shift_combo.addItems(["10%", "30%", "50%"])
        self.shift_combo.setCurrentIndex(0)  # 預設選擇 10%

        # 添加zoom按鈕
        zoom_in_button = QPushButton("-")
        zoom_in_button.clicked.connect(lambda: self.zoom_chart(zoom_in=False))
        zoom_out_button = QPushButton("+")
        zoom_out_button.clicked.connect(lambda: self.zoom_chart(zoom_in=True))

        # 添加平移按鈕
        left_shift_button = QPushButton("<--")
        left_shift_button.clicked.connect(lambda: self.shift_chart(left_shift=True))
        right_shift_button = QPushButton("-->")
        right_shift_button.clicked.connect(lambda: self.shift_chart(left_shift=False))

        # 將 zoom/shift 按鈕添加到佈局中
        msc_zoom_btn_h_layout = QHBoxLayout()
        msc_zoom_btn_h_layout.addWidget(zoom_in_button)
        msc_zoom_btn_h_layout.addWidget(zoom_out_button)
        msc_zoom_btn_h_layout.addWidget(left_shift_button)
        msc_zoom_btn_h_layout.addWidget(right_shift_button)
        msc_zoom_btn_h_layout.addWidget(self.shift_combo)
        msc_zoom_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignRight)

        msc_v_layout.addLayout(msc_zoom_btn_h_layout)
        msc_v_layout.addLayout(msc_btn_h_layout)        
        
    def create_menu_bar(self):
        menu_bar = self.menuBar()

        # File menu
        file_menu = menu_bar.addMenu("File")
        open_action = file_menu.addAction("Open Folder")
        open_action.triggered.connect(self.select_folder)

        # 新增 Log menu
        log_menu = menu_bar.addMenu("Log")
        self.log_actions = {
            "LogService": log_menu.addAction("LogService"),
            "Install": log_menu.addAction("Install"),
            "mjnxtdebug": log_menu.addAction("mjnxtdebug"),
            "Protocol": log_menu.addAction("Protocol")
        }

        for action in self.log_actions.values():
            action.setCheckable(True)  # 設置為可勾選
            action.setChecked(True)  # 預設為勾選

        # Info menu
        info_menu = menu_bar.addMenu("Info")
        about_action = info_menu.addAction("About")
        about_action.triggered.connect(self.show_about_dialog)

    def update_processing_label(self):
        self.dot_count = (self.dot_count + 1) % 4  # 使點數在 0 到 3 之間循環
        dots = '.' * self.dot_count  # 根據計數生成點數
        self.display_status(f"Processing {dots}", "ongoing")
    
    def show_about_dialog(self):
        QMessageBox.about(self, "About", "Log Analyser v1.0.0\nAuthor : Davian Kuo\nE-mail : davian.kuo@zeiss.com")

    def display_status(self, msg, type):
        self.status_label.setText(msg)
        if type == "done":
            self.status_label.setStyleSheet("background-color: lightgreen;")
        elif type == "fail":
            self.status_label.setStyleSheet("background-color: lightpink;")
        elif type == "ongoing":
            self.status_label.setStyleSheet("background-color: lightyellow;")

    def select_folder_from_tree(self):
        # 獲取當前選中的索引
        selected_index = self.tree_view.currentIndex()
        if selected_index.isValid():  # 確保選中的索引有效
            # 獲取選中的資料夾路徑
            folder_path = self.file_system_model.filePath(selected_index)
            # 將路徑填入 folder_path_edit
            self.folder_path_edit.setText(folder_path)
            self.display_status("Folder selected !", "done")
        else:
            self.display_status("Please select folder !", "fail")
            QMessageBox.information(self, "Log Analyser", "Please select folder !", QMessageBox.StandardButton.Ok)

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.folder_path_edit.setText(folder)
            self.display_status("Folder selected !", "done")

    def set_time_filter_ctrl(self, en):
        self.start_date_edit.setEnabled(en)
        self.start_hour_spinbox.setEnabled(en)
        self.start_minute_spinbox.setEnabled(en)
        self.start_second_spinbox.setEnabled(en)

        self.end_date_edit.setEnabled(en)
        self.end_hour_spinbox.setEnabled(en)
        self.end_minute_spinbox.setEnabled(en)
        self.end_second_spinbox.setEnabled(en)

    def on_chart_source_selected(self):
        if self.multijet_chart_radio.isChecked():
            self.selected_chart = "mjnxtdebug"
            self.set_time_filter_ctrl(True)
        elif self.protocol_chart_radio.isChecked():
            self.selected_chart = "protocol"
            self.set_time_filter_ctrl(False)

        for gauge_id in self.gauge_types:
            self.gauge_checkboxes[gauge_id].setEnabled(True)
            self.gauge_checkboxes[gauge_id].setChecked(True)
    
    def open_protocol_reader(self, start_date, end_date, folder_path):
        if start_date == None or end_date == None or folder_path == None:
            self.display_status("Protocol date is None !", "fail")
            QMessageBox.information(self, "Log Analyser", "Protocol date is None !", QMessageBox.StandardButton.Ok)
            return

        # 獲取當前腳本的路徑
        excel_file_path = os.path.join(folder_path, "Protocol reader_V5.xlsm")  # 替換為你的檔案名

        # 檢查檔案是否存在
        if os.path.exists(excel_file_path):
            try:
                # 啟動 Excel 應用程式
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True

                workbook = excel.Workbooks.Open(excel_file_path)

                excel.Application.Run("SetInformationInputValues", start_date, end_date, folder_path)
                excel.Application.Run("Log_analysis")

                # 保存並關閉工作簿
                #workbook.Save()
                #workbook.Close()
                #excel.Quit()
            except Exception as e:
                print(f"Failed to fill Excel file: {str(e)}")
        else:
            print("The specified Excel file does not exist !")
            self.display_status("The specified Excel file does not exist !", "fail")
    
    def find_log_files(self):
        found_path = self.folder_path_edit.text()
        if not found_path:
            self.display_status("Please choose folder before parsing !", "fail")
            QMessageBox.information(self, "Log Analyser", "Please choose folder before parsing !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None

        logsrvc_log = []
        install_log = []
        msc_log = []
        protocol_log = []
        process_log = []

        # 遞迴搜尋所有子資料夾
        for root, dirs, files in os.walk(found_path):
            for file in files:
                if self.log_actions["LogService"].isChecked():
                    if file == "LogService.txt" or re.match(r"LogService\d{4}-\d{2}-\d{2}_\d{6}\.txt", file):
                        logsrvc_log.append(os.path.join(root, file))

                if self.log_actions["Install"].isChecked():
                    if file == "Install.txt":
                        install_log.append(os.path.join(root, file))

                if self.log_actions["mjnxtdebug"].isChecked():
                    if re.match(r"^mjnxtdebug\d{8}\.log$", file):
                        msc_log.append(os.path.join(root, file))

                if self.log_actions["Protocol"].isChecked():
                    if re.match(r"Protocol_\d{4}-\d{2}-\d{2}_\d{4}_\d{4}\.txt", file):
                        protocol_log.append(os.path.join(root, file))
                    if re.match(r"ProcessLog_\d{4}-\d{2}-\d{2}_\d{4}_\d{4}\.csv", file):
                        process_log.append(os.path.join(root, file))

        # 檢查是否找到有效的日誌文件
        if self.log_actions["LogService"].isChecked() and not logsrvc_log:
            self.display_status("❌ Can't find valid LogService !", "fail")
            QMessageBox.information(self, "Log Analyser", "❌ Can't find valid LogService !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None
        if self.log_actions["Install"].isChecked() and not install_log:
            self.display_status("❌ Can't find valid Install !", "fail")
            QMessageBox.information(self, "Log Analyser", "❌ Can't find valid Install !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None
        if self.log_actions["mjnxtdebug"].isChecked() and not msc_log:
            self.display_status("❌ Can't find valid mjnxtdebug !", "fail")
            QMessageBox.information(self, "Log Analyser", "❌ Can't find valid mjnxtdebug !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None    
        if self.log_actions["Protocol"].isChecked():
            if not protocol_log:
                self.display_status("❌ Can't find valid protocol file !", "fail")
                QMessageBox.information(self, "Log Analyser", "❌ Can't find valid protocol file !", QMessageBox.StandardButton.Ok)
                return None, None, None, None, None, None
            if not process_log:
                self.display_status("❌ Can't find valid ProcessLog file !", "fail")
                QMessageBox.information(self, "Log Analyser", "❌ Can't find valid ProcessLog file !", QMessageBox.StandardButton.Ok)
                return None, None, None, None, None, None

        # 進度條初始化
        total_files = len(logsrvc_log) + len(install_log) + len(msc_log) + len(protocol_log) + len(process_log)
        self.progress_bar.setRange(0, total_files)
        self.progress_bar.setValue(0)

        return found_path, logsrvc_log, install_log, msc_log, protocol_log, process_log

    def process_raw_log(self):
        self.mask_events, self.workflow_data = {}, {}
        self.log_data, time_range, self.valve_log_data = {}, [], {}
        self.protocol_data = {}

        if not any(action.isChecked() for action in self.log_actions.values()):
            self.display_status("Please select at least one log to parse !", "fail")
            QMessageBox.information(self, "Log Analyser", "Please select at least one log to parse !", QMessageBox.StandardButton.Ok)
            return

        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        
        self.prot_folder_path = os.path.normpath(folder_path)
        self.process_button.setEnabled(False)
        self.process_button.setStyleSheet("background-color: lightgray; color: darkgray;")
        self.processed_files = 0
        self.display_status("Processing .", "ongoing")

        self.timer.start(100)

        # 1. LogService.txt 
        if logsrvc_file:
            parsed_result = self.parse_logsvr(folder_path, logsrvc_file)
            if parsed_result is None:
                self.timer.stop()
                self.display_status("LogService*.txt parse failed !", "fail")
                QMessageBox.information(self, "Log Analyser", "LogService*.txt parse failed !", QMessageBox.StandardButton.Ok)
                return
            self.mask_events, self.workflow_data = parsed_result

            # Show Mask ID Load/Unload Info
            self.display_load_unload_info()
        else:
            self.mask_events, self.workflow_data = {}, {}
            self.lwt_mask_info.clear()
            self.lwt_wk_flow.clear()

        # 2. Install.txt
        if install_log:
            self.parse_install(folder_path)
        else:
            self.install_info_list.clear()

        # 3. mjnxtdebug.log
        if msc_file:
            self.log_data, time_range, self.valve_log_data = self.parse_msc(folder_path, msc_file)
            if not self.log_data or not self.valve_log_data:
                self.timer.stop()
                self.display_status("mjnxtdebug*.log parse failed !", "fail")
                QMessageBox.information(self, "Log Analyser", "mjnxtdebug*.log parse failed !", QMessageBox.StandardButton.Ok)
                return

            # 獲取時間範圍
            self.min_time = min(time_range)
            self.max_time = max(time_range)
            self.start_date_edit.setMinimumDate(self.min_time.date())
            self.start_date_edit.setMaximumDate(self.max_time.date())
            self.end_date_edit.setMinimumDate(self.min_time.date())
            self.end_date_edit.setMaximumDate(self.max_time.date())
        else:
            self.log_data, time_range, self.valve_log_data = {}, [], {}
            self.figure_ctr.clear()

        # 4. Protocol.txt
        if protocol_file and process_log:
            self.prot_start_date, self.prot_end_date = self.parse_protocol(folder_path, protocol_file, process_log)
        else:
            self.protocol_data = {}
 
        self.timer.stop()
        self.display_status("Process done !", "done")
        QMessageBox.information(self, "Log Analyser", "Process done !", QMessageBox.StandardButton.Ok)
        self.process_button.setEnabled(True)
        self.process_button.setStyleSheet("""
            QPushButton {
                font-weight: bold;
                background-color: white;
                color: black; 
                border: 2px solid #000;
            }
            QPushButton:hover {
                font-weight: bold;
                background-color: lightblue;
                color: black;
            }
        """)

    def parse_logsvr(self, folder_path, logsrvc_file):
        logsrvc_file = sorted(
            logsrvc_file,
            key=lambda f: os.path.getmtime(os.path.join(folder_path, f)),
            reverse=True
        )

        # 儲存數據
        self.mask_events = defaultdict(list)
        self.workflow_data = defaultdict(dict)
        workflow_id = 1

        mask_event_pattern = r"Mask with Id: (\w+\.\d+) (loaded|unloaded)"
        workflow_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into '([^']+)'"
        end_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'Setup.ProtocolDefinition'"
        loaded_start_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'MaskLoading.LoadingMask'"
        unloaded_start_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'MaskUnloading.UnloadingMask'"

        loaded_start_time = None
        loaded_end_time = None
        unloaded_start_time = None
        unloaded_end_time = None

        current_workflow = []
        mask_id = None

        for i, file in enumerate(logsrvc_file):
            file_path = os.path.join(folder_path, file)
            with open(file_path, 'r', encoding='iso-8859-1') as f:
                lines = f.readlines()

            for line in reversed(lines):  # 倒序處理
                # 提取 Mask ID 事件
                mask_match = re.search(mask_event_pattern, line)
                if mask_match:
                    mask_id = mask_match.group(1)
                    event_type = mask_match.group(2)
                    timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    #self.mask_events[mask_id].append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, "N/A"))

                    if event_type == "loaded":
                        loaded_end_time = timestamp
                    elif event_type == "unloaded":
                        unloaded_end_time = timestamp

                # 提取 workflow_pattern
                workflow_match = re.search(workflow_pattern, line)
                if workflow_match and mask_id:
                    timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    current_workflow.append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], workflow_match.group(1)))

                    if re.search(loaded_start_pattern, line):
                        loaded_start_time = timestamp
                        if loaded_start_time and loaded_end_time and event_type == "loaded":
                            cost_time = str((loaded_end_time - loaded_start_time).total_seconds())
                        else:
                            cost_time = "N/A"
                        self.mask_events[mask_id].append((loaded_end_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, cost_time))
                    elif re.search(unloaded_start_pattern, line):
                        unloaded_start_time = timestamp
                        if unloaded_start_time and unloaded_end_time and event_type == "unloaded":
                            cost_time = str((unloaded_end_time - unloaded_start_time).total_seconds())
                        else:
                            cost_time = "N/A"
                        self.mask_events[mask_id].append((unloaded_end_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, cost_time))

                # 偵測到結束條件
                if re.search(end_pattern, line) and mask_id:
                    #timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    #timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    #current_workflow.append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], line.strip()))

                    # 確保 mask_id 的列表已初始化
                    if mask_id not in self.workflow_data[workflow_id]:
                        self.workflow_data[workflow_id][mask_id] = []
                    self.workflow_data[workflow_id][mask_id].extend(reversed(current_workflow))
                    
                    workflow_id += 1
                    current_workflow = []
                    mask_id = None
                    loaded_start_time = loaded_end_time = None
                    unloaded_start_time = unloaded_end_time = None

            # 更新進度條
            self.progress_bar.setValue(i + 1)
            self.processed_files += 1
            QApplication.processEvents()

        return self.mask_events, self.workflow_data
    
    def display_load_unload_info(self):
        # 清空 Mask ID 列表
        self.lwt_mask_info.clear()

        self.duration_dict = defaultdict(list)

        # 收集所有事件並排序
        all_events = [(mask_id, timestamp, event_type, cost_time)
                    for mask_id, events in self.mask_events.items()
                    for timestamp, event_type, cost_time in events]

        sorted_events = sorted(all_events, key=lambda x: x[1])  # 按時間排序

        loaded_timestamp = None  # 用於計算載入與卸載之間的時間

        for mask_id, timestamp, event_type, cost_time in sorted_events:
            self.lwt_mask_info.addItem(f"{mask_id} | {timestamp} | {event_type} | => cost {cost_time} secs")

            timestamp_dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S.%f")

            if event_type == "loaded":
                loaded_timestamp = timestamp_dt

            if event_type == "unloaded":
                if loaded_timestamp is None:
                    self.lwt_mask_info.addItem(f"duration : N/A")
                else:
                    duration_seconds = (timestamp_dt - loaded_timestamp).total_seconds()
                    hours, remainder = divmod(duration_seconds, 3600)  # 計算小時和剩餘秒數
                    minutes, seconds = divmod(remainder, 60)  # 計算分鐘和秒數
                    # hh:mm:ss 格式
                    duration_str = f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
                    self.lwt_mask_info.addItem(f"duration : {duration_str}")

                    # 將持續時間記錄到 duration_dict 中
                    self.duration_dict[mask_id].append(duration_seconds)

                    loaded_timestamp = None  # 重置載入時間
                self.lwt_mask_info.addItem(f"===========================================")

        if loaded_timestamp is not None:
            self.lwt_mask_info.addItem(f"duration : N/A")
        
    def display_workflow(self, item):
        selected_text = item.text()
        split_data = selected_text.split(" | ")
        if len(split_data) >= 4:
            mask_id, timestamp_str, event_type, cost_time = split_data
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S.%f")

            mask_event_loaded = r"MaskLoading.MaskLoaded"

            # 遍歷 workflow_data，找到與 mask_id 相符的記錄
            self.lwt_wk_flow.clear()
            wf_print_out = False
            for workflow_id, data in self.workflow_data.items():
                if mask_id in data and not wf_print_out:  # 確保 mask_id 存在於該 workflow_id 的數據中
                    for wf_timestamp_str, wf_line in data[mask_id]:  # 遍歷該 mask_id 的工作流
                        wf_timestamp = datetime.strptime(wf_timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
                        if wf_timestamp <= timestamp:  # 範圍條件
                            # 顯示第一組符合條件的資料後退出
                            #self.lwt_wk_flow.addItem(f"{workflow_id} | {wf_timestamp_str} | {wf_line}")
                            self.lwt_wk_flow.addItem(f"{wf_timestamp_str} | {wf_line}")
                            evnt_match = re.search(mask_event_loaded, wf_line)
                            if evnt_match:
                                self.lwt_wk_flow.addItem(f"-----------------------------------------------------------")
                            wf_print_out = True
            self.lwt_wk_flow.addItem(f"=================================")

    def generate_load_unload_time_chart(self):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        if not self.mask_events or not self.workflow_data:
            self.display_status("No parse for LogService.txt !", "fail")
            QMessageBox.information(self, "Log Analyser", "No parse for LogService.txt !", QMessageBox.StandardButton.Ok)
            return

        # 收集所有 loaded 事件
        all_loaded_events = [
            (mask_id, timestamp, event_type, cost_time)
            for mask_id, events in self.mask_events.items()
            for timestamp, event_type, cost_time in events if event_type == "loaded"
        ]
        # 收集所有 unloaded 事件
        all_unloaded_events = [
            (mask_id, timestamp, event_type, cost_time)
            for mask_id, events in self.mask_events.items()
            for timestamp, event_type, cost_time in events if event_type == "unloaded"
        ]

        # 按時間排序
        sorted_loaded_events = sorted(all_loaded_events, key=lambda x: datetime.strptime(x[1], "%Y-%m-%d %H:%M:%S.%f"))
        sorted_unloaded_events = sorted(all_unloaded_events, key=lambda x: datetime.strptime(x[1], "%Y-%m-%d %H:%M:%S.%f"))

        # 準備數據以繪製圖表
        loaded_times = {}
        for mask_id, timestamp, event_type, cost_time in sorted_loaded_events:
            if mask_id not in loaded_times:
                loaded_times[mask_id] = []
            loaded_times[mask_id].append(float(cost_time))

        unloaded_times = {}
        for mask_id, timestamp, event_type, cost_time in sorted_unloaded_events:
            if mask_id not in unloaded_times:
                unloaded_times[mask_id] = []
            unloaded_times[mask_id].append(float(cost_time)) 

        self.plot_load_unload_chart(loaded_times, unloaded_times)

    def plot_load_unload_chart(self, loaded_data, unloaded_data):
        self.figure_mask_analysis.clear()

        # 使用 gridspec 設置子圖的高度比例
        gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1])  # 上子圖佔 2， 下子圖佔 1

        ax1 = self.figure_mask_analysis.add_subplot(gs[0])  # 上半部子圖
        ax2 = self.figure_mask_analysis.add_subplot(gs[1])  # 下半部子圖

        # 繪製載入時間圖
        load_times = {mask_id: [cost_time for cost_time in cost_times if cost_time > 0] for mask_id, cost_times in loaded_data.items()}
        self.plot_chart(load_times, "Load Times", "Mask ID", "Cost Time (minutes)", event_type="loaded", ax=ax1)

        # 繪製卸載時間圖
        unload_times = {mask_id: [cost_time for cost_time in cost_times if cost_time > 0] for mask_id, cost_times in unloaded_data.items()}
        self.plot_chart(unload_times, "Unload Times", "Mask ID", "Cost Time (minutes)", event_type="unloaded", ax=ax2)

    def generate_duration_time_chart(self):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        if not self.mask_events or not self.workflow_data:
            self.display_status("No parse for LogService.txt !", "fail")
            QMessageBox.information(self, "Log Analyser", "No parse for LogService.txt !", QMessageBox.StandardButton.Ok)
            return
        
        self.plot_duration_chart(self.duration_dict)
    
    def plot_chart(self, data, title, xlabel, ylabel, event_type, ax):
        # 準備數據以繪製直條圖
        mask_ids = []
        duration_values = []
        mask_id_count = {}  # 用於計數每個 Mask ID 的出現次數

        for mask_id, cost_times in data.items():
            if cost_times:  # 確保有資料才繪製
                for cost_time in cost_times:
                    # 檢查 Mask ID 是否已經存在，並進行計數
                    if mask_id in mask_id_count:
                        mask_id_count[mask_id] += 1
                        new_mask_id = f"{mask_id}_{mask_id_count[mask_id]}"
                    else:
                        mask_id_count[mask_id] = 1
                        new_mask_id = mask_id

                    mask_ids.append(new_mask_id)
                    duration_values.append(cost_time / 60)  # 將秒轉換為分鐘

        if not mask_ids or not duration_values:
            print("No data to plot.")
            return  # 如果沒有數據，則不繪製圖表

        bar_width = 0.4

        # 繪製直條圖
        bars = ax.bar(mask_ids, duration_values, width=bar_width, color='skyblue')

        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)  # 單位為分鐘
        
        ax.grid(axis='y')  # 只顯示 y 軸的網格線

        # 在每根直條上顯示數值
        for bar in bars:
            yval = bar.get_height()  # 獲取直條的高度（即數值）
            ax.text(bar.get_x() + bar.get_width() / 2, yval, f'{yval:.2f}', ha='center', va='bottom')  # 在直條上方顯示數值

        # 在 y 軸為 10 的地方畫一條水平線
        #ax.axhline(y=10, color='red', linestyle='--', label='Threshold Line at 10') 

        # 調整 x 軸標籤的顯示
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 調整 y 軸範圍（如果需要）
        ax.set_ylim(bottom=0)  # 確保 y 軸從 0 開始

        # 設置 y 軸範圍
        ax.set_ylim(0, 20)

        # 調整邊距
        self.figure_mask_analysis.tight_layout()
        self.canvas_mask_analysis.draw()

    def plot_duration_chart(self, data):
        self.figure_mask_analysis.clear()
        ax = self.figure_mask_analysis.add_subplot(111)

        mask_ids = []
        duration_values = []
        mask_id_count = {}  # 用於計數每個 Mask ID 的出現次數

        for mask_id, durations in data.items():
            if durations:  # 確保有資料才繪製
                for duration in durations:
                    # 檢查 Mask ID 是否已經存在，並進行計數
                    if mask_id in mask_id_count:
                        mask_id_count[mask_id] += 1
                        new_mask_id = f"{mask_id}_{mask_id_count[mask_id]}"
                    else:
                        mask_id_count[mask_id] = 1
                        new_mask_id = mask_id

                    mask_ids.append(new_mask_id)
                    duration_values.append(duration)

        # 將持續時間從秒轉換為小時
        duration_values_in_hours = [duration / 3600 for duration in duration_values]
        
        # 繪製直條圖
        bars = ax.bar(mask_ids, duration_values_in_hours, color='skyblue')

        ax.set_title("Duration Time")
        ax.set_xlabel("Mask ID")
        ax.set_ylabel("Duration Time (hours)")
        
        ax.grid(axis='y')  # 只顯示 y 軸的網格線

        # 在每根直條上顯示數值
        for bar in bars:
            yval = bar.get_height()  # 獲取直條的高度（即數值）
            ax.text(bar.get_x() + bar.get_width() / 2, yval, f'{yval:.2f}', ha='center', va='bottom')  # 在直條上方顯示數值

        # 調整 x 軸標籤的顯示
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 調整 y 軸範圍（如果需要）
        ax.set_ylim(bottom=0)  # 確保 y 軸從 0 開始

        # 調整邊距
        self.figure_mask_analysis.tight_layout()

        self.canvas_mask_analysis.draw()

    def parse_install(self, folder_path):
        install_file = os.path.join(folder_path, "Install.txt")
        if os.path.exists(install_file):            
            with open(install_file, 'r', encoding='iso-8859-1') as file:
                lines = file.readlines()
                for line in lines:
                    item = QListWidgetItem(line.strip())
                    self.install_info_list.addItem(item)

                # 將最後一行設置為黃色
                if self.install_info_list.count() > 0:  # 確保有項目存在
                    last_item = self.install_info_list.item(self.install_info_list.count() - 1)
                    last_item.setBackground(QColor("yellow"))  # 設置背景顏色為黃色

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()

    def parse_msc(self, folder_path, msc_file):
        gauge_ids = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3", "MFC9-5", "MFC10-5"]
        valve_ids = ["V0-3", "V0-4", "V0-5", "V0-7", "V0-10", "V0-13", "V0-16", "V0-31", "V0-32", "V0-33", "V0-34", "V0-35",
                     "V1-3", "V1-4", "V1-9", "V1-10", "V1-11", "V1-13", "V1-21", "V1-22",
                     "V2-3", "V2-4", "V2-9", "V2-11", "V2-13", "V2-21",
                     "V3-3", "V3-4", "V3-9", "V3-11", "V3-13", "V3-21",
                     "V4-3", "V4-4", "V4-5", "V4-9", "V4-11", "V4-13", "V4-21",
                     "V9-3", "V9-4", "V9-5", "V9-7", "V9-8", "V9-11", "V9-12", "V9-13", "V9-14", "V9-15", "V9-16", "V9-17", "V9-21",
                     "V10-3", "V10-5", "V10-7", "V10-8", "V10-11", "V10-12", "V10-13", "V10-14", "V10-15", "V10-16", "V10-17", "V10-21"]

        selected_format = self.selected_format
        if selected_format == "MSC 2.x":
            ctr_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(\d+\) PressEvTh\(\): Sent MULTIJET_EVENT_CODE_CURRENT_PRESSURE_HAS_CHANGED\((\d+), (\d+), press=(-?\d+\.\d+), array=(-?\d+\.\d+)\)")
            valve_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \((\d+)\) .*?: Calling SafeMediaCabinetSetValveDefaultState\((\d+),(\d+),(\d+)\)")
        elif selected_format == "MSC 3.x":
            ctr_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(.*?\) MultiJetImpl::MCPressCurrentValueChangedEvent\((\d+),(\d+)\), .*?pressure = (-?\d+\.\d+) mbar.*")
            mfc_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(\d+\) MCMFCCurrValueChangedEvent\(ch=(\d+), row=(\d+), value=([\d\.-]+), ValueSent=[\d\.-]+\)")
            valve_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(\d+\) .*?: Calling MediaCabinetSetValveDefaultState\((\d+),(\d+),(\d+)\)")
        else:
            self.display_status("Unsupported log format selected !", "fail")
            QMessageBox.information(self, "Log Analyser", "Unsupported log format selected !", QMessageBox.StandardButton.Ok)
            return None, None

        # 讀取 mjnxtdebugXXXXXXXX.log 文件以獲取初始時間
        initial_time = None

        for file_name in msc_file:
            file_path = os.path.join(folder_path, file_name)
            with open(file_path, 'r') as file:
                for line in file:
                    # 使用正則表達式提取時間字符串
                    time_pattern = re.search(r'(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3})', line)
                    if time_pattern:
                        date_time_str = time_pattern.group(1)  # 獲取匹配的時間字符串
                        initial_time = datetime.strptime(date_time_str, "%Y/%m/%d, %H:%M:%S.%f")
                        break  # 只需獲取第一行的時間
            break  # 只需處理第一個日誌文件

        valve_log_data = {valve_id: [(initial_time, 0)] for valve_id in valve_ids}  # 每個閥門的初始狀態為關閉
        
        time_stamps = []
        gauge_log_data = {gauge_id: [] for gauge_id in gauge_ids}

        for file_name in msc_file:
            file_path = os.path.join(folder_path, file_name)

            with open(file_path, 'r') as file:
                for line in file:
                    # CTR/MFC value
                    gauge_match = ctr_pattern.search(line) or mfc_pattern.search(line)
                    if gauge_match:
                        date_time_str, main_id, sub_id, press_value = gauge_match.groups()[:4]
                        parsed_time = datetime.strptime(date_time_str, "%Y/%m/%d, %H:%M:%S.%f")
                        time_stamps.append(parsed_time)
                        
                        if (main_id == '9' and sub_id == '5') or (main_id == '10' and sub_id == '5'):
                            gauge_id = f"MFC{main_id}-{sub_id}"
                        else:
                            gauge_id = f"P{main_id}-{sub_id}"
                        if gauge_id in gauge_log_data:
                            gauge_log_data[gauge_id].append((parsed_time, float(press_value)))

                    # valve status
                    valve_match = valve_pattern.search(line)
                    if valve_match:
                        date_time_str, main_id, sub_id, valve_status = valve_match.groups()[:4]
                        parsed_time = datetime.strptime(date_time_str, "%Y/%m/%d, %H:%M:%S.%f")

                        valve_id = f"V{main_id}-{sub_id}"
                        if valve_id in valve_log_data:
                            # 將新的狀態附加到閥門狀態列表中
                            valve_log_data[valve_id].append((parsed_time, int(valve_status)))

                    if self.processed_files % 100 == 0:  # 每 100 行更新一次
                        QApplication.processEvents()

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
        
        # 檢查是否有任何閥門狀態或量測儀數據被記錄
        if not any(valve_log_data.values()) and not any(gauge_log_data.values()):
            return None, None, None
        
        return gauge_log_data, sorted(set(time_stamps)), valve_log_data
    
    def parse_protocol(self, folder_path, protocol_file, process_log):
        self.mask_project_combobox.clear()
        self.protocol_combobox.clear()
        
        start_date = None
        end_date = None
        gauge_ids = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3"]
        

        added_projects = set()  # 用於跟踪已添加的項目

        self.protocol_data = {}
        process_data = {gauge_id: [] for gauge_id in gauge_ids}

        # 設定閥門的起始值
        valve_ids = {
            "P1-1": "V1-21|22",
            "P2-1": "V2-21",
            "P3-1": "V3-21",
            "P4-1": "V4-21",
            "P9-1": "V9-21",
            "P9-2": "V9-21",
            "P9-3": "V9-21",
            "P10-1": "V10-21",
            "P10-2": "V10-21",
            "P10-3": "V10-21",
        }

        # Parse ProcessLog files
        for file in process_log:
            date_match = re.search(r'ProcessLog_(\d{4})-(\d{2})-(\d{2})', file)
            if date_match:
                date_str = f"{date_match.group(1)}/{date_match.group(2)}/{date_match.group(3)}"  # yyyy/mm/dd
                if start_date is None:
                    start_date = date_str
                end_date = date_str
            with open(file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                process_filename = os.path.splitext(os.path.basename(file))[0]  # 去除擴展名

                # 從檔名中提取基準時間
                base_time_str = process_filename.split("_")[1:3]  # 提取檔名中的日期和時間部分
                base_time_str = f"{base_time_str[0]} {base_time_str[1]}"  # 合併為 "YYYY-MM-DD HHMM"
                base_time = datetime.strptime(base_time_str, "%Y-%m-%d %H%M")  # 轉換為 datetime 對象

                gauge_values = {gauge_id: [] for gauge_id in gauge_ids}  # 用於存儲每個量測儀的值
                valve_values = {valve_id: [] for valve_id in valve_ids.values()}  # 用於存儲閥門的值

                for line in lines:
                    line = line.strip()  # 去除行首和行尾的空白字符
                    if line.startswith("\"Elapsed Seconds\""):
                        continue  # 跳過標題行

                    parts = line.split(",")
                    if len(parts) < len(gauge_ids) + 2:  # 確保有足夠的欄位, 跳過不完整的行
                        continue

                    # 獲取 CSV 標題行中的量測儀欄位索引
                    header = [col.strip().strip('"') for col in lines[0].strip().split(",")]  # 去除引號
                    gauge_indices = {gauge_id: header.index(gauge_id) for gauge_id in gauge_ids if gauge_id in header}
                    valve_indices = {valve_id: header.index(valve_id) for valve_id in valve_ids.values() if valve_id in header}

                    # 依序提取量測儀的值
                    for gauge_id, index in gauge_indices.items():
                        value = float(parts[index].strip())  # 根據索引獲取對應的值
                        gauge_values[gauge_id].append(value)  # 將值存入對應的 gauge_id 列表

                    # 依序提取閥門的值
                    for valve_id, index in valve_indices.items():
                        valve_value = int(parts[index].strip())  # 根據索引獲取對應的閥門值
                        valve_values[valve_id].append(valve_value)  # 將值存入對應的閥門列表

                # 計算每個量測儀的平均值並存入 process_data
                for gauge_id in gauge_ids:
                    if gauge_values[gauge_id]:  # 確保有值
                        # 獲取對應閥門的值
                        valve_id = valve_ids[gauge_id]
                        start_calculating = False
                        total_value = 0
                        count = 0

                        for i in range(len(valve_values[valve_id])):
                            if valve_values[valve_id][i] == 1:
                                start_calculating = True  # 開始計算
                            if start_calculating:
                                total_value += gauge_values[gauge_id][i]
                                count += 1

                        if count > 0:
                            average_value = total_value / count  # 計算平均值
                            average_value = round(average_value, 6)
                            process_data[gauge_id].append((base_time.strftime("%Y-%m-%d %H%M"), average_value))  # 存入時間戳和平均值

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
            
        # Parse Protocol.txt file
        for file in protocol_file:
            with open(file, 'r', encoding='utf-8') as f:
                lines = f.readlines()

                protocol_filename = os.path.splitext(os.path.basename(file))[0] # 去除擴展名
                self.protocol_combobox.addItem(protocol_filename)

                process_names = []
                u_value = None
                v_value = None
                current_recipe = None
                mask_project_name = None

                for line in lines:
                    line = line.strip()  # 去除行首和行尾的空白字符

                    if line.startswith("[Recipe_"):
                        current_recipe = line.split("[")[1].split("]")[0]  # 獲取 Recipe_#
                        continue

                    if current_recipe:
                        if "ApplicationModule=" in line:
                            am_path = line.split("ApplicationModule=")[1].strip()
                            process_name = os.path.basename(am_path)  # 獲取檔名，例如 Opaque.am
                            process_names.append(f"{current_recipe} => {process_name}")
                            current_recipe = None

                    if "U=" in line:  # 檢查行中是否包含 U=
                        u_match = re.search(r"U\s*=\s*([\d\.]+)", line)  # 匹配 U 的值
                        if u_match:
                            u_value = u_match.group(1)  # 更新 U 值

                    if "V=" in line:  # 檢查行中是否包含 V=
                        v_match = re.search(r"V\s*=\s*([\d\.]+)", line)  # 匹配 V 的值
                        if v_match:
                            v_value = v_match.group(1)  # 更新 V 值
                
                    # 提取 PreRepairImage 路徑中的項目
                    if "PreRepairImage=" in line:
                        image_path = line.split("PreRepairImage=")[1].strip()
                        project_match = re.search(r'\\([^\\]+)\\[^\\]+$', image_path)  # 匹配最後一個目錄名稱
                        if project_match:
                            mask_project_name = project_match.group(1)  # 提取目錄名稱
                            if mask_project_name not in added_projects:
                                self.mask_project_combobox.addItem(mask_project_name)  # 添加到 mask_project_combobox
                                added_projects.add(mask_project_name)  # 將項目添加到集合中
                
                # 從 protocol_filename 提取時間戳
                protocol_base_time_split = protocol_filename.split("_")[1:3]  # 提取檔名中的日期和時間部分
                protocol_base_time_str = f"{protocol_base_time_split[0]} {protocol_base_time_split[1]}"  # 合併為 "YYYY-MM-DD HHMM"

                # 將資料存入字典
                if mask_project_name:
                    if mask_project_name not in self.protocol_data:
                        self.protocol_data[mask_project_name] = {}
                    self.protocol_data[mask_project_name][protocol_filename] = {
                        "process": process_names,
                        "U": u_value,
                        "V": v_value,
                        "process_data": {}
                    }

                    # 只將與 protocol_filename 時間匹配的 gauge 平均值存入
                    for gauge_id in gauge_ids:
                        if gauge_id in process_data and process_data[gauge_id]:
                            for timestamp, value in process_data[gauge_id]:
                                if timestamp == protocol_base_time_str:
                                    self.protocol_data[mask_project_name][protocol_filename]["process_data"][gauge_id] = [(timestamp, value)]

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
            
        self.mask_project_combobox.currentIndexChanged.connect(self.update_protocol_combobox)
        self.protocol_combobox.currentIndexChanged.connect(self.update_protocol_info)

        return start_date, end_date

    def on_tab_changed(self, index):
        if index == 3:  # MSC Info 是第四個 Tab，索引為 3
            self.update_protocol_info()

    def update_protocol_combobox(self):
        selected_project = self.mask_project_combobox.currentText()  # 獲取選中的 mask_project_name
        self.protocol_combobox.clear()  # 清空 protocol_combobox

        if selected_project in self.protocol_data:
            for protocol_filename in self.protocol_data[selected_project]:
                self.protocol_combobox.addItem(protocol_filename)  # 添加相關的 protocol_filename

    def update_protocol_info(self):
        selected_project = self.mask_project_combobox.currentText()  # 獲取選中的 mask_project_name
        selected_protocol = self.protocol_combobox.currentText()  # 獲取選中的 Protocol 文件名

        if selected_project in self.protocol_data and selected_protocol in self.protocol_data[selected_project]:
            process_names = self.protocol_data[selected_project][selected_protocol]["process"]
            u_value = self.protocol_data[selected_project][selected_protocol]["U"]
            v_value = self.protocol_data[selected_project][selected_protocol]["V"]

            # 更新 Label
            process_display = "\n".join(process_names) if process_names else "N/A"
            self.process_label.setText(f"Processes:\n{process_display}")
            self.u_label.setText(f"\tU: {u_value if u_value is not None else 'N/A'}")
            self.v_label.setText(f"\tV: {v_value if v_value is not None else 'N/A'}")
        else:
            # 如果沒有找到對應的資料，顯示 N/A
            self.process_label.setText("\tProcess: N/A")
            self.u_label.setText("\tU: N/A")
            self.v_label.setText("\tV: N/A")

    def on_format_selected(self):
        if self.msc_v2_radio.isChecked():
            self.selected_format = "MSC 2.x"
        elif self.msc_v3_radio.isChecked():
            self.selected_format = "MSC 3.x"

    def on_gauge_checkbox_changed(self):
        # 獲取當前觸發事件的複選框
        #checkbox = self.sender()
        
        # 獲取複選框的名稱或 ID
        #gauge_id = None
        #for gauge, cb in self.gauge_checkboxes.items():
        #    if cb == checkbox:
        #        gauge_id = gauge
        #        break

        # 檢查複選框的狀態
        #if checkbox.isChecked():
            #print(f"{gauge_id} 被選中")  # 當前複選框被選中
        #    self.last_unchecked_gauge = None  # 如果選中，重置最後一個未選中的 gauge
        #else:
            #print(f"{gauge_id} 被取消選擇")  # 當前複選框被取消選擇
        #    self.last_unchecked_gauge = gauge_id  # 更新最後一個被取消選擇的 gauge

        #self.create_chart(generate_chart=True)

        pass
    
    def create_chart(self, generate_chart=False):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        
        if self.selected_chart == "mjnxtdebug" and not self.log_data:
            self.display_status("mjnxtdebug*.log parse failed !", "fail")
            QMessageBox.information(self, "Log Analyser", "mjnxtdebug*.log parse failed !", QMessageBox.StandardButton.Ok)
            return
        elif self.selected_chart == "protocol" and not self.protocol_data:
            self.display_status("Protocol.txt or ProcessLog.csv parse failed !", "fail")
            QMessageBox.information(self, "Log Analyser", "Protocol.txt or ProcessLog.csv parse failed !", QMessageBox.StandardButton.Ok)
            return
            
        # 選擇開始和結束時間
        start_date = self.start_date_edit.text()
        start_time = f"{self.start_hour_spinbox.value():02}:{self.start_minute_spinbox.value():02}:{self.start_second_spinbox.value():02}"
        end_date = self.end_date_edit.text()
        end_time = f"{self.end_hour_spinbox.value():02}:{self.end_minute_spinbox.value():02}:{self.end_second_spinbox.value():02}"
        
        # 轉換為 Python datetime
        start_datetime = datetime.strptime(f"{start_date} {start_time}", "%Y/%m/%d %H:%M:%S")
        end_datetime = datetime.strptime(f"{end_date} {end_time}", "%Y/%m/%d %H:%M:%S")
        
        # 過濾數據
        filtered_data = self.filter_data(start_datetime, end_datetime)

        if self.selected_chart == "mjnxtdebug" and all(not values for values in filtered_data.values()):
            self.display_status("No data in selected range !", "fail")
            QMessageBox.information(self, "Log Analyser", "No CTR or Valve data in selected range !\n Please set the time filter.", QMessageBox.StandardButton.Ok)
            
            for gauge_id in self.gauge_types:
                self.gauge_checkboxes[gauge_id].setEnabled(True)
                self.gauge_checkboxes[gauge_id].setChecked(True)
            
            return
        
        # 獲取選擇的量測儀類型
        selected_gauges = [gauge for gauge, checkbox in self.gauge_checkboxes.items() if checkbox.isChecked()]
        
        # 根據選擇的量測儀過濾數據
        if selected_gauges:
            filtered_data = {gauge: filtered_data.get(gauge, []) for gauge in selected_gauges}
        else:
            self.display_status("No gauge selected !", "fail")
            QMessageBox.information(self, "Log Analyser", "No gauge selected !", QMessageBox.StandardButton.Ok)

            # 恢復最後一個被取消選擇的 gauge
            #if self.last_unchecked_gauge:
            #    self.gauge_checkboxes[self.last_unchecked_gauge].setChecked(True)  # 恢復該複選框為選中狀態
                #print(f"恢復選中: {self.last_unchecked_gauge}")  # 顯示恢復的 gauge

            return
        
        if generate_chart:
            self.plot_ctr_chart(filtered_data)
        else:
            self.save_to_excel(filtered_data)
        
    def filter_data(self, start_datetime, end_datetime):
        filtered_data = {}
        gauges_to_disable = []

        if self.selected_chart == "mjnxtdebug":
            #gauges_to_disable = []
            for gauge_id, data in self.log_data.items():
                #filtered_data[gauge_id] = [(dt, val) for dt, val in data if start_datetime <= dt <= end_datetime]
                filtered_values = [(dt, val) for dt, val in data if start_datetime <= dt <= end_datetime]
                filtered_data[gauge_id] = filtered_values

                if not filtered_values:
                    gauges_to_disable.append(gauge_id)
            
        elif self.selected_chart == "protocol":
            # 獲取當前選擇的 project file
            current_project = self.mask_project_combobox.currentText()

            # 整理所有 protocol file 的 gauge 平均值
            for _, protocol_info in self.protocol_data[current_project].items():
                process_data = protocol_info["process_data"]
                for gauge_id, gauge_values in process_data.items():
                    for timestamp, value in gauge_values:
                        # 將時間戳轉換為 datetime 對象
                        dt = datetime.strptime(timestamp, "%Y-%m-%d %H%M")
  
                        if gauge_id not in filtered_data:
                            filtered_data[gauge_id] = []
                        filtered_data[gauge_id].append((dt, value))

            # 檢查 filtered_data 中的 gauge_id 是否存在，若不存在則添加到 gauges_to_disable
            for dis_gauge in self.gauge_checkboxes.keys():
                if dis_gauge not in filtered_data or not filtered_data[gauge_id]:
                    gauges_to_disable.append(dis_gauge)

        # 禁用不符合條件的複選框並設置為未選中
        for gauge_id in self.gauge_checkboxes.keys():
            if gauge_id in gauges_to_disable:
                self.gauge_checkboxes[gauge_id].setEnabled(False)
                self.gauge_checkboxes[gauge_id].setChecked(False)
            else:
                self.gauge_checkboxes[gauge_id].setEnabled(True)
                #self.gauge_checkboxes[gauge_id].setChecked(True)

        return filtered_data
    
    def save_to_excel(self, filtered_data):
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel files (*.xlsx)")
        # 檢查文件是否存在且是否可寫入
        if os.path.exists(save_path):
            try:
                # 嘗試以寫入模式打開文件
                with open(save_path, 'a'):
                    pass  # 如果成功，則文件未被佔用
            except PermissionError:
                # 如果出現 PermissionError，則顯示警告
                QMessageBox.warning(self, "Warning", f"The file '{save_path}' is currently open or cannot be accessed. Please close it and try again.", QMessageBox.StandardButton.Ok)
                self.display_status("Excel save error !", "fail")
                return

        # 建立新的 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "All Gauge Data"  # 設定工作表名稱

        # 設定表頭
        ws.append(["Date Time"] + [f"{gauge_id} Press Value" for gauge_id in filtered_data.keys()])  # 依序為每個量測儀添加壓力數據的標題

        # 收集所有時間點
        all_times = sorted(set(time for data in filtered_data.values() for time, _ in data))

        # 填充數據
        for row_idx, time in enumerate(all_times, start=2):
            # 將時間格式化為字符串，包含毫秒
            formatted_time = time.strftime("%Y/%m/%d, %H:%M:%S.%f")[:-3]  # 只保留到毫秒
            ws.cell(row=row_idx, column=1, value=formatted_time)  # 寫入時間
            for col_idx, gauge_id in enumerate(filtered_data.keys(), start=2):
                # 嘗試找到該時間點對應的壓力值
                press_value = next((val for dt, val in filtered_data[gauge_id] if dt == time), None)
                if press_value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=press_value)
                else:
                    # 若無對應壓力值，計算前後壓力值的平均值
                    earlier_values = [val for dt, val in filtered_data[gauge_id] if dt < time]
                    later_values = [val for dt, val in filtered_data[gauge_id] if dt > time]

                    # 獲取最近的前後值
                    earlier_value = earlier_values[-1] if earlier_values else None
                    later_value = later_values[0] if later_values else None

                    # 計算平均值並填入
                    if earlier_value is not None and later_value is not None:
                        average_value = (earlier_value + later_value) / 2
                        ws.cell(row=row_idx, column=col_idx, value=average_value)
                    elif earlier_value is not None:  # 若只有前值
                        ws.cell(row=row_idx, column=col_idx, value=earlier_value)
                    elif later_value is not None:  # 若只有後值
                        ws.cell(row=row_idx, column=col_idx, value=later_value)

        # 建立趨勢圖
        chart = LineChart()
        chart.title = "CTR Pressure Trends"
        chart.x_axis.title = "Date Time"
        chart.y_axis.title = "Press Value (mbar)"

        # 設定數據範圍，從第二列開始，以包含所有量測儀的數據
        data_ref = Reference(ws, min_col=2, min_row=1, max_col=1 + len(filtered_data), max_row=len(all_times) + 1)
        chart.add_data(data_ref, titles_from_data=True)  # 包含標題

        # 設定 X 軸標籤（時間範圍）
        time_ref = Reference(ws, min_col=1, min_row=2, max_row=len(all_times) + 1)
        chart.set_categories(time_ref)

        # 設定 X 軸標籤格式
        chart.x_axis.number_format = "yyyy/mm/dd hh:mm:ss.000"
        chart.x_axis.majorTimeUnit = "days"
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickLblPos = "low"

        # 添加趨勢圖到工作表
        ws.add_chart(chart, "N2")  # 設定圖表顯示位置

        # 儲存 Excel 檔案
        wb.save(save_path)

        QMessageBox.information(self, "Log Analyser", "Data is exported to Excel.", QMessageBox.StandardButton.Ok)
        self.display_status("Data is exported to Excel.", "done")

    def plot_ctr_chart(self, filtered_data):
        self.ctr_char_gen = False
        # 清空畫布
        self.figure_ctr.clear()
        ax = self.figure_ctr.add_subplot(111)
        #ax2 = ax.twinx()

        # 一開始將所有 gauge 設成 N/A
        for gauge_id in self.gauge_types:
            getattr(self, f'vline_{gauge_id.replace("-", "_")}_label').setText(f'<b>{gauge_id}:</b> N/A ')

        # 準備數據
        self.all_times = sorted(set(time.replace(tzinfo=None) for data in filtered_data.values() for time, _ in data))  # 將所有時間轉換為 offset-naive
        #print("All Times:", all_times)  # 調試輸出
        # 列印最小值和最大值
        #if self.all_times:  # 確保 all_times 不為空
        #    print("X 軸範圍: 最小值 =", min(self.all_times), ", 最大值 =", max(self.all_times))  # 列印 X 軸範圍
        #else:
        #    print("all_times 為空，無法計算最小值和最大值。")
        
        press_values = {gauge_id: [] for gauge_id in filtered_data.keys()}

        for gauge_id, data in filtered_data.items():
            for dt, val in data:
                press_values[gauge_id].append((dt.replace(tzinfo=None), val))  # 將 dt 轉換為 offset-naive

        # 繪製數據
        for gauge_id, values in press_values.items():
            if values:
                times, vals = zip(*values)
                if self.selected_chart == "protocol":
                    ax.plot(times, vals, label=gauge_id, marker='o')  # 使用 marker 繪製數據點
                    # 在每個數據點上顯示數值和時間
                    for time, val in zip(times, vals):
                        ax.text(time, val, f'{val:.6f}\n{time.strftime("%H:%M:%S.%f")}', fontsize=8, ha='center', va='bottom', color='red')
                elif self.selected_chart == "mjnxtdebug":
                    ax.plot(times, vals, label=gauge_id)
                    #if gauge_id in ["MFC9-5", "MFC10-5"]:
                    #    ax2.plot(times, vals, label=gauge_id, linestyle='--')  # 使用虛線繪製 SCCM 數據
    
        ax.set_title("MultiJet Trend Chart")
        ax.set_xlabel("Date Time")
        if (self.gauge_checkboxes["MFC9-5"].isChecked() and not self.gauge_checkboxes["MFC10-5"].isChecked() and
            all(not self.gauge_checkboxes[gauge].isChecked() for gauge in self.gauge_checkboxes if gauge not in ["MFC9-5", "MFC10-5"])):
            ax.set_ylabel("Flow Rate (SCCM)")  # 只有當 MFC9-5 被選中且其他都未選中時設置為 SCCM
        elif (self.gauge_checkboxes["MFC10-5"].isChecked() and not self.gauge_checkboxes["MFC9-5"].isChecked() and
            all(not self.gauge_checkboxes[gauge].isChecked() for gauge in self.gauge_checkboxes if gauge not in ["MFC9-5", "MFC10-5"])):
            ax.set_ylabel("Flow Rate (SCCM)")  # 只有當 MFC10-5 被選中且其他都未選中時設置為 SCCM
        elif (self.gauge_checkboxes["MFC10-5"].isChecked() and self.gauge_checkboxes["MFC9-5"].isChecked() and
            all(not self.gauge_checkboxes[gauge].isChecked() for gauge in self.gauge_checkboxes if gauge not in ["MFC9-5", "MFC10-5"])):
            ax.set_ylabel("Flow Rate (SCCM)")
        else:
            ax.set_ylabel("Press Value (mbar)")  # 否則設置為 mbar
        #ax2.set_ylabel("Flow Rate (SCCM)")
        
        # 調整圖例的位置
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1), framealpha=0.5)
        #ax2.legend(loc='lower right', bbox_to_anchor=(1, 1), framealpha=0.5)
        ax.grid()

        # 設置 X 軸格式
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d %H:%M:%S"))  # 只顯示時間
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 添加可移動的縱線
        if self.all_times:
            self.vertical_line = ax.axvline(x=self.all_times[0], color='r', linestyle='--')  # 初始位置

            def on_mouse_move(event):
                try:
                    if event.inaxes == ax:  # 確保鼠標在正確的坐標軸上
                        if event.xdata is not None:  # 確保 xdata 是有效的
                            # 將 event.xdata 轉換為 datetime
                            date_time = mdates.num2date(event.xdata).replace(tzinfo=None)

                            # 格式化 datetime 為字符串
                            formatted_time = date_time.strftime("%Y-%m-%d %H:%M:%S.%f")

                            new_x_values = [event.xdata]  # 將 event.xdata 放入列表中
                            self.vertical_line.set_xdata(new_x_values)  # 更新縱線位置

                            # show the datatime
                            self.vline_time_label.setText(f'<b>Time:</b> {formatted_time}')
                            self.figure_ctr.canvas.draw()

		                    # 獲取當前時間點的 gauge 值
                            for gauge_id in self.gauge_checkboxes.keys():
                                if gauge_id in filtered_data:
                                    # 獲取該 gauge 的數據
                                    gauge_data = filtered_data[gauge_id]
                                    # 找到最近的兩個點
                                    lower_point = None
                                    upper_point = None

                                    for dt, val in gauge_data:
                                        dt = dt.replace(tzinfo=None)
                                        if dt < date_time:
                                            lower_point = (dt, val)
                                        elif dt > date_time and upper_point is None:
                                            upper_point = (dt, val)
                                            break

                                    # 如果找到了兩個點，進行插值
                                    if lower_point and upper_point:
                                        dt1, val1 = lower_point
                                        dt2, val2 = upper_point

                                        # 線性插值計算
                                        slope = (val2 - val1) / (dt2 - dt1).total_seconds()  # 計算斜率
                                        interpolated_value = val1 + slope * (date_time - dt1).total_seconds()  # 計算插值

                                        # 更新相應的 QLabel
                                        getattr(self, f'vline_{gauge_id.replace("-", "_")}_label').setText(f'<b>{gauge_id}:</b> {interpolated_value:.2f} ')
                                    else:
                                        # 若沒有找到足夠的點，則顯示 N/A
                                        getattr(self, f'vline_{gauge_id.replace("-", "_")}_label').setText(f'<b>{gauge_id}:</b> N/A ')

                            # 更新閥門狀態顯示
                            if self.valve_status_dialog is not None:
                                current_valve_states = self.get_valve_states_at_time(date_time)
                                self.valve_status_dialog.valve_states = current_valve_states  # 更新 ValveStatusDialog 的閥門狀態
                                self.valve_status_dialog.update_valve_display()  # 調用 ValveStatusDialog 的更新方法

                except Exception as e:
                    print(f"An error occurred: {e}")  # 輸出錯誤信息

            self.figure_ctr.canvas.mpl_connect('motion_notify_event', on_mouse_move)

        # 自動調整 X 軸範圍
        if self.all_times:
            ax.set_xlim([min(self.all_times), max(self.all_times)])  # 設置 X 軸範圍

        # 調整邊距
        plt.subplots_adjust(bottom=0.2)  # 調整底部邊距

        # 顯示圖表
        self.figure_ctr.tight_layout()  # 自動調整子圖參數
        self.canvas_ctr.draw()

        self.ctr_char_gen = True
        self.display_status("Trend chart is generated", "done")

    def get_valve_states_at_time(self, date_time):
        # 確保 date_time 是 offset-naive
        if date_time.tzinfo is not None:
            date_time = date_time.replace(tzinfo=None)

        current_states = {}
        
        for valve_id, states in self.valve_log_data.items():
            # 遍歷狀態列表，找到最新的狀態
            for timestamp, state in reversed(states):
                # 確保 timestamp 是 offset-naive
                if timestamp.tzinfo is not None:
                    timestamp = timestamp.replace(tzinfo=None)
                    
                if timestamp <= date_time:
                    current_states[valve_id] = state
                    break  # 找到最新狀態後退出循環

        return current_states

    def zoom_chart(self, zoom_in=True):
        if self.ctr_char_gen == False:
            self.display_status("Trend Chart is unavailabe !", "fail")
            QMessageBox.information(self, "Log Analyser", "Trend Chart is unavailabe !", QMessageBox.StandardButton.Ok)
            return
        
        ax = self.figure_ctr.get_axes()[0]  # 獲取第一個坐標軸
        xlim = ax.get_xlim()  # 獲取當前 x 軸範圍

        # 獲取選擇的倍率
        #selected_multiplier = self.zoom_combo.currentText().replace('%', '')
        #multiplier = int(selected_multiplier) / 100  # 轉換為小數
        multiplier = 0.1
        
        if zoom_in:
            ax.set_xlim([xlim[0] + (xlim[1] - xlim[0]) * multiplier, xlim[1] - (xlim[1] - xlim[0]) * multiplier])  # 放大 10%
        else:
            ax.set_xlim([xlim[0] - (xlim[1] - xlim[0]) * multiplier, xlim[1] + (xlim[1] - xlim[0]) * multiplier])  # 縮小 10%
        
        self.canvas_ctr.draw()  # 重新繪製圖形

    def shift_chart(self, left_shift=True):
        if self.ctr_char_gen == False:
            self.display_status("Trend Chart is unavailabe !", "fail")
            QMessageBox.information(self, "Log Analyser", "Trend Chart is unavailabe !", QMessageBox.StandardButton.Ok)
            return
        
        ax = self.figure_ctr.get_axes()[0]  # 獲取第一個坐標軸
        xlim = ax.get_xlim()  # 獲取當前 x 軸範圍

        # 獲取選擇的倍率
        selected_multiplier = self.shift_combo.currentText().replace('%', '')
        multiplier = int(selected_multiplier) / 100  # 轉換為小數
        
        if left_shift:
            ax.set_xlim([xlim[0] - (xlim[1] - xlim[0]) * multiplier, xlim[1] - (xlim[1] - xlim[0]) * multiplier])  # 向左平移 10%
        else:
            ax.set_xlim([xlim[0] + (xlim[1] - xlim[0]) * multiplier, xlim[1] + (xlim[1] - xlim[0]) * multiplier])  # 向右平移 10%
        
        self.canvas_ctr.draw()  # 重新繪製圖形

    def multijet_show(self):
        # 獲取當前腳本的路徑
        current_directory = os.path.dirname(os.path.abspath(__file__))
        image_path = os.path.join(current_directory, "multijet_pipe_sample.png")  # 替換為您的圖片檔名

        dialog = MultijetImgDialog(image_path, self)
        dialog.show()

    def valve_state_show(self):
        # 獲取當前腳本的路徑
        current_directory = os.path.dirname(os.path.abspath(__file__))
        multijet_image_path = os.path.join(current_directory, "multijet_pipe.png")
        if not os.path.exists(multijet_image_path):
            #self.display_status("multijet_pipe.png not found !", "fail")
            print("multijet_pipe.png not found")
            return

        valve_open_image_path = os.path.join(current_directory, "valve_open.png")
        if not os.path.exists(valve_open_image_path):
            #self.display_status("valve_open.png not found !", "fail")
            print("valve_open.png not found")
            return

        valve_close_image_path = os.path.join(current_directory, "valve_close.png")
        if not os.path.exists(valve_close_image_path):
            #self.display_status("valve_close.png not found !", "fail")
            print("valve_close.png not found")
            return

        valve_positions = {
            "V0-3": (174, 163),
            "V0-4": (174, 581),
            "V0-5": (118, 118),
            "V0-7": (118, 163),
            "V0-10": (174, 22),
            "V0-13": (118, 227),
            "V0-16": (174, 286),
            "V0-31": (80, 335),
            "V0-32": (80, 376),
            "V0-33": (80, 416),
            "V0-34": (80, 460),
            "V0-35": (80, 499),
            # CH1
            "V1-3": (896, 22),
            "V1-4": (520, 57),
            "V1-9": (588, 57),
            "V1-10": (804, 56),
            "V1-11": (247, 22),
            "V1-13": (304, 56),
            "V1-21": (956, 22),
            "V1-22": (956, 57),
            #CH2
            "V2-3": (896, 91),
            "V2-4": (520, 125),
            "V2-9": (588, 126),
            "V2-11": (247, 91),
            "V2-13": (304, 125),
            "V2-21": (956, 92),
            #CH3
            "V3-3": (896, 152),
            "V3-4": (520, 186),
            "V3-9": (588, 186),
            "V3-11": (247, 152),
            "V3-13": (304, 186),
            "V3-21": (956, 152),
            #CH4
            "V4-3": (896, 216),
            "V4-4": (478, 245),
            "V4-5": (386, 278),
            "V4-9": (588, 244),
            "V4-11": (247, 216),
            "V4-13": (304, 277),
            "V4-21": (956, 216),
            #CH9
            "V9-3": (896, 343),
            "V9-4": (804, 344),
            "V9-5": (459, 370),
            "V9-7": (626, 387),
            "V9-8": (705, 387),
            "V9-11": (247, 344),
            "V9-12": (784, 428),
            "V9-13": (304, 427),
            "V9-14": (386, 427),
            "V9-15": (386, 387),
            "V9-16": (386, 344),
            "V9-17": (495, 386),
            "V9-21": (956, 344),
            #CH10
            "V10-3": (896, 470),
            "V10-5": (458, 498),
            "V10-7": (626, 511),
            "V10-8": (705, 512),
            "V10-11": (247, 470),
            "V10-12": (788, 552),
            "V10-13": (304, 552),
            "V10-14": (386, 553),
            "V10-15": (386, 511),
            "V10-16": (386, 471),
            "V10-17": (495, 512),
            "V10-21": (956, 470),
        }

        # 創建 ValveStatusDialog 實例並儲存
        self.valve_status_dialog = ValveStatusDialog(multijet_image_path, valve_open_image_path, valve_close_image_path, valve_positions, self)
        self.valve_status_dialog.show()

if __name__ == "__main__":
    app = QApplication([])
    window = LogAnalyser()
    window.show()
    app.exec()'
'''
import sys
import os
import re
import win32com.client
from datetime import datetime
from collections import defaultdict
import subprocess
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QPushButton, QLabel,
    QLineEdit, QListWidget, QVBoxLayout, QHBoxLayout, QWidget, QProgressBar, QTabWidget,
    QRadioButton, QMessageBox, QSizePolicy, QDateEdit, QSpinBox, QGridLayout, QCheckBox,
    QComboBox, QDialog, QTreeView, QHeaderView, QScrollArea, QButtonGroup, QListWidgetItem
)
from PyQt6.QtCore import Qt, QTimer, QSize, QPoint
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.gridspec as gridspec
from datetime import datetime
from PyQt6.QtGui import QPainter, QPixmap, QIcon, QFileSystemModel, QColor

class MultijetImgDialog(QDialog):
    def __init__(self, image_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("MultiJet Pipe")
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinMaxButtonsHint)  # 添加最小化和最大化按鈕

        # 載入圖片並獲取其大小
        pixmap = QPixmap(image_path)
        self.image_label = QLabel()
        self.image_label.setPixmap(pixmap)  # 設置圖片到 QLabel

        # 設置對話框大小為圖片大小
        self.resize(pixmap.size())  # 將對話框大小設置為圖片大小

        layout = QVBoxLayout()
        layout.addWidget(self.image_label)
        self.setLayout(layout)

class ValveImageLabel(QLabel):
    def __init__(self, pixmap, valve_positions, valve_close_img, parent=None):
        super().__init__(parent)
        self.pixmap = pixmap
        self.valve_positions = valve_positions  # 儲存所有閥門的位置
        self.valve_images = {valve_id: QPixmap(valve_close_img) for valve_id in valve_positions.keys()}  # 儲存閥門圖片

    def paintEvent(self, event):
        painter = QPainter(self)  # 在 QLabel 本身上繪製
        painter.drawPixmap(0, 0, self.pixmap)  # 繪製主圖片

        # 繪製所有閥門圖片
        for valve_id, position in self.valve_positions.items():
            # 確保 position 是 QPoint
            if isinstance(position, tuple):
                position = QPoint(*position)  # 使用 * 解包元組
            painter.drawPixmap(position, self.valve_images[valve_id])  # 繪製閥門圖片

    def sizeHint(self):
        return self.pixmap.size()  # 返回主圖片的大小

    def set_valve_image(self, valve_id, valve_image_path):
        self.valve_images[valve_id] = QPixmap(valve_image_path)
        self.update()  # 更新 QLabel 以顯示新的閥門圖片

class ValveStatusDialog(QDialog):
    def __init__(self, multijet_image_path, valve_open_image_path, valve_close_image_path, valve_positions, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Valve Status")
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinMaxButtonsHint)  # 添加最小化和最大化按鈕

        # 載入主圖片
        self.pixmap = QPixmap(multijet_image_path)
        self.valve_open_image_path = valve_open_image_path
        self.valve_close_image_path = valve_close_image_path
        self.valve_positions = valve_positions

        # 創建自定義 QLabel 來顯示圖片
        self.image_label = ValveImageLabel(self.pixmap, self.valve_positions, self.valve_close_image_path)

        # 創建滾動區域
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(False)  # 使滾動區域可調整大小
        self.scroll_area.setWidget(self.image_label)  # 將自定義 QLabel 設置為滾動區域的內容

        # 設置對話框大小為圖片大小
        self.resize(self.pixmap.size())

        # 創建顯示座標的 QLabel
        self.coordinate_label = QLabel("座標: (0, 0)")
        self.coordinate_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        layout = QVBoxLayout()
        layout.addWidget(self.scroll_area)
        layout.addWidget(self.coordinate_label)  # 添加座標顯示
        self.setLayout(layout)

        # 連接滑鼠移動事件
        self.image_label.setMouseTracking(True)  # 啟用滑鼠追蹤
        self.image_label.mouseMoveEvent = self.mouse_move_event  # 自定義滑鼠移動事件

        # 初始化閥門狀態
        self.valve_states = {valve_id: 0 for valve_id in self.valve_positions.keys()}  # 預設為關閉狀態

        # 繪製閥門
        self.update_valve_display()

    def closeEvent(self, event):
        # 在關閉對話框時釋放資源或進行清理
        print("Closing ValveStatusDialog")  # 可選，顯示關閉信息
        event.accept()  # 確保事件被接受

    def update_valve_display(self):
        for valve_id, position in self.valve_positions.items():
            status = self.valve_states[valve_id]
            valve_image = self.valve_close_image_path if status == 0 else self.valve_open_image_path
            self.image_label.set_valve_image(valve_id, valve_image)  # 更新閥門圖片

    def mouse_move_event(self, event):
        # 獲取滑鼠座標
        x = event.pos().x()
        y = event.pos().y()
        self.coordinate_label.setText(f"座標: ({x}, {y})")  # 更新座標顯示

class LogAnalyser(QMainWindow):
    def __init__(self):
        super().__init__()
        # 設置窗口圖標
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "zeiss-logo.png")  # 替換為您的圖標文件名
        self.setWindowIcon(QIcon(icon_path))
        self.processed_files = 0  # 已解析的檔案數
        self.log_data = None
        self.min_time = None
        self.max_time = None
        self.valve_log_data = {}
        self.valve_status_dialog = None
        # 初始化計時器
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_processing_label)  # 連接計時器的超時信號到更新函數
        self.dot_count = 0
        self.mask_events = {}
        self.workflow_data = {}
        self.ctr_char_gen = False
        self.all_times = []
        self.protocol_data = {}
        self.prot_folder_path = None
        self.prot_start_date = None
        self.prot_end_date = None
        self.last_unchecked_gauge = None
        self.gauge_types = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3", "MFC9-5", "MFC10-5"]

        self.initUI()

    def initUI(self):
        self.setWindowTitle("Log Analyser v1.0.0")
        self.setGeometry(100, 100, 1200, 930)

        # 主 Widget 和 Layout
        central_widget = QWidget()
        central_layout = QVBoxLayout(central_widget)
        self.setCentralWidget(central_widget)
        
        # 創建 QLabel 用於顯示 logo
        logo_label = QLabel()
        if getattr(sys, 'frozen', False):
            # 如果應用程序是從可執行檔運行
            logo_pixmap = QPixmap(os.path.join(sys._MEIPASS, "zeiss-logo.png"))
        else:
            # 如果應用程序是從源代碼運行
            logo_pixmap = QPixmap("zeiss-logo.png")
        logo_label.setPixmap(logo_pixmap)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignRight)

        # 將 logo 添加到佈局
        central_layout.addWidget(logo_label)

        # 功能表
        self.create_menu_bar()

        # 顯示資料夾路徑
        self.folder_path_edit = QLineEdit()
        self.folder_path_edit.setFixedHeight(30)
        self.folder_path_edit.setFixedWidth(550)
        central_layout.addWidget(self.folder_path_edit)

        # 建立 Tab
        self.tab_widget = QTabWidget()
        central_layout.addWidget(self.tab_widget)

        #### Tab 1 - 主要資料處理
        wf_widget = QWidget()
        wf_layout = QVBoxLayout(wf_widget)

        # Tab 1 完成，加入到 TabWidget
        self.tab_widget.addTab(wf_widget, "Work Flow")
        
        load_unload_label = QLabel("Load/Unload information")
        work_step_label = QLabel("Work Steps")

        self.lwt_mask_info = QListWidget()
        self.lwt_mask_info.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.lwt_mask_info.itemClicked.connect(self.display_workflow)

        self.lwt_wk_flow = QListWidget()
        self.lwt_wk_flow.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        left_v_layout = QVBoxLayout()
        left_v_layout.addWidget(load_unload_label)
        left_v_layout.addWidget(self.lwt_mask_info)

        right_v_layout = QVBoxLayout()
        right_v_layout.addWidget(work_step_label)
        right_v_layout.addWidget(self.lwt_wk_flow)

        list_h_layout = QHBoxLayout()
        list_h_layout.addLayout(left_v_layout)
        list_h_layout.addLayout(right_v_layout)
        list_h_layout.setStretchFactor(left_v_layout, 1)
        list_h_layout.setStretchFactor(right_v_layout, 1)

        # 創建 QFileSystemModel
        self.file_system_model = QFileSystemModel()

        # 獲取所有可用的磁碟驅動器
        drives = [f"{d}:\\" for d in range(65, 91) if os.path.exists(f"{chr(d)}:\\")]  # ASCII 65-90 對應 A-Z

        # 創建 QTreeView
        self.tree_view = QTreeView()
        self.tree_view.setFixedHeight(170)
        self.tree_view.setModel(self.file_system_model)

        # 添加「本機」作為根節點
        local_machine_index = self.file_system_model.setRootPath('')  # 設置根路徑為空以顯示所有磁碟驅動器
        self.tree_view.setRootIndex(local_machine_index)  # 設置樹的根節點

        # 設置樹狀結構的顯示屬性
        self.tree_view.setHeaderHidden(False)  # 顯示標題
        self.tree_view.setAlternatingRowColors(True)  # 交替行顏色

        # 設置欄位大小可調整
        header = self.tree_view.header()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # 設置欄位為可互動調整大小

        # 將可用的磁碟驅動器添加到樹狀結構
        for drive in drives:
            drive_index = self.file_system_model.index(drive)  # 獲取驅動器的索引
            self.tree_view.setRootIndex(drive_index)  # 設置樹的根節點為該驅動器

        select_folder_button = QPushButton("Select Folder")
        select_folder_button.clicked.connect(self.select_folder_from_tree)

        wf_layout.addLayout(list_h_layout)
        wf_layout.addWidget(self.tree_view)  # 將樹狀結構添加到工作流程佈局中
        wf_layout.addWidget(select_folder_button, alignment=Qt.AlignmentFlag.AlignLeft)

        # 處理按鈕
        self.process_button = QPushButton("Process")
        self.process_button.setFixedSize(80, 40)
        self.process_button.setStyleSheet("""
            QPushButton {
                font-weight: bold;
                background-color: white;
                color: black; 
                border: 2px solid #000;
            }
            QPushButton:hover {
                font-weight: bold;
                background-color: lightblue;
                color: black;
            }
        """)
        self.process_button.clicked.connect(self.process_raw_log)
        central_layout.addWidget(self.process_button)

        # 新增進度條
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)  # 預設範圍為 0 到 100
        self.progress_bar.setValue(0)  # 初始值為 0
        self.progress_bar.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        self.progress_bar.setStyleSheet("QProgressBar::chunk { background-color: lightgreen; }")
        central_layout.addWidget(self.progress_bar)

        process_layout = QHBoxLayout()
        process_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        process_layout.addWidget(self.process_button, alignment=Qt.AlignmentFlag.AlignLeft)
        process_layout.addWidget(self.progress_bar, alignment=Qt.AlignmentFlag.AlignLeft)
        central_layout.addLayout(process_layout)

        # 完成訊息
        self.status_label = QLabel("")
        central_layout.addWidget(self.status_label, alignment=Qt.AlignmentFlag.AlignLeft)

        # Tab 2 - Mask ID Analysis
        mask_analysis_widget = QWidget()
        mask_analysis_v_layout = QVBoxLayout(mask_analysis_widget)
        self.tab_widget.addTab(mask_analysis_widget, "Mask ID Analysis")

        # Add buttons to generate charts
        load_unload_chart_button = QPushButton("Load/Unload Time Chart")
        load_unload_chart_button.clicked.connect(self.generate_load_unload_time_chart)

        duration_chart_button = QPushButton("Duration Time Chart")
        duration_chart_button.clicked.connect(self.generate_duration_time_chart)

        mask_analysis_btn_h_layout = QHBoxLayout()
        mask_analysis_btn_h_layout.addWidget(load_unload_chart_button)
        mask_analysis_btn_h_layout.addWidget(duration_chart_button)
        mask_analysis_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        mask_analysis_v_layout.addLayout(mask_analysis_btn_h_layout)

        # Canvas for displaying charts
        self.figure_mask_analysis = Figure(figsize=(10, 6))
        self.canvas_mask_analysis  = FigureCanvas(self.figure_mask_analysis)
        mask_analysis_v_layout.addWidget(self.canvas_mask_analysis)

        ### Tab 3 - Install Info
        install_widge = QWidget()
        install_layout = QVBoxLayout(install_widge)

        # Tab 3 完成，加入到 TabWidget
        self.tab_widget.addTab(install_widge, "Install Log")

        # Install Info 列表
        self.install_info_list = QListWidget()
        self.install_info_list.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        install_layout.addWidget(self.install_info_list)

        ### Tab 4 - MSC Info
        msc_widge = QWidget()
        msc_v_layout = QVBoxLayout(msc_widge)

        # Tab 4 完成，加入到 TabWidget
        self.tab_widget.addTab(msc_widge, "MSC Info")

        self.tab_widget.currentChanged.connect(self.on_tab_changed)

        # Radio Buttons for msc version 
        label_msc_ver = QLabel("MSC ver : ")
        self.msc_v2_radio = QRadioButton("MSC 2.x")
        self.msc_v2_radio.toggled.connect(self.on_format_selected)
        self.selected_format = "MSC 3.x"  # 預設選擇
        self.msc_v3_radio = QRadioButton("MSC 3.x")
        self.msc_v3_radio.setChecked(True)
        self.msc_v3_radio.toggled.connect(self.on_format_selected)

        msc_version_group = QButtonGroup(self)
        msc_version_group.addButton(self.msc_v2_radio)
        msc_version_group.addButton(self.msc_v3_radio)

        msc_ver_h_layout = QHBoxLayout()

        msc_ver_h_layout.addWidget(label_msc_ver)
        msc_ver_h_layout.addWidget(self.msc_v2_radio)
        msc_ver_h_layout.addWidget(self.msc_v3_radio)

        # Gauge checkbox
        label_gauge_tp = QLabel("   CTR : ")
        self.gauge_checkboxes = {}
        msc_ver_h_layout.addWidget(label_gauge_tp)

        for gauge in self.gauge_types:
            checkbox = QCheckBox(gauge)
            checkbox.setChecked(True)  # 預設為選中
            self.gauge_checkboxes[gauge] = checkbox  # 將複選框存儲在字典中
            msc_ver_h_layout.addWidget(checkbox)  # 將複選框添加到布局中

            # 連接複選框的狀態變化事件
            checkbox.stateChanged.connect(self.on_gauge_checkbox_changed)

        msc_ver_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(msc_ver_h_layout)

        # Radio Buttons for chart source 
        label_chart_source = QLabel("Chart source : ")
        self.multijet_chart_radio = QRadioButton("mjnxtdebug")
        self.multijet_chart_radio.setChecked(True)
        self.multijet_chart_radio.toggled.connect(self.on_chart_source_selected)
        self.selected_chart = "mjnxtdebug"  # 預設選擇
        self.protocol_chart_radio = QRadioButton("protocol")
        self.protocol_chart_radio.toggled.connect(self.on_chart_source_selected)

        chart_source_group = QButtonGroup(self)
        chart_source_group.addButton(self.multijet_chart_radio)
        chart_source_group.addButton(self.protocol_chart_radio)
        
        chart_source_h_layout = QHBoxLayout()
        chart_source_h_layout.addWidget(label_chart_source)
        chart_source_h_layout.addWidget(self.multijet_chart_radio)
        chart_source_h_layout.addWidget(self.protocol_chart_radio)

        chart_source_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(chart_source_h_layout)

        # 時間篩選器
        time_filter_layout = QGridLayout()
        #time_filter_layout.setHorizontalSpacing(1)  # 設置水平間距
        #time_filter_layout.setVerticalSpacing(1)     # 設置垂直間距
        #time_filter_layout.setContentsMargins(1, 1, 1, 1)  # 設置邊距為 0

        # 標題列
        header_labels = ["", "Date", "Hr", "Min", "Sec"]
        for col, text in enumerate(header_labels):
            label = QLabel(text)
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)  # 置中對齊
            time_filter_layout.addWidget(label, 0, col)

        # Start Time
        time_filter_layout.addWidget(QLabel("Start:"), 1, 0)

        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("yyyy/MM/dd")
        time_filter_layout.addWidget(self.start_date_edit, 1, 1)

        self.start_hour_spinbox = QSpinBox()
        self.start_hour_spinbox.setRange(0, 23)
        self.start_hour_spinbox.setFixedWidth(60)
        self.start_hour_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_hour_spinbox, 1, 2)

        self.start_minute_spinbox = QSpinBox()
        self.start_minute_spinbox.setRange(0, 59)
        self.start_minute_spinbox.setFixedWidth(60)
        self.start_minute_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_minute_spinbox, 1, 3)

        self.start_second_spinbox = QSpinBox()
        self.start_second_spinbox.setRange(0, 59)
        self.start_second_spinbox.setFixedWidth(60)
        self.start_second_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.start_second_spinbox, 1, 4)

        # End Time
        time_filter_layout.addWidget(QLabel("End:"), 2, 0)

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy/MM/dd")
        time_filter_layout.addWidget(self.end_date_edit, 2, 1)

        self.end_hour_spinbox = QSpinBox()
        self.end_hour_spinbox.setRange(0, 23)
        self.end_hour_spinbox.setFixedWidth(60)
        self.end_hour_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_hour_spinbox, 2, 2)

        self.end_minute_spinbox = QSpinBox()
        self.end_minute_spinbox.setRange(0, 59)
        self.end_minute_spinbox.setFixedWidth(60)
        self.end_minute_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_minute_spinbox, 2, 3)

        self.end_second_spinbox = QSpinBox()
        self.end_second_spinbox.setRange(0, 59)
        self.end_second_spinbox.setFixedWidth(60)
        self.end_second_spinbox.setStyleSheet("""
            QSpinBox::up-button {
                min-height: 10px;
                max-height: 10px;
            }
            QSpinBox::down-button {
                min-height: 10px;
                max-height: 10px;
            }
        """)
        time_filter_layout.addWidget(self.end_second_spinbox, 2, 4)

        # Protocol 
        self.mask_project_combobox = QComboBox()
        self.mask_project_combobox.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents) # 自動調整大小
        self.mask_project_combobox.setMinimumWidth(280)
        self.protocol_combobox = QComboBox()
        self.protocol_combobox.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents) # 自動調整大小
        self.protocol_combobox.setMinimumWidth(220)
        self.mask_project_label = QLabel("Project:")
        self.protocol_label = QLabel("Protocol:")
        self.process_label = QLabel("Process:")
        self.u_label = QLabel("U: ")
        self.v_label = QLabel("V: ")

        protocol_h_layout = QHBoxLayout()
        protocol_h_layout.setContentsMargins(0, 0, 0, 0)
        protocol_h_layout.addWidget(self.mask_project_label)
        protocol_h_layout.addWidget(self.mask_project_combobox)
        protocol_h_layout.addWidget(self.protocol_label)
        protocol_h_layout.addWidget(self.protocol_combobox)
        protocol_h_layout.addWidget(self.process_label)
        protocol_h_layout.addWidget(self.u_label)
        protocol_h_layout.addWidget(self.v_label)
        protocol_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(protocol_h_layout)

        # Protocol reader
        protocol_reader_h_layout = QHBoxLayout()
        protocol_reader_button = QPushButton("Protocol Reader V5")
        protocol_reader_button.clicked.connect(lambda: self.open_protocol_reader(self.prot_start_date, self.prot_end_date, self.prot_folder_path))

        protocol_reader_h_layout.addWidget(protocol_reader_button)
        protocol_reader_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(protocol_reader_h_layout)
     
        # Multijet Chart
        multijet_chart_h_layout = QHBoxLayout()
        multijet_pipe_button = QPushButton("MultiJet Pipe")
        multijet_pipe_button.clicked.connect(self.multijet_show)
        valve_status_button = QPushButton("Valve State")
        valve_status_button.clicked.connect(self.valve_state_show)

        multijet_chart_h_layout.addWidget(multijet_pipe_button)
        multijet_chart_h_layout.addWidget(valve_status_button)

        vline_h_layout = QHBoxLayout()
        # Display time data
        self.vline_time_label = QLabel("")
        # Display CTR data
        self.vline_P1_1_label = QLabel("")
        self.vline_P2_1_label = QLabel("")
        self.vline_P3_1_label = QLabel("")
        self.vline_P4_1_label = QLabel("")
        self.vline_P9_1_label = QLabel("")
        self.vline_P9_2_label = QLabel("")
        self.vline_P9_3_label = QLabel("")
        self.vline_P10_1_label = QLabel("")
        self.vline_P10_2_label = QLabel("")
        self.vline_P10_3_label = QLabel("")
        self.vline_MFC9_5_label = QLabel("")
        self.vline_MFC10_5_label = QLabel("")
        vline_h_layout.addWidget(self.vline_time_label)
        vline_h_layout.addWidget(self.vline_P1_1_label)
        vline_h_layout.addWidget(self.vline_P2_1_label)
        vline_h_layout.addWidget(self.vline_P3_1_label)
        vline_h_layout.addWidget(self.vline_P4_1_label)
        vline_h_layout.addWidget(self.vline_P9_1_label)
        vline_h_layout.addWidget(self.vline_P9_2_label)
        vline_h_layout.addWidget(self.vline_P9_3_label)
        vline_h_layout.addWidget(self.vline_P10_1_label)
        vline_h_layout.addWidget(self.vline_P10_2_label)
        vline_h_layout.addWidget(self.vline_P10_3_label)
        vline_h_layout.addWidget(self.vline_MFC9_5_label)
        vline_h_layout.addWidget(self.vline_MFC10_5_label)

        # 讓時間篩選器靠左對齊
        time_filter_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(time_filter_layout)
        multijet_chart_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(multijet_chart_h_layout)
        vline_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        msc_v_layout.addLayout(vline_h_layout)

        # 添加畫布以顯示圖表
        self.figure_ctr = Figure(figsize=(10, 6))
        self.canvas_ctr = FigureCanvas(self.figure_ctr)
        msc_v_layout.addWidget(self.canvas_ctr)

        # 添加按鈕以生成趨勢圖
        gen_ctr_button = QPushButton("")
        gen_ctr_button.setToolTip("Generate Chart")  # 設置工具提示
        line_chart_icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "line_chart_icon.png")
        gen_ctr_button.setIcon(QIcon(line_chart_icon_path))  # 設置按鈕圖標
        gen_ctr_button.setIconSize(QSize(30, 30))  # 設置圖標大小
        gen_ctr_button.clicked.connect(lambda: self.create_chart(generate_chart=True))
        expo_button = QPushButton("")
        expo_button.setToolTip("Export to Excel")  # 設置工具提示
        excel_icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_icon.png")
        expo_button.setIcon(QIcon(excel_icon_path))  # 設置按鈕圖標
        expo_button.setIconSize(QSize(30, 30))  # 設置圖標大小
        expo_button.clicked.connect(lambda: self.create_chart(generate_chart=False))

        msc_btn_h_layout = QHBoxLayout()
        msc_btn_h_layout.addWidget(gen_ctr_button)
        msc_btn_h_layout.addWidget(expo_button)
        msc_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)

        # 添加 QComboBox 來選擇移動倍率
        self.shift_combo = QComboBox()
        self.shift_combo.addItems(["10%", "30%", "50%"])
        self.shift_combo.setCurrentIndex(0)  # 預設選擇 10%

        # 添加zoom按鈕
        zoom_in_button = QPushButton("-")
        zoom_in_button.clicked.connect(lambda: self.zoom_chart(zoom_in=False))
        zoom_out_button = QPushButton("+")
        zoom_out_button.clicked.connect(lambda: self.zoom_chart(zoom_in=True))

        # 添加平移按鈕
        left_shift_button = QPushButton("<--")
        left_shift_button.clicked.connect(lambda: self.shift_chart(left_shift=True))
        right_shift_button = QPushButton("-->")
        right_shift_button.clicked.connect(lambda: self.shift_chart(left_shift=False))

        # 將 zoom/shift 按鈕添加到佈局中
        msc_zoom_btn_h_layout = QHBoxLayout()
        msc_zoom_btn_h_layout.addWidget(zoom_in_button)
        msc_zoom_btn_h_layout.addWidget(zoom_out_button)
        msc_zoom_btn_h_layout.addWidget(left_shift_button)
        msc_zoom_btn_h_layout.addWidget(right_shift_button)
        msc_zoom_btn_h_layout.addWidget(self.shift_combo)
        msc_zoom_btn_h_layout.setAlignment(Qt.AlignmentFlag.AlignRight)

        msc_v_layout.addLayout(msc_zoom_btn_h_layout)
        msc_v_layout.addLayout(msc_btn_h_layout)        
        
    def create_menu_bar(self):
        menu_bar = self.menuBar()

        # File menu
        file_menu = menu_bar.addMenu("File")
        open_action = file_menu.addAction("Open Folder")
        open_action.triggered.connect(self.select_folder)

        # 新增 Log menu
        log_menu = menu_bar.addMenu("Log")
        self.log_actions = {
            "LogService": log_menu.addAction("LogService"),
            "Install": log_menu.addAction("Install"),
            "mjnxtdebug": log_menu.addAction("mjnxtdebug"),
            "Protocol": log_menu.addAction("Protocol")
        }

        for action in self.log_actions.values():
            action.setCheckable(True)  # 設置為可勾選
            action.setChecked(True)  # 預設為勾選

        # Info menu
        info_menu = menu_bar.addMenu("Info")
        about_action = info_menu.addAction("About")
        about_action.triggered.connect(self.show_about_dialog)

    def update_processing_label(self):
        self.dot_count = (self.dot_count + 1) % 4  # 使點數在 0 到 3 之間循環
        dots = '.' * self.dot_count  # 根據計數生成點數
        self.display_status(f"Processing {dots}", "ongoing")
    
    def show_about_dialog(self):
        QMessageBox.about(self, "About", "Log Analyser v1.0.0\nAuthor : Davian Kuo\nE-mail : davian.kuo@zeiss.com")

    def display_status(self, msg, type):
        self.status_label.setText(msg)
        if type == "done":
            self.status_label.setStyleSheet("background-color: lightgreen;")
        elif type == "fail":
            self.status_label.setStyleSheet("background-color: lightpink;")
        elif type == "ongoing":
            self.status_label.setStyleSheet("background-color: lightyellow;")

    def select_folder_from_tree(self):
        # 獲取當前選中的索引
        selected_index = self.tree_view.currentIndex()
        if selected_index.isValid():  # 確保選中的索引有效
            # 獲取選中的資料夾路徑
            folder_path = self.file_system_model.filePath(selected_index)
            # 將路徑填入 folder_path_edit
            self.folder_path_edit.setText(folder_path)
            self.display_status("Folder selected !", "done")
        else:
            self.display_status("Please select folder !", "fail")
            QMessageBox.information(self, "Log Analyser", "Please select folder !", QMessageBox.StandardButton.Ok)

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.folder_path_edit.setText(folder)
            self.display_status("Folder selected !", "done")

    def set_time_filter_ctrl(self, en):
        self.start_date_edit.setEnabled(en)
        self.start_hour_spinbox.setEnabled(en)
        self.start_minute_spinbox.setEnabled(en)
        self.start_second_spinbox.setEnabled(en)

        self.end_date_edit.setEnabled(en)
        self.end_hour_spinbox.setEnabled(en)
        self.end_minute_spinbox.setEnabled(en)
        self.end_second_spinbox.setEnabled(en)

    def on_chart_source_selected(self):
        if self.multijet_chart_radio.isChecked():
            self.selected_chart = "mjnxtdebug"
            self.set_time_filter_ctrl(True)
        elif self.protocol_chart_radio.isChecked():
            self.selected_chart = "protocol"
            self.set_time_filter_ctrl(False)

        for gauge_id in self.gauge_types:
            self.gauge_checkboxes[gauge_id].setEnabled(True)
            self.gauge_checkboxes[gauge_id].setChecked(True)
    
    def open_protocol_reader(self, start_date, end_date, folder_path):
        if start_date == None or end_date == None or folder_path == None:
            self.display_status("Protocol date is None !", "fail")
            QMessageBox.information(self, "Log Analyser", "Protocol date is None !", QMessageBox.StandardButton.Ok)
            return

        # 獲取當前腳本的路徑
        excel_file_path = os.path.join(folder_path, "Protocol reader_V5.xlsm")  # 替換為你的檔案名

        # 檢查檔案是否存在
        if os.path.exists(excel_file_path):
            try:
                # 啟動 Excel 應用程式
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True

                workbook = excel.Workbooks.Open(excel_file_path)

                excel.Application.Run("SetInformationInputValues", start_date, end_date, folder_path)
                excel.Application.Run("Log_analysis")

                # 保存並關閉工作簿
                #workbook.Save()
                #workbook.Close()
                #excel.Quit()
            except Exception as e:
                print(f"Failed to fill Excel file: {str(e)}")
        else:
            print("The specified Excel file does not exist !")
            self.display_status("The specified Excel file does not exist !", "fail")
    
    def find_log_files(self):
        found_path = self.folder_path_edit.text()
        if not found_path:
            self.display_status("Please choose folder before parsing !", "fail")
            QMessageBox.information(self, "Log Analyser", "Please choose folder before parsing !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None

        logsrvc_log = []
        install_log = []
        msc_log = []
        protocol_log = []
        process_log = []

        # 遞迴搜尋所有子資料夾
        for root, dirs, files in os.walk(found_path):
            for file in files:
                if self.log_actions["LogService"].isChecked():
                    if file == "LogService.txt" or re.match(r"LogService\d{4}-\d{2}-\d{2}_\d{6}\.txt", file):
                        logsrvc_log.append(os.path.join(root, file))

                if self.log_actions["Install"].isChecked():
                    if file == "Install.txt":
                        install_log.append(os.path.join(root, file))

                if self.log_actions["mjnxtdebug"].isChecked():
                    if re.match(r"^mjnxtdebug\d{8}\.log$", file):
                        msc_log.append(os.path.join(root, file))

                if self.log_actions["Protocol"].isChecked():
                    if re.match(r"Protocol_\d{4}-\d{2}-\d{2}_\d{4}_\d{4}\.txt", file):
                        protocol_log.append(os.path.join(root, file))
                    if re.match(r"ProcessLog_\d{4}-\d{2}-\d{2}_\d{4}_\d{4}\.csv", file):
                        process_log.append(os.path.join(root, file))

        # 檢查是否找到有效的日誌文件
        if self.log_actions["LogService"].isChecked() and not logsrvc_log:
            self.display_status("❌ Can't find valid LogService !", "fail")
            QMessageBox.information(self, "Log Analyser", "❌ Can't find valid LogService !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None
        if self.log_actions["Install"].isChecked() and not install_log:
            self.display_status("❌ Can't find valid Install !", "fail")
            QMessageBox.information(self, "Log Analyser", "❌ Can't find valid Install !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None
        if self.log_actions["mjnxtdebug"].isChecked() and not msc_log:
            self.display_status("❌ Can't find valid mjnxtdebug !", "fail")
            QMessageBox.information(self, "Log Analyser", "❌ Can't find valid mjnxtdebug !", QMessageBox.StandardButton.Ok)
            return None, None, None, None, None, None    
        if self.log_actions["Protocol"].isChecked():
            if not protocol_log:
                self.display_status("❌ Can't find valid protocol file !", "fail")
                QMessageBox.information(self, "Log Analyser", "❌ Can't find valid protocol file !", QMessageBox.StandardButton.Ok)
                return None, None, None, None, None, None
            if not process_log:
                self.display_status("❌ Can't find valid ProcessLog file !", "fail")
                QMessageBox.information(self, "Log Analyser", "❌ Can't find valid ProcessLog file !", QMessageBox.StandardButton.Ok)
                return None, None, None, None, None, None

        # 進度條初始化
        total_files = len(logsrvc_log) + len(install_log) + len(msc_log) + len(protocol_log) + len(process_log)
        self.progress_bar.setRange(0, total_files)
        self.progress_bar.setValue(0)

        return found_path, logsrvc_log, install_log, msc_log, protocol_log, process_log

    def process_raw_log(self):
        self.mask_events, self.workflow_data = {}, {}
        self.log_data, time_range, self.valve_log_data = {}, [], {}
        self.protocol_data = {}

        if not any(action.isChecked() for action in self.log_actions.values()):
            self.display_status("Please select at least one log to parse !", "fail")
            QMessageBox.information(self, "Log Analyser", "Please select at least one log to parse !", QMessageBox.StandardButton.Ok)
            return

        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        
        self.prot_folder_path = os.path.normpath(folder_path)
        self.process_button.setEnabled(False)
        self.process_button.setStyleSheet("background-color: lightgray; color: darkgray;")
        self.processed_files = 0
        self.display_status("Processing .", "ongoing")

        self.timer.start(100)

        # 1. LogService.txt 
        if logsrvc_file:
            parsed_result = self.parse_logsvr(folder_path, logsrvc_file)
            if parsed_result is None:
                self.timer.stop()
                self.display_status("LogService*.txt parse failed !", "fail")
                QMessageBox.information(self, "Log Analyser", "LogService*.txt parse failed !", QMessageBox.StandardButton.Ok)
                return
            self.mask_events, self.workflow_data = parsed_result

            # Show Mask ID Load/Unload Info
            self.display_load_unload_info()
        else:
            self.mask_events, self.workflow_data = {}, {}
            self.lwt_mask_info.clear()
            self.lwt_wk_flow.clear()

        # 2. Install.txt
        if install_log:
            self.parse_install(folder_path)
        else:
            self.install_info_list.clear()

        # 3. mjnxtdebug.log
        if msc_file:
            self.log_data, time_range, self.valve_log_data = self.parse_msc(folder_path, msc_file)
            if not self.log_data or not self.valve_log_data:
                self.timer.stop()
                self.display_status("mjnxtdebug*.log parse failed !", "fail")
                QMessageBox.information(self, "Log Analyser", "mjnxtdebug*.log parse failed !", QMessageBox.StandardButton.Ok)
                return

            # 獲取時間範圍
            self.min_time = min(time_range)
            self.max_time = max(time_range)
            self.start_date_edit.setMinimumDate(self.min_time.date())
            self.start_date_edit.setMaximumDate(self.max_time.date())
            self.end_date_edit.setMinimumDate(self.min_time.date())
            self.end_date_edit.setMaximumDate(self.max_time.date())
        else:
            self.log_data, time_range, self.valve_log_data = {}, [], {}
            self.figure_ctr.clear()

        # 4. Protocol.txt
        if protocol_file and process_log:
            self.prot_start_date, self.prot_end_date = self.parse_protocol(folder_path, protocol_file, process_log)
        else:
            self.protocol_data = {}
 
        self.timer.stop()
        self.display_status("Process done !", "done")
        QMessageBox.information(self, "Log Analyser", "Process done !", QMessageBox.StandardButton.Ok)
        self.process_button.setEnabled(True)
        self.process_button.setStyleSheet("""
            QPushButton {
                font-weight: bold;
                background-color: white;
                color: black; 
                border: 2px solid #000;
            }
            QPushButton:hover {
                font-weight: bold;
                background-color: lightblue;
                color: black;
            }
        """)

    def parse_logsvr(self, folder_path, logsrvc_file):
        logsrvc_file = sorted(
            logsrvc_file,
            key=lambda f: os.path.getmtime(os.path.join(folder_path, f)),
            reverse=True
        )

        # 儲存數據
        self.mask_events = defaultdict(list)
        self.workflow_data = defaultdict(dict)
        workflow_id = 1

        mask_event_pattern = r"Mask with Id: (\w+\.\d+) (loaded|unloaded)"
        workflow_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into '([^']+)'"
        end_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'Setup.ProtocolDefinition'"
        loaded_start_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'MaskLoading.LoadingMask'"
        unloaded_start_pattern = r"State Machine state changed: Jump from state 'Special.Idle' into 'MaskUnloading.UnloadingMask'"

        loaded_start_time = None
        loaded_end_time = None
        unloaded_start_time = None
        unloaded_end_time = None

        current_workflow = []
        mask_id = None

        for i, file in enumerate(logsrvc_file):
            file_path = os.path.join(folder_path, file)
            with open(file_path, 'r', encoding='iso-8859-1') as f:
                lines = f.readlines()

            for line in reversed(lines):  # 倒序處理
                # 提取 Mask ID 事件
                mask_match = re.search(mask_event_pattern, line)
                if mask_match:
                    mask_id = mask_match.group(1)
                    event_type = mask_match.group(2)
                    timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    #self.mask_events[mask_id].append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, "N/A"))

                    if event_type == "loaded":
                        loaded_end_time = timestamp
                    elif event_type == "unloaded":
                        unloaded_end_time = timestamp

                # 提取 workflow_pattern
                workflow_match = re.search(workflow_pattern, line)
                if workflow_match and mask_id:
                    timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    current_workflow.append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], workflow_match.group(1)))

                    if re.search(loaded_start_pattern, line):
                        loaded_start_time = timestamp
                        if loaded_start_time and loaded_end_time and event_type == "loaded":
                            cost_time = str((loaded_end_time - loaded_start_time).total_seconds())
                        else:
                            cost_time = "N/A"
                        self.mask_events[mask_id].append((loaded_end_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, cost_time))
                    elif re.search(unloaded_start_pattern, line):
                        unloaded_start_time = timestamp
                        if unloaded_start_time and unloaded_end_time and event_type == "unloaded":
                            cost_time = str((unloaded_end_time - unloaded_start_time).total_seconds())
                        else:
                            cost_time = "N/A"
                        self.mask_events[mask_id].append((unloaded_end_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], event_type, cost_time))

                # 偵測到結束條件
                if re.search(end_pattern, line) and mask_id:
                    #timestamp_str = re.search(r"(\d+-\w+-\d+ \d+:\d+:\d+\.\d+)", line).group(1)
                    #timestamp = datetime.strptime(timestamp_str, "%d-%b-%Y %H:%M:%S.%f")
                    #current_workflow.append((timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], line.strip()))

                    # 確保 mask_id 的列表已初始化
                    if mask_id not in self.workflow_data[workflow_id]:
                        self.workflow_data[workflow_id][mask_id] = []
                    self.workflow_data[workflow_id][mask_id].extend(reversed(current_workflow))
                    
                    workflow_id += 1
                    current_workflow = []
                    mask_id = None
                    loaded_start_time = loaded_end_time = None
                    unloaded_start_time = unloaded_end_time = None

            # 更新進度條
            self.progress_bar.setValue(i + 1)
            self.processed_files += 1
            QApplication.processEvents()

        return self.mask_events, self.workflow_data
    
    def display_load_unload_info(self):
        # 清空 Mask ID 列表
        self.lwt_mask_info.clear()

        self.duration_dict = defaultdict(list)

        # 收集所有事件並排序
        all_events = [(mask_id, timestamp, event_type, cost_time)
                    for mask_id, events in self.mask_events.items()
                    for timestamp, event_type, cost_time in events]

        sorted_events = sorted(all_events, key=lambda x: x[1])  # 按時間排序

        loaded_timestamp = None  # 用於計算載入與卸載之間的時間

        for mask_id, timestamp, event_type, cost_time in sorted_events:
            self.lwt_mask_info.addItem(f"{mask_id} | {timestamp} | {event_type} | => cost {cost_time} secs")

            timestamp_dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S.%f")

            if event_type == "loaded":
                loaded_timestamp = timestamp_dt

            if event_type == "unloaded":
                if loaded_timestamp is None:
                    self.lwt_mask_info.addItem(f"duration : N/A")
                else:
                    duration_seconds = (timestamp_dt - loaded_timestamp).total_seconds()
                    hours, remainder = divmod(duration_seconds, 3600)  # 計算小時和剩餘秒數
                    minutes, seconds = divmod(remainder, 60)  # 計算分鐘和秒數
                    # hh:mm:ss 格式
                    duration_str = f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
                    self.lwt_mask_info.addItem(f"duration : {duration_str}")

                    # 將持續時間記錄到 duration_dict 中
                    self.duration_dict[mask_id].append(duration_seconds)

                    loaded_timestamp = None  # 重置載入時間
                self.lwt_mask_info.addItem(f"===========================================")

        if loaded_timestamp is not None:
            self.lwt_mask_info.addItem(f"duration : N/A")
        
    def display_workflow(self, item):
        selected_text = item.text()
        split_data = selected_text.split(" | ")
        if len(split_data) >= 4:
            mask_id, timestamp_str, event_type, cost_time = split_data
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S.%f")

            mask_event_loaded = r"MaskLoading.MaskLoaded"

            # 遍歷 workflow_data，找到與 mask_id 相符的記錄
            self.lwt_wk_flow.clear()
            wf_print_out = False
            for workflow_id, data in self.workflow_data.items():
                if mask_id in data and not wf_print_out:  # 確保 mask_id 存在於該 workflow_id 的數據中
                    for wf_timestamp_str, wf_line in data[mask_id]:  # 遍歷該 mask_id 的工作流
                        wf_timestamp = datetime.strptime(wf_timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
                        if wf_timestamp <= timestamp:  # 範圍條件
                            # 顯示第一組符合條件的資料後退出
                            #self.lwt_wk_flow.addItem(f"{workflow_id} | {wf_timestamp_str} | {wf_line}")
                            self.lwt_wk_flow.addItem(f"{wf_timestamp_str} | {wf_line}")
                            evnt_match = re.search(mask_event_loaded, wf_line)
                            if evnt_match:
                                self.lwt_wk_flow.addItem(f"-----------------------------------------------------------")
                            wf_print_out = True
            self.lwt_wk_flow.addItem(f"=================================")

    def generate_load_unload_time_chart(self):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        if not self.mask_events or not self.workflow_data:
            self.display_status("No parse for LogService.txt !", "fail")
            QMessageBox.information(self, "Log Analyser", "No parse for LogService.txt !", QMessageBox.StandardButton.Ok)
            return

        # 收集所有 loaded 事件
        all_loaded_events = [
            (mask_id, timestamp, event_type, cost_time)
            for mask_id, events in self.mask_events.items()
            for timestamp, event_type, cost_time in events if event_type == "loaded"
        ]
        # 收集所有 unloaded 事件
        all_unloaded_events = [
            (mask_id, timestamp, event_type, cost_time)
            for mask_id, events in self.mask_events.items()
            for timestamp, event_type, cost_time in events if event_type == "unloaded"
        ]

        # 按時間排序
        sorted_loaded_events = sorted(all_loaded_events, key=lambda x: datetime.strptime(x[1], "%Y-%m-%d %H:%M:%S.%f"))
        sorted_unloaded_events = sorted(all_unloaded_events, key=lambda x: datetime.strptime(x[1], "%Y-%m-%d %H:%M:%S.%f"))

        # 準備數據以繪製圖表
        loaded_times = {}
        for mask_id, timestamp, event_type, cost_time in sorted_loaded_events:
            if mask_id not in loaded_times:
                loaded_times[mask_id] = []
            loaded_times[mask_id].append(float(cost_time))

        unloaded_times = {}
        for mask_id, timestamp, event_type, cost_time in sorted_unloaded_events:
            if mask_id not in unloaded_times:
                unloaded_times[mask_id] = []
            unloaded_times[mask_id].append(float(cost_time)) 

        self.plot_load_unload_chart(loaded_times, unloaded_times)

    def plot_load_unload_chart(self, loaded_data, unloaded_data):
        self.figure_mask_analysis.clear()

        # 使用 gridspec 設置子圖的高度比例
        gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1])  # 上子圖佔 2， 下子圖佔 1

        ax1 = self.figure_mask_analysis.add_subplot(gs[0])  # 上半部子圖
        ax2 = self.figure_mask_analysis.add_subplot(gs[1])  # 下半部子圖

        # 繪製載入時間圖
        load_times = {mask_id: [cost_time for cost_time in cost_times if cost_time > 0] for mask_id, cost_times in loaded_data.items()}
        self.plot_chart(load_times, "Load Times", "Mask ID", "Cost Time (minutes)", event_type="loaded", ax=ax1)

        # 繪製卸載時間圖
        unload_times = {mask_id: [cost_time for cost_time in cost_times if cost_time > 0] for mask_id, cost_times in unloaded_data.items()}
        self.plot_chart(unload_times, "Unload Times", "Mask ID", "Cost Time (minutes)", event_type="unloaded", ax=ax2)

    def generate_duration_time_chart(self):
        folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log = self.find_log_files()
        if any(var is None for var in [folder_path, logsrvc_file, install_log, msc_file, protocol_file, process_log]):
            return
        if not self.mask_events or not self.workflow_data:
            self.display_status("No parse for LogService.txt !", "fail")
            QMessageBox.information(self, "Log Analyser", "No parse for LogService.txt !", QMessageBox.StandardButton.Ok)
            return
        
        self.plot_duration_chart(self.duration_dict)
    
    def plot_chart(self, data, title, xlabel, ylabel, event_type, ax):
        # 準備數據以繪製直條圖
        mask_ids = []
        duration_values = []
        mask_id_count = {}  # 用於計數每個 Mask ID 的出現次數

        for mask_id, cost_times in data.items():
            if cost_times:  # 確保有資料才繪製
                for cost_time in cost_times:
                    # 檢查 Mask ID 是否已經存在，並進行計數
                    if mask_id in mask_id_count:
                        mask_id_count[mask_id] += 1
                        new_mask_id = f"{mask_id}_{mask_id_count[mask_id]}"
                    else:
                        mask_id_count[mask_id] = 1
                        new_mask_id = mask_id

                    mask_ids.append(new_mask_id)
                    duration_values.append(cost_time / 60)  # 將秒轉換為分鐘

        if not mask_ids or not duration_values:
            print("No data to plot.")
            return  # 如果沒有數據，則不繪製圖表

        bar_width = 0.4

        # 繪製直條圖
        bars = ax.bar(mask_ids, duration_values, width=bar_width, color='skyblue')

        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)  # 單位為分鐘
        
        ax.grid(axis='y')  # 只顯示 y 軸的網格線

        # 在每根直條上顯示數值
        for bar in bars:
            yval = bar.get_height()  # 獲取直條的高度（即數值）
            ax.text(bar.get_x() + bar.get_width() / 2, yval, f'{yval:.2f}', ha='center', va='bottom')  # 在直條上方顯示數值

        # 在 y 軸為 10 的地方畫一條水平線
        #ax.axhline(y=10, color='red', linestyle='--', label='Threshold Line at 10') 

        # 調整 x 軸標籤的顯示
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 調整 y 軸範圍（如果需要）
        ax.set_ylim(bottom=0)  # 確保 y 軸從 0 開始

        # 設置 y 軸範圍
        ax.set_ylim(0, 20)

        # 調整邊距
        self.figure_mask_analysis.tight_layout()
        self.canvas_mask_analysis.draw()

    def plot_duration_chart(self, data):
        self.figure_mask_analysis.clear()
        ax = self.figure_mask_analysis.add_subplot(111)

        mask_ids = []
        duration_values = []
        mask_id_count = {}  # 用於計數每個 Mask ID 的出現次數

        for mask_id, durations in data.items():
            if durations:  # 確保有資料才繪製
                for duration in durations:
                    # 檢查 Mask ID 是否已經存在，並進行計數
                    if mask_id in mask_id_count:
                        mask_id_count[mask_id] += 1
                        new_mask_id = f"{mask_id}_{mask_id_count[mask_id]}"
                    else:
                        mask_id_count[mask_id] = 1
                        new_mask_id = mask_id

                    mask_ids.append(new_mask_id)
                    duration_values.append(duration)

        # 將持續時間從秒轉換為小時
        duration_values_in_hours = [duration / 3600 for duration in duration_values]
        
        # 繪製直條圖
        bars = ax.bar(mask_ids, duration_values_in_hours, color='skyblue')

        ax.set_title("Duration Time")
        ax.set_xlabel("Mask ID")
        ax.set_ylabel("Duration Time (hours)")
        
        ax.grid(axis='y')  # 只顯示 y 軸的網格線

        # 在每根直條上顯示數值
        for bar in bars:
            yval = bar.get_height()  # 獲取直條的高度（即數值）
            ax.text(bar.get_x() + bar.get_width() / 2, yval, f'{yval:.2f}', ha='center', va='bottom')  # 在直條上方顯示數值

        # 調整 x 軸標籤的顯示
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 調整 y 軸範圍（如果需要）
        ax.set_ylim(bottom=0)  # 確保 y 軸從 0 開始

        # 調整邊距
        self.figure_mask_analysis.tight_layout()

        self.canvas_mask_analysis.draw()

    def parse_install(self, folder_path):
        install_file = os.path.join(folder_path, "Install.txt")
        if os.path.exists(install_file):            
            with open(install_file, 'r', encoding='iso-8859-1') as file:
                lines = file.readlines()
                for line in lines:
                    item = QListWidgetItem(line.strip())
                    self.install_info_list.addItem(item)

                # 將最後一行設置為黃色
                if self.install_info_list.count() > 0:  # 確保有項目存在
                    last_item = self.install_info_list.item(self.install_info_list.count() - 1)
                    last_item.setBackground(QColor("yellow"))  # 設置背景顏色為黃色

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()

    def parse_msc(self, folder_path, msc_file):
        valve_ids = ["V0-3", "V0-4", "V0-5", "V0-7", "V0-10", "V0-13", "V0-16", "V0-31", "V0-32", "V0-33", "V0-34", "V0-35",
                     "V1-3", "V1-4", "V1-9", "V1-10", "V1-11", "V1-13", "V1-21", "V1-22",
                     "V2-3", "V2-4", "V2-9", "V2-11", "V2-13", "V2-21",
                     "V3-3", "V3-4", "V3-9", "V3-11", "V3-13", "V3-21",
                     "V4-3", "V4-4", "V4-5", "V4-9", "V4-11", "V4-13", "V4-21",
                     "V9-3", "V9-4", "V9-5", "V9-7", "V9-8", "V9-11", "V9-12", "V9-13", "V9-14", "V9-15", "V9-16", "V9-17", "V9-21",
                     "V10-3", "V10-5", "V10-7", "V10-8", "V10-11", "V10-12", "V10-13", "V10-14", "V10-15", "V10-16", "V10-17", "V10-21"]

        selected_format = self.selected_format
        if selected_format == "MSC 2.x":
            ctr_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(\d+\) PressEvTh\(\): Sent MULTIJET_EVENT_CODE_CURRENT_PRESSURE_HAS_CHANGED\((\d+), (\d+), press=(-?\d+\.\d+), array=(-?\d+\.\d+)\)")
            valve_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \((\d+)\) .*?: Calling SafeMediaCabinetSetValveDefaultState\((\d+),(\d+),(\d+)\)")
        elif selected_format == "MSC 3.x":
            ctr_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(.*?\) MultiJetImpl::MCPressCurrentValueChangedEvent\((\d+),(\d+)\), .*?pressure = (-?\d+\.\d+) mbar.*")
            mfc_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(\d+\) MCMFCCurrValueChangedEvent\(ch=(\d+), row=(\d+), value=([\d\.-]+), ValueSent=[\d\.-]+\)")
            valve_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3}): \(\d+\) .*?: Calling MediaCabinetSetValveDefaultState\((\d+),(\d+),(\d+)\)")
        else:
            self.display_status("Unsupported log format selected !", "fail")
            QMessageBox.information(self, "Log Analyser", "Unsupported log format selected !", QMessageBox.StandardButton.Ok)
            return None, None

        # 讀取 mjnxtdebugXXXXXXXX.log 文件以獲取初始時間
        initial_time = None

        for file_name in msc_file:
            file_path = os.path.join(folder_path, file_name)
            with open(file_path, 'r') as file:
                for line in file:
                    # 使用正則表達式提取時間字符串
                    time_pattern = re.search(r'(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d{3})', line)
                    if time_pattern:
                        date_time_str = time_pattern.group(1)  # 獲取匹配的時間字符串
                        initial_time = datetime.strptime(date_time_str, "%Y/%m/%d, %H:%M:%S.%f")
                        break  # 只需獲取第一行的時間
            break  # 只需處理第一個日誌文件

        valve_log_data = {valve_id: [(initial_time, 0)] for valve_id in valve_ids}  # 每個閥門的初始狀態為關閉
        
        time_stamps = []
        #gauge_log_data = {gauge_id: [] for gauge_id in gauge_ids}
        gauge_log_data = {gauge_id: [] for gauge_id in self.gauge_types}

        for file_name in msc_file:
            file_path = os.path.join(folder_path, file_name)

            with open(file_path, 'r') as file:
                for line in file:
                    # CTR/MFC value
                    gauge_match = ctr_pattern.search(line) or mfc_pattern.search(line)
                    if gauge_match:
                        date_time_str, main_id, sub_id, press_value = gauge_match.groups()[:4]
                        parsed_time = datetime.strptime(date_time_str, "%Y/%m/%d, %H:%M:%S.%f")
                        time_stamps.append(parsed_time)
                        
                        if (main_id == '9' and sub_id == '5') or (main_id == '10' and sub_id == '5'):
                            gauge_id = f"MFC{main_id}-{sub_id}"
                        else:
                            gauge_id = f"P{main_id}-{sub_id}"
                        if gauge_id in gauge_log_data:
                            gauge_log_data[gauge_id].append((parsed_time, float(press_value)))

                    # valve status
                    valve_match = valve_pattern.search(line)
                    if valve_match:
                        date_time_str, main_id, sub_id, valve_status = valve_match.groups()[:4]
                        parsed_time = datetime.strptime(date_time_str, "%Y/%m/%d, %H:%M:%S.%f")

                        valve_id = f"V{main_id}-{sub_id}"
                        if valve_id in valve_log_data:
                            # 將新的狀態附加到閥門狀態列表中
                            valve_log_data[valve_id].append((parsed_time, int(valve_status)))

                    if self.processed_files % 100 == 0:  # 每 100 行更新一次
                        QApplication.processEvents()

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
        
        # 檢查是否有任何閥門狀態或量測儀數據被記錄
        if not any(valve_log_data.values()) and not any(gauge_log_data.values()):
            return None, None, None
        
        return gauge_log_data, sorted(set(time_stamps)), valve_log_data
    
    def parse_protocol(self, folder_path, protocol_file, process_log):
        self.mask_project_combobox.clear()
        self.protocol_combobox.clear()
        
        start_date = None
        end_date = None
        gauge_ids = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3"]
        

        added_projects = set()  # 用於跟踪已添加的項目

        self.protocol_data = {}
        process_data = {gauge_id: [] for gauge_id in gauge_ids}

        # 設定閥門的起始值
        valve_ids = {
            "P1-1": "V1-21|22",
            "P2-1": "V2-21",
            "P3-1": "V3-21",
            "P4-1": "V4-21",
            "P9-1": "V9-21",
            "P9-2": "V9-21",
            "P9-3": "V9-21",
            "P10-1": "V10-21",
            "P10-2": "V10-21",
            "P10-3": "V10-21",
        }

        # Parse ProcessLog files
        for file in process_log:
            date_match = re.search(r'ProcessLog_(\d{4})-(\d{2})-(\d{2})', file)
            if date_match:
                date_str = f"{date_match.group(1)}/{date_match.group(2)}/{date_match.group(3)}"  # yyyy/mm/dd
                if start_date is None:
                    start_date = date_str
                end_date = date_str
            with open(file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                process_filename = os.path.splitext(os.path.basename(file))[0]  # 去除擴展名

                # 從檔名中提取基準時間
                base_time_str = process_filename.split("_")[1:3]  # 提取檔名中的日期和時間部分
                base_time_str = f"{base_time_str[0]} {base_time_str[1]}"  # 合併為 "YYYY-MM-DD HHMM"
                base_time = datetime.strptime(base_time_str, "%Y-%m-%d %H%M")  # 轉換為 datetime 對象

                gauge_values = {gauge_id: [] for gauge_id in gauge_ids}  # 用於存儲每個量測儀的值
                valve_values = {valve_id: [] for valve_id in valve_ids.values()}  # 用於存儲閥門的值

                for line in lines:
                    line = line.strip()  # 去除行首和行尾的空白字符
                    if line.startswith("\"Elapsed Seconds\""):
                        continue  # 跳過標題行

                    parts = line.split(",")
                    if len(parts) < len(gauge_ids) + 2:  # 確保有足夠的欄位, 跳過不完整的行
                        continue

                    # 獲取 CSV 標題行中的量測儀欄位索引
                    header = [col.strip().strip('"') for col in lines[0].strip().split(",")]  # 去除引號
                    gauge_indices = {gauge_id: header.index(gauge_id) for gauge_id in gauge_ids if gauge_id in header}
                    valve_indices = {valve_id: header.index(valve_id) for valve_id in valve_ids.values() if valve_id in header}

                    # 依序提取量測儀的值
                    for gauge_id, index in gauge_indices.items():
                        value = float(parts[index].strip())  # 根據索引獲取對應的值
                        gauge_values[gauge_id].append(value)  # 將值存入對應的 gauge_id 列表

                    # 依序提取閥門的值
                    for valve_id, index in valve_indices.items():
                        valve_value = int(parts[index].strip())  # 根據索引獲取對應的閥門值
                        valve_values[valve_id].append(valve_value)  # 將值存入對應的閥門列表

                # 計算每個量測儀的平均值並存入 process_data
                for gauge_id in gauge_ids:
                    if gauge_values[gauge_id]:  # 確保有值
                        # 獲取對應閥門的值
                        valve_id = valve_ids[gauge_id]
                        start_calculating = False
                        total_value = 0
                        count = 0

                        for i in range(len(valve_values[valve_id])):
                            if valve_values[valve_id][i] == 1:
                                start_calculating = True  # 開始計算
                            if start_calculating:
                                total_value += gauge_values[gauge_id][i]
                                count += 1

                        if count > 0:
                            average_value = total_value / count  # 計算平均值
                            average_value = round(average_value, 6)
                            process_data[gauge_id].append((base_time.strftime("%Y-%m-%d %H%M"), average_value))  # 存入時間戳和平均值

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
            
        # Parse Protocol.txt file
        for file in protocol_file:
            with open(file, 'r', encoding='utf-8') as f:
                lines = f.readlines()

                protocol_filename = os.path.splitext(os.path.basename(file))[0] # 去除擴展名
                self.protocol_combobox.addItem(protocol_filename)

                process_names = []
                u_value = None
                v_value = None
                current_recipe = None
                mask_project_name = None

                for line in lines:
                    line = line.strip()  # 去除行首和行尾的空白字符

                    if line.startswith("[Recipe_"):
                        current_recipe = line.split("[")[1].split("]")[0]  # 獲取 Recipe_#
                        continue

                    if current_recipe:
                        if "ApplicationModule=" in line:
                            am_path = line.split("ApplicationModule=")[1].strip()
                            process_name = os.path.basename(am_path)  # 獲取檔名，例如 Opaque.am
                            process_names.append(f"{current_recipe} => {process_name}")
                            current_recipe = None

                    if "U=" in line:  # 檢查行中是否包含 U=
                        u_match = re.search(r"U\s*=\s*([\d\.]+)", line)  # 匹配 U 的值
                        if u_match:
                            u_value = u_match.group(1)  # 更新 U 值

                    if "V=" in line:  # 檢查行中是否包含 V=
                        v_match = re.search(r"V\s*=\s*([\d\.]+)", line)  # 匹配 V 的值
                        if v_match:
                            v_value = v_match.group(1)  # 更新 V 值
                
                    # 提取 PreRepairImage 路徑中的項目
                    if "PreRepairImage=" in line:
                        image_path = line.split("PreRepairImage=")[1].strip()
                        project_match = re.search(r'\\([^\\]+)\\[^\\]+$', image_path)  # 匹配最後一個目錄名稱
                        if project_match:
                            mask_project_name = project_match.group(1)  # 提取目錄名稱
                            if mask_project_name not in added_projects:
                                self.mask_project_combobox.addItem(mask_project_name)  # 添加到 mask_project_combobox
                                added_projects.add(mask_project_name)  # 將項目添加到集合中
                
                # 從 protocol_filename 提取時間戳
                protocol_base_time_split = protocol_filename.split("_")[1:3]  # 提取檔名中的日期和時間部分
                protocol_base_time_str = f"{protocol_base_time_split[0]} {protocol_base_time_split[1]}"  # 合併為 "YYYY-MM-DD HHMM"

                # 將資料存入字典
                if mask_project_name:
                    if mask_project_name not in self.protocol_data:
                        self.protocol_data[mask_project_name] = {}
                    self.protocol_data[mask_project_name][protocol_filename] = {
                        "process": process_names,
                        "U": u_value,
                        "V": v_value,
                        "process_data": {}
                    }

                    # 只將與 protocol_filename 時間匹配的 gauge 平均值存入
                    for gauge_id in gauge_ids:
                        if gauge_id in process_data and process_data[gauge_id]:
                            for timestamp, value in process_data[gauge_id]:
                                if timestamp == protocol_base_time_str:
                                    self.protocol_data[mask_project_name][protocol_filename]["process_data"][gauge_id] = [(timestamp, value)]

            self.processed_files += 1
            self.progress_bar.setValue(self.processed_files)
            QApplication.processEvents()
            
        self.mask_project_combobox.currentIndexChanged.connect(self.update_protocol_combobox)
        self.protocol_combobox.currentIndexChanged.connect(self.update_protocol_info)

        return start_date, end_date

    def on_tab_changed(self, index):
        if index == 3:  # MSC Info 是第四個 Tab，索引為 3
            self.update_protocol_info()

    def update_protocol_combobox(self):
        selected_project = self.mask_project_combobox.currentText()  # 獲取選中的 mask_project_name
        self.protocol_combobox.clear()  # 清空 protocol_combobox

        if selected_project in self.protocol_data:
            for protocol_filename in self.protocol_data[selected_project]:
                self.protocol_combobox.addItem(protocol_filename)  # 添加相關的 protocol_filename

    def update_protocol_info(self):
        selected_project = self.mask_project_combobox.currentText()  # 獲取選中的 mask_project_name
        selected_protocol = self.protocol_combobox.currentText()  # 獲取選中的 Protocol 文件名

        if selected_project in self.protocol_data and selected_protocol in self.protocol_data[selected_project]:
            process_names = self.protocol_data[selected_project][selected_protocol]["process"]
            u_value = self.protocol_data[selected_project][selected_protocol]["U"]
            v_value = self.protocol_data[selected_project][selected_protocol]["V"]

            # 更新 Label
            process_display = "\n".join(process_names) if process_names else "N/A"
            self.process_label.setText(f"Processes:\n{process_display}")
            self.u_label.setText(f"\tU: {u_value if u_value is not None else 'N/A'}")
            self.v_label.setText(f"\tV: {v_value if v_value is not None else 'N/A'}")
        else:
            # 如果沒有找到對應的資料，顯示 N/A
            self.process_label.setText("\tProcess: N/A")
            self.u_label.setText("\tU: N/A")
            self.v_label.setText("\tV: N/A")

    def on_format_selected(self):
        if self.msc_v2_radio.isChecked():
            self.selected_format = "MSC 2.x"
        elif self.msc_v3_radio.isChecked():
            self.selected_format = "MSC 3.x"

    def on_gauge_checkbox_changed(self):
        # 獲取當前觸發事件的複選框
        #checkbox = self.sender()
        
        # 獲取複選框的名稱或 ID
        #gauge_id = None
        #for gauge, cb in self.gauge_checkboxes.items():
        #    if cb == checkbox:
        #        gauge_id = gauge
        #        break

        # 檢查複選框的狀態
        #if checkbox.isChecked():
            #print(f"{gauge_id} 被選中")  # 當前複選框被選中
        #    self.last_unchecked_gauge = None  # 如果選中，重置最後一個未選中的 gauge
        #else:
            #print(f"{gauge_id} 被取消選擇")  # 當前複選框被取消選擇
        #    self.last_unchecked_gauge = gauge_id  # 更新最後一個被取消選擇的 gauge

        #self.create_chart(generate_chart=True)

        pass
    
    def create_chart(self, generate_chart=False):
        _, _, _, msc_file, protocol_file, process_log = self.find_log_files()
        
        if all(var is None for var in [msc_file, protocol_file, process_log]):
            self.display_status("File path, mjnxtdebug*.log, and Protocol.txt cannot be empty!", "fail")
            QMessageBox.information(self, "Log Analyser", "Please check File path, mjnxtdebug.log, and Protocol.txt !", QMessageBox.StandardButton.Ok)
            return
        
        if self.selected_chart == "mjnxtdebug" and not self.log_data:
            self.display_status("mjnxtdebug*.log parse failed !", "fail")
            QMessageBox.information(self, "Log Analyser", "mjnxtdebug*.log parse failed !", QMessageBox.StandardButton.Ok)
            return
        elif self.selected_chart == "protocol" and not self.protocol_data:
            self.display_status("Protocol.txt or ProcessLog.csv parse failed !", "fail")
            QMessageBox.information(self, "Log Analyser", "Protocol.txt or ProcessLog.csv parse failed !", QMessageBox.StandardButton.Ok)
            return
            
        # 選擇開始和結束時間
        start_date = self.start_date_edit.text()
        start_time = f"{self.start_hour_spinbox.value():02}:{self.start_minute_spinbox.value():02}:{self.start_second_spinbox.value():02}"
        end_date = self.end_date_edit.text()
        end_time = f"{self.end_hour_spinbox.value():02}:{self.end_minute_spinbox.value():02}:{self.end_second_spinbox.value():02}"
        
        # 轉換為 Python datetime
        start_datetime = datetime.strptime(f"{start_date} {start_time}", "%Y/%m/%d %H:%M:%S")
        end_datetime = datetime.strptime(f"{end_date} {end_time}", "%Y/%m/%d %H:%M:%S")
        
        # 過濾數據
        filtered_data = self.filter_data(start_datetime, end_datetime)

        if self.selected_chart == "mjnxtdebug" and all(not values for values in filtered_data.values()):
            self.display_status("No data in selected range !", "fail")
            QMessageBox.information(self, "Log Analyser", "No CTR or Valve data in selected range !\n Please set the time filter.", QMessageBox.StandardButton.Ok)
            
            for gauge_id in self.gauge_types:
                self.gauge_checkboxes[gauge_id].setEnabled(True)
                self.gauge_checkboxes[gauge_id].setChecked(True)
            
            return
        
        # 獲取選擇的量測儀類型
        selected_gauges = [gauge for gauge, checkbox in self.gauge_checkboxes.items() if checkbox.isChecked()]
        
        # 根據選擇的量測儀過濾數據
        if selected_gauges:
            filtered_data = {gauge: filtered_data.get(gauge, []) for gauge in selected_gauges}
        else:
            self.display_status("No gauge selected !", "fail")
            QMessageBox.information(self, "Log Analyser", "No gauge selected !", QMessageBox.StandardButton.Ok)

            # 恢復最後一個被取消選擇的 gauge
            #if self.last_unchecked_gauge:
            #    self.gauge_checkboxes[self.last_unchecked_gauge].setChecked(True)  # 恢復該複選框為選中狀態
                #print(f"恢復選中: {self.last_unchecked_gauge}")  # 顯示恢復的 gauge

            return
        
        if generate_chart:
            self.plot_ctr_chart(filtered_data)
        else:
            self.save_to_excel(filtered_data)
        
    def filter_data(self, start_datetime, end_datetime):
        filtered_data = {}
        gauges_to_disable = []

        if self.selected_chart == "mjnxtdebug":
            #gauges_to_disable = []
            for gauge_id, data in self.log_data.items():
                #filtered_data[gauge_id] = [(dt, val) for dt, val in data if start_datetime <= dt <= end_datetime]
                filtered_values = [(dt, val) for dt, val in data if start_datetime <= dt <= end_datetime]
                filtered_data[gauge_id] = filtered_values

                if not filtered_values:
                    gauges_to_disable.append(gauge_id)
            
        elif self.selected_chart == "protocol":
            # 獲取當前選擇的 project file
            current_project = self.mask_project_combobox.currentText()

            # 整理所有 protocol file 的 gauge 平均值
            for _, protocol_info in self.protocol_data[current_project].items():
                process_data = protocol_info["process_data"]
                for gauge_id, gauge_values in process_data.items():
                    for timestamp, value in gauge_values:
                        # 將時間戳轉換為 datetime 對象
                        dt = datetime.strptime(timestamp, "%Y-%m-%d %H%M")
  
                        if gauge_id not in filtered_data:
                            filtered_data[gauge_id] = []
                        filtered_data[gauge_id].append((dt, value))

            # 檢查 filtered_data 中的 gauge_id 是否存在，若不存在則添加到 gauges_to_disable
            for dis_gauge in self.gauge_checkboxes.keys():
                if dis_gauge not in filtered_data or not filtered_data[gauge_id]:
                    gauges_to_disable.append(dis_gauge)

        # 禁用不符合條件的複選框並設置為未選中
        for gauge_id in self.gauge_checkboxes.keys():
            if gauge_id in gauges_to_disable:
                self.gauge_checkboxes[gauge_id].setEnabled(False)
                self.gauge_checkboxes[gauge_id].setChecked(False)
            else:
                self.gauge_checkboxes[gauge_id].setEnabled(True)
                #self.gauge_checkboxes[gauge_id].setChecked(True)

        return filtered_data
    
    def save_to_excel(self, filtered_data):
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel files (*.xlsx)")
        # 檢查文件是否存在且是否可寫入
        if os.path.exists(save_path):
            try:
                # 嘗試以寫入模式打開文件
                with open(save_path, 'a'):
                    pass  # 如果成功，則文件未被佔用
            except PermissionError:
                # 如果出現 PermissionError，則顯示警告
                QMessageBox.warning(self, "Warning", f"The file '{save_path}' is currently open or cannot be accessed. Please close it and try again.", QMessageBox.StandardButton.Ok)
                self.display_status("Excel save error !", "fail")
                return

        # 建立新的 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "All Gauge Data"  # 設定工作表名稱

        # 設定表頭
        ws.append(["Date Time"] + [f"{gauge_id} Press Value" for gauge_id in filtered_data.keys()])  # 依序為每個量測儀添加壓力數據的標題

        # 收集所有時間點
        all_times = sorted(set(time for data in filtered_data.values() for time, _ in data))

        # 填充數據
        for row_idx, time in enumerate(all_times, start=2):
            # 將時間格式化為字符串，包含毫秒
            formatted_time = time.strftime("%Y/%m/%d, %H:%M:%S.%f")[:-3]  # 只保留到毫秒
            ws.cell(row=row_idx, column=1, value=formatted_time)  # 寫入時間
            for col_idx, gauge_id in enumerate(filtered_data.keys(), start=2):
                # 嘗試找到該時間點對應的壓力值
                press_value = next((val for dt, val in filtered_data[gauge_id] if dt == time), None)
                if press_value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=press_value)
                else:
                    # 若無對應壓力值，計算前後壓力值的平均值
                    earlier_values = [val for dt, val in filtered_data[gauge_id] if dt < time]
                    later_values = [val for dt, val in filtered_data[gauge_id] if dt > time]

                    # 獲取最近的前後值
                    earlier_value = earlier_values[-1] if earlier_values else None
                    later_value = later_values[0] if later_values else None

                    # 計算平均值並填入
                    if earlier_value is not None and later_value is not None:
                        average_value = (earlier_value + later_value) / 2
                        ws.cell(row=row_idx, column=col_idx, value=average_value)
                    elif earlier_value is not None:  # 若只有前值
                        ws.cell(row=row_idx, column=col_idx, value=earlier_value)
                    elif later_value is not None:  # 若只有後值
                        ws.cell(row=row_idx, column=col_idx, value=later_value)

        # 建立趨勢圖
        chart = LineChart()
        chart.title = "CTR Pressure Trends"
        chart.x_axis.title = "Date Time"
        chart.y_axis.title = "Press Value (mbar)"

        # 設定數據範圍，從第二列開始，以包含所有量測儀的數據
        data_ref = Reference(ws, min_col=2, min_row=1, max_col=1 + len(filtered_data), max_row=len(all_times) + 1)
        chart.add_data(data_ref, titles_from_data=True)  # 包含標題

        # 設定 X 軸標籤（時間範圍）
        time_ref = Reference(ws, min_col=1, min_row=2, max_row=len(all_times) + 1)
        chart.set_categories(time_ref)

        # 設定 X 軸標籤格式
        chart.x_axis.number_format = "yyyy/mm/dd hh:mm:ss.000"
        chart.x_axis.majorTimeUnit = "days"
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickLblPos = "low"

        # 添加趨勢圖到工作表
        ws.add_chart(chart, "N2")  # 設定圖表顯示位置

        # 儲存 Excel 檔案
        wb.save(save_path)

        QMessageBox.information(self, "Log Analyser", "Data is exported to Excel.", QMessageBox.StandardButton.Ok)
        self.display_status("Data is exported to Excel.", "done")

    def plot_ctr_chart(self, filtered_data):
        self.ctr_char_gen = False
        # 清空畫布
        self.figure_ctr.clear()
        ax = self.figure_ctr.add_subplot(111)
        #ax2 = ax.twinx()

        # 一開始將所有 gauge 設成 N/A
        for gauge_id in self.gauge_types:
            getattr(self, f'vline_{gauge_id.replace("-", "_")}_label').setText(f'<b>{gauge_id}:</b> N/A ')

        # 準備數據
        self.all_times = sorted(set(time.replace(tzinfo=None) for data in filtered_data.values() for time, _ in data))  # 將所有時間轉換為 offset-naive
        #print("All Times:", all_times)  # 調試輸出
        # 列印最小值和最大值
        #if self.all_times:  # 確保 all_times 不為空
        #    print("X 軸範圍: 最小值 =", min(self.all_times), ", 最大值 =", max(self.all_times))  # 列印 X 軸範圍
        #else:
        #    print("all_times 為空，無法計算最小值和最大值。")
        
        press_values = {gauge_id: [] for gauge_id in filtered_data.keys()}

        for gauge_id, data in filtered_data.items():
            for dt, val in data:
                press_values[gauge_id].append((dt.replace(tzinfo=None), val))  # 將 dt 轉換為 offset-naive

        # 繪製數據
        for gauge_id, values in press_values.items():
            if values:
                times, vals = zip(*values)
                if self.selected_chart == "protocol":
                    ax.plot(times, vals, label=gauge_id, marker='o')  # 使用 marker 繪製數據點
                    # 在每個數據點上顯示數值和時間
                    for time, val in zip(times, vals):
                        ax.text(time, val, f'{val:.6f}\n{time.strftime("%H:%M:%S.%f")}', fontsize=8, ha='center', va='bottom', color='red')
                elif self.selected_chart == "mjnxtdebug":
                    ax.plot(times, vals, label=gauge_id)
                    #if gauge_id in ["MFC9-5", "MFC10-5"]:
                    #    ax2.plot(times, vals, label=gauge_id, linestyle='--')  # 使用虛線繪製 SCCM 數據
    
        ax.set_title("MultiJet Trend Chart")
        ax.set_xlabel("Date Time")
        if (self.gauge_checkboxes["MFC9-5"].isChecked() and not self.gauge_checkboxes["MFC10-5"].isChecked() and
            all(not self.gauge_checkboxes[gauge].isChecked() for gauge in self.gauge_checkboxes if gauge not in ["MFC9-5", "MFC10-5"])):
            ax.set_ylabel("Flow Rate (SCCM)")  # 只有當 MFC9-5 被選中且其他都未選中時設置為 SCCM
        elif (self.gauge_checkboxes["MFC10-5"].isChecked() and not self.gauge_checkboxes["MFC9-5"].isChecked() and
            all(not self.gauge_checkboxes[gauge].isChecked() for gauge in self.gauge_checkboxes if gauge not in ["MFC9-5", "MFC10-5"])):
            ax.set_ylabel("Flow Rate (SCCM)")  # 只有當 MFC10-5 被選中且其他都未選中時設置為 SCCM
        elif (self.gauge_checkboxes["MFC10-5"].isChecked() and self.gauge_checkboxes["MFC9-5"].isChecked() and
            all(not self.gauge_checkboxes[gauge].isChecked() for gauge in self.gauge_checkboxes if gauge not in ["MFC9-5", "MFC10-5"])):
            ax.set_ylabel("Flow Rate (SCCM)")
        else:
            ax.set_ylabel("Press Value (mbar)")  # 否則設置為 mbar
        #ax2.set_ylabel("Flow Rate (SCCM)")
        
        # 調整圖例的位置
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1), framealpha=0.5)
        #ax2.legend(loc='lower right', bbox_to_anchor=(1, 1), framealpha=0.5)
        ax.grid()

        # 設置 X 軸格式
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d %H:%M:%S"))  # 只顯示時間
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        # 添加可移動的縱線
        if self.all_times:
            self.vertical_line = ax.axvline(x=self.all_times[0], color='r', linestyle='--')  # 初始位置

            def on_mouse_move(event):
                try:
                    if event.inaxes == ax:  # 確保鼠標在正確的坐標軸上
                        if event.xdata is not None:  # 確保 xdata 是有效的
                            # 將 event.xdata 轉換為 datetime
                            date_time = mdates.num2date(event.xdata).replace(tzinfo=None)

                            # 格式化 datetime 為字符串
                            formatted_time = date_time.strftime("%Y-%m-%d %H:%M:%S.%f")

                            new_x_values = [event.xdata]  # 將 event.xdata 放入列表中
                            self.vertical_line.set_xdata(new_x_values)  # 更新縱線位置

                            # show the datatime
                            self.vline_time_label.setText(f'<b>Time:</b> {formatted_time}')
                            self.figure_ctr.canvas.draw()

		                    # 獲取當前時間點的 gauge 值
                            for gauge_id in self.gauge_checkboxes.keys():
                                if gauge_id in filtered_data:
                                    # 獲取該 gauge 的數據
                                    gauge_data = filtered_data[gauge_id]
                                    # 找到最近的兩個點
                                    lower_point = None
                                    upper_point = None

                                    for dt, val in gauge_data:
                                        dt = dt.replace(tzinfo=None)
                                        if dt < date_time:
                                            lower_point = (dt, val)
                                        elif dt > date_time and upper_point is None:
                                            upper_point = (dt, val)
                                            break

                                    # 如果找到了兩個點，進行插值
                                    if lower_point and upper_point:
                                        dt1, val1 = lower_point
                                        dt2, val2 = upper_point

                                        # 線性插值計算
                                        slope = (val2 - val1) / (dt2 - dt1).total_seconds()  # 計算斜率
                                        interpolated_value = val1 + slope * (date_time - dt1).total_seconds()  # 計算插值

                                        # 更新相應的 QLabel
                                        getattr(self, f'vline_{gauge_id.replace("-", "_")}_label').setText(f'<b>{gauge_id}:</b> {interpolated_value:.2f} ')
                                    else:
                                        # 若沒有找到足夠的點，則顯示 N/A
                                        getattr(self, f'vline_{gauge_id.replace("-", "_")}_label').setText(f'<b>{gauge_id}:</b> N/A ')

                            # 更新閥門狀態顯示
                            if self.valve_status_dialog is not None:
                                current_valve_states = self.get_valve_states_at_time(date_time)
                                self.valve_status_dialog.valve_states = current_valve_states  # 更新 ValveStatusDialog 的閥門狀態
                                self.valve_status_dialog.update_valve_display()  # 調用 ValveStatusDialog 的更新方法

                except Exception as e:
                    print(f"An error occurred: {e}")  # 輸出錯誤信息

            self.figure_ctr.canvas.mpl_connect('motion_notify_event', on_mouse_move)

        # 自動調整 X 軸範圍
        if self.all_times:
            ax.set_xlim([min(self.all_times), max(self.all_times)])  # 設置 X 軸範圍

        # 調整邊距
        plt.subplots_adjust(bottom=0.2)  # 調整底部邊距

        # 顯示圖表
        self.figure_ctr.tight_layout()  # 自動調整子圖參數
        self.canvas_ctr.draw()

        self.ctr_char_gen = True
        self.display_status("Trend chart is generated", "done")

    def get_valve_states_at_time(self, date_time):
        # 確保 date_time 是 offset-naive
        if date_time.tzinfo is not None:
            date_time = date_time.replace(tzinfo=None)

        current_states = {}
        
        for valve_id, states in self.valve_log_data.items():
            # 遍歷狀態列表，找到最新的狀態
            for timestamp, state in reversed(states):
                # 確保 timestamp 是 offset-naive
                if timestamp.tzinfo is not None:
                    timestamp = timestamp.replace(tzinfo=None)
                    
                if timestamp <= date_time:
                    current_states[valve_id] = state
                    break  # 找到最新狀態後退出循環

        return current_states

    def zoom_chart(self, zoom_in=True):
        if self.ctr_char_gen == False:
            self.display_status("Trend Chart is unavailabe !", "fail")
            QMessageBox.information(self, "Log Analyser", "Trend Chart is unavailabe !", QMessageBox.StandardButton.Ok)
            return
        
        ax = self.figure_ctr.get_axes()[0]  # 獲取第一個坐標軸
        xlim = ax.get_xlim()  # 獲取當前 x 軸範圍

        # 獲取選擇的倍率
        #selected_multiplier = self.zoom_combo.currentText().replace('%', '')
        #multiplier = int(selected_multiplier) / 100  # 轉換為小數
        multiplier = 0.1
        
        if zoom_in:
            ax.set_xlim([xlim[0] + (xlim[1] - xlim[0]) * multiplier, xlim[1] - (xlim[1] - xlim[0]) * multiplier])  # 放大 10%
        else:
            ax.set_xlim([xlim[0] - (xlim[1] - xlim[0]) * multiplier, xlim[1] + (xlim[1] - xlim[0]) * multiplier])  # 縮小 10%
        
        self.canvas_ctr.draw()  # 重新繪製圖形

    def shift_chart(self, left_shift=True):
        if self.ctr_char_gen == False:
            self.display_status("Trend Chart is unavailabe !", "fail")
            QMessageBox.information(self, "Log Analyser", "Trend Chart is unavailabe !", QMessageBox.StandardButton.Ok)
            return
        
        ax = self.figure_ctr.get_axes()[0]  # 獲取第一個坐標軸
        xlim = ax.get_xlim()  # 獲取當前 x 軸範圍

        # 獲取選擇的倍率
        selected_multiplier = self.shift_combo.currentText().replace('%', '')
        multiplier = int(selected_multiplier) / 100  # 轉換為小數
        
        if left_shift:
            ax.set_xlim([xlim[0] - (xlim[1] - xlim[0]) * multiplier, xlim[1] - (xlim[1] - xlim[0]) * multiplier])  # 向左平移 10%
        else:
            ax.set_xlim([xlim[0] + (xlim[1] - xlim[0]) * multiplier, xlim[1] + (xlim[1] - xlim[0]) * multiplier])  # 向右平移 10%
        
        self.canvas_ctr.draw()  # 重新繪製圖形

    def multijet_show(self):
        # 獲取當前腳本的路徑
        current_directory = os.path.dirname(os.path.abspath(__file__))
        image_path = os.path.join(current_directory, "multijet_pipe_sample.png")  # 替換為您的圖片檔名

        dialog = MultijetImgDialog(image_path, self)
        dialog.show()

    def valve_state_show(self):
        # 獲取當前腳本的路徑
        current_directory = os.path.dirname(os.path.abspath(__file__))
        multijet_image_path = os.path.join(current_directory, "multijet_pipe.png")
        if not os.path.exists(multijet_image_path):
            #self.display_status("multijet_pipe.png not found !", "fail")
            print("multijet_pipe.png not found")
            return

        valve_open_image_path = os.path.join(current_directory, "valve_open.png")
        if not os.path.exists(valve_open_image_path):
            #self.display_status("valve_open.png not found !", "fail")
            print("valve_open.png not found")
            return

        valve_close_image_path = os.path.join(current_directory, "valve_close.png")
        if not os.path.exists(valve_close_image_path):
            #self.display_status("valve_close.png not found !", "fail")
            print("valve_close.png not found")
            return

        valve_positions = {
            "V0-3": (174, 163),
            "V0-4": (174, 581),
            "V0-5": (118, 118),
            "V0-7": (118, 163),
            "V0-10": (174, 22),
            "V0-13": (118, 227),
            "V0-16": (174, 286),
            "V0-31": (80, 335),
            "V0-32": (80, 376),
            "V0-33": (80, 416),
            "V0-34": (80, 460),
            "V0-35": (80, 499),
            # CH1
            "V1-3": (896, 22),
            "V1-4": (520, 57),
            "V1-9": (588, 57),
            "V1-10": (804, 56),
            "V1-11": (247, 22),
            "V1-13": (304, 56),
            "V1-21": (956, 22),
            "V1-22": (956, 57),
            #CH2
            "V2-3": (896, 91),
            "V2-4": (520, 125),
            "V2-9": (588, 126),
            "V2-11": (247, 91),
            "V2-13": (304, 125),
            "V2-21": (956, 92),
            #CH3
            "V3-3": (896, 152),
            "V3-4": (520, 186),
            "V3-9": (588, 186),
            "V3-11": (247, 152),
            "V3-13": (304, 186),
            "V3-21": (956, 152),
            #CH4
            "V4-3": (896, 216),
            "V4-4": (478, 245),
            "V4-5": (386, 278),
            "V4-9": (588, 244),
            "V4-11": (247, 216),
            "V4-13": (304, 277),
            "V4-21": (956, 216),
            #CH9
            "V9-3": (896, 343),
            "V9-4": (804, 344),
            "V9-5": (459, 370),
            "V9-7": (626, 387),
            "V9-8": (705, 387),
            "V9-11": (247, 344),
            "V9-12": (784, 428),
            "V9-13": (304, 427),
            "V9-14": (386, 427),
            "V9-15": (386, 387),
            "V9-16": (386, 344),
            "V9-17": (495, 386),
            "V9-21": (956, 344),
            #CH10
            "V10-3": (896, 470),
            "V10-5": (458, 498),
            "V10-7": (626, 511),
            "V10-8": (705, 512),
            "V10-11": (247, 470),
            "V10-12": (788, 552),
            "V10-13": (304, 552),
            "V10-14": (386, 553),
            "V10-15": (386, 511),
            "V10-16": (386, 471),
            "V10-17": (495, 512),
            "V10-21": (956, 470),
        }

        # 創建 ValveStatusDialog 實例並儲存
        self.valve_status_dialog = ValveStatusDialog(multijet_image_path, valve_open_image_path, valve_close_image_path, valve_positions, self)
        self.valve_status_dialog.show()

if __name__ == "__main__":
    app = QApplication([])
    window = LogAnalyser()
    window.show()
    app.exec()
